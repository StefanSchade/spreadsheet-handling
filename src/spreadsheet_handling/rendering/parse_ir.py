"""
parse_ir  –  Reconstruct a WorkbookIR from an XLSX file produced by our renderer.

Scenario A (P1):
    Flat table(s) without named ranges.  Each table has a start anchor
    (default A1) and grows until the first fully-empty row/column.

Discovery priority for table structure:
    1. Embedded _meta sheet  (hidden kv pairs written by MetaPass)
    2. Explicit anchors passed by the caller
    3. Heuristic fallback (first non-empty header row + empty-row sentinel)
"""
from __future__ import annotations

import ast
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

from .ir import WorkbookIR, SheetIR, TableBlock, DataValidationSpec


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_ir(
    path: str | Path,
    *,
    anchors: Dict[str, List[Tuple[int, int]]] | None = None,
    stop_on_empty_row: bool = True,
    stop_on_empty_col: bool = False,
) -> WorkbookIR:
    """
    Read an XLSX file and reconstruct a :class:`WorkbookIR`.

    Parameters
    ----------
    path
        Path to the .xlsx file.
    anchors
        Optional dict mapping sheet names to list of (row, col) table
        start positions (1-based).  Overrides embedded meta for those sheets.
    stop_on_empty_row
        Scenario-A: stop table when a fully-empty row is encountered.
    stop_on_empty_col
        Scenario-A: stop table when a fully-empty column is encountered.

    Returns
    -------
    WorkbookIR
        Reconstructed workbook intermediate representation.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    try:
        ir = WorkbookIR()

        # ---- 1. Read embedded meta ----------------------------------------
        embedded_meta = _read_meta_sheet(wb)

        # ---- 2. Process visible sheets ------------------------------------
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            if ws.sheet_state == "hidden":
                # Reconstruct hidden sheet (e.g. _meta) into hidden_sheets
                sh = _parse_hidden_sheet(ws)
                ir.hidden_sheets[ws_name] = sh
                continue

            sheet_anchors = (anchors or {}).get(ws_name)
            sheet_meta_hints = (embedded_meta.get("sheets") or {}).get(ws_name, {})

            sh = _parse_visible_sheet(
                ws,
                sheet_name=ws_name,
                anchors=sheet_anchors,
                meta_hints=sheet_meta_hints,
                stop_on_empty_row=stop_on_empty_row,
                stop_on_empty_col=stop_on_empty_col,
            )
            ir.sheets[ws_name] = sh

        return ir
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Embedded meta reader
# ---------------------------------------------------------------------------

def _read_meta_sheet(wb: openpyxl.Workbook) -> Dict[str, Any]:
    """Read the hidden _meta sheet if present.  Returns a dict."""
    if "_meta" not in wb.sheetnames:
        return {}

    ws = wb["_meta"]
    kv: Dict[str, str] = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        key = row[0]
        val = row[1] if len(row) > 1 else ""
        if key is not None:
            kv[str(key)] = str(val) if val is not None else ""

    # Try to parse workbook_meta_blob as JSON or Python literal
    blob_str = kv.get("workbook_meta_blob", "")
    if blob_str:
        try:
            return json.loads(blob_str)
        except (json.JSONDecodeError, TypeError):
            pass
        try:
            result = ast.literal_eval(blob_str)
            if isinstance(result, dict):
                return result
        except (ValueError, SyntaxError):
            pass

    return kv


# ---------------------------------------------------------------------------
# Hidden sheet parser
# ---------------------------------------------------------------------------

def _parse_hidden_sheet(ws: Worksheet) -> SheetIR:
    """Parse a hidden sheet (e.g. _meta) into SheetIR with kv meta."""
    sh = SheetIR(name=ws.title)
    sh.meta["_hidden"] = True
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        key = row[0]
        val = row[1] if len(row) > 1 else ""
        if key is not None:
            sh.meta[str(key)] = str(val) if val is not None else ""
    return sh


# ---------------------------------------------------------------------------
# Visible sheet / Scenario A parser
# ---------------------------------------------------------------------------

def _parse_visible_sheet(
    ws: Worksheet,
    *,
    sheet_name: str,
    anchors: List[Tuple[int, int]] | None,
    meta_hints: Dict[str, Any],
    stop_on_empty_row: bool,
    stop_on_empty_col: bool,
) -> SheetIR:
    """Parse a visible sheet into SheetIR with one or more TableBlocks."""
    sh = SheetIR(name=sheet_name)

    # Restore per-sheet options from meta hints
    options = {}
    for key in ("freeze_header", "auto_filter", "header_fill_rgb", "helper_prefix"):
        if key in meta_hints:
            options[key] = meta_hints[key]
    if options:
        sh.meta["options"] = options

    # Determine table anchors
    if anchors:
        table_starts = anchors
    else:
        # Default: single table at A1
        table_starts = [(1, 1)]

    for idx, (top, left) in enumerate(table_starts):
        tbl = _parse_table_block(
            ws,
            frame_name=sheet_name,
            top=top,
            left=left,
            meta_hints=meta_hints,
            stop_on_empty_row=stop_on_empty_row,
            stop_on_empty_col=stop_on_empty_col,
        )
        sh.tables.append(tbl)

    # Extract merge regions relevant to the table headers
    header_merges = _extract_header_merges(ws, sh.tables)
    if header_merges:
        sh.meta["__header_merges"] = header_merges

    # Extract header grid from actual cell values
    for tbl in sh.tables:
        grid = _extract_header_grid(ws, tbl)
        if grid and tbl.header_rows > 1:
            sh.meta["__header_grid"] = grid

    # Extract validations
    validations = _extract_validations(ws)
    sh.validations = validations

    # Detect freeze panes
    if ws.freeze_panes:
        _extract_freeze(ws, sh)

    # Detect auto filter
    if ws.auto_filter and ws.auto_filter.ref:
        sh.meta["__autofilter_ref"] = ws.auto_filter.ref

    return sh


# ---------------------------------------------------------------------------
# Table block parser
# ---------------------------------------------------------------------------

def _parse_table_block(
    ws: Worksheet,
    *,
    frame_name: str,
    top: int,
    left: int,
    meta_hints: Dict[str, Any],
    stop_on_empty_row: bool,
    stop_on_empty_col: bool,
) -> TableBlock:
    """Discover a single table starting at (top, left)."""

    # Detect header_rows: count merge depth at (top, left) or use meta hint
    header_rows = _detect_header_rows(ws, top, left)

    # Find right boundary (n_cols): scan first header row for non-empty cells
    n_cols = _find_col_extent(ws, top, left, stop_on_empty_col)

    # Find bottom boundary (n_data_rows): scan from first data row down
    data_start_row = top + header_rows
    n_data_rows = _find_row_extent(ws, data_start_row, left, n_cols, stop_on_empty_row)

    n_rows = header_rows + n_data_rows

    # Extract headers from last header row (leaf level)
    headers: List[str] = []
    leaf_row = top + header_rows - 1
    for c in range(left, left + n_cols):
        cell = ws.cell(row=leaf_row, column=c)
        val = _cell_value(ws, leaf_row, c)
        headers.append(str(val) if val else "")

    # If MultiIndex (header_rows > 1), build flattened header strings
    if header_rows > 1:
        flat_headers = []
        for c in range(left, left + n_cols):
            parts = []
            for r in range(top, top + header_rows):
                val = _cell_value(ws, r, c)
                parts.append(str(val) if val else "")
            flat_headers.append(" / ".join(p for p in parts if p))
        headers = flat_headers

    header_map = {h: idx + 1 for idx, h in enumerate(headers)}

    # Extract data cell values
    data: List[List] = []
    for r in range(data_start_row, data_start_row + n_data_rows):
        row_vals: List = []
        for c in range(left, left + n_cols):
            val = ws.cell(row=r, column=c).value
            row_vals.append(str(val) if val is not None else "")
        data.append(row_vals)

    return TableBlock(
        frame_name=frame_name,
        top=top,
        left=left,
        header_rows=header_rows,
        header_cols=1,
        n_rows=n_rows,
        n_cols=n_cols,
        headers=headers,
        header_map=header_map,
        data=data,
    )


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Get cell value, resolving merged cells to their top-left value."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # Find the master cell for this merged region
        for mr in ws.merged_cells.ranges:
            if (row, col) in [(r, c) for r in range(mr.min_row, mr.max_row + 1)
                              for c in range(mr.min_col, mr.max_col + 1)]:
                return ws.cell(row=mr.min_row, column=mr.min_col).value
        return None
    return cell.value


def _detect_header_rows(ws: Worksheet, top: int, left: int) -> int:
    """
    Detect number of header rows by examining merge regions.

    Rules:
    - Vertical merges (spanning rows from anchor) → header depth = max span.
    - Horizontal merges at anchor row → at least 2 header rows (group + leaf).
    - No merges → 1 header row.
    """
    max_header_row = top
    has_horizontal_merge_at_top = False

    for mr in ws.merged_cells.ranges:
        if mr.min_row < top or mr.min_col < left:
            continue

        # Vertical merge → direct evidence of multi-row header
        if mr.max_row > mr.min_row and mr.min_row >= top:
            if mr.max_row > max_header_row:
                max_header_row = mr.max_row

        # Horizontal merge at anchor row → group header, leaf row below
        if mr.min_row == top and mr.max_col > mr.min_col and mr.min_row == mr.max_row:
            has_horizontal_merge_at_top = True

    if max_header_row > top:
        return min(max_header_row - top + 1, 10)

    if has_horizontal_merge_at_top:
        return 2

    return 1


def _find_col_extent(ws: Worksheet, top: int, left: int, stop_on_empty: bool) -> int:
    """Find number of columns by scanning the first header row."""
    n_cols = 0
    for c in range(left, left + 16384):  # max Excel columns
        val = _cell_value(ws, top, c)
        if val is None or (isinstance(val, str) and val.strip() == ""):
            if stop_on_empty:
                break
            # Check if there's content further right (limited lookahead)
            has_more = False
            for lookahead in range(1, 4):
                v2 = _cell_value(ws, top, c + lookahead)
                if v2 is not None and str(v2).strip():
                    has_more = True
                    break
            if not has_more:
                break
        n_cols = c - left + 1
    return max(n_cols, 1)


def _find_row_extent(
    ws: Worksheet,
    data_start_row: int,
    left: int,
    n_cols: int,
    stop_on_empty: bool,
) -> int:
    """Find number of data rows by scanning down from data_start_row."""
    n_data_rows = 0
    for r in range(data_start_row, data_start_row + 1048576):  # max Excel rows
        # Check if entire row is empty across known columns
        row_empty = True
        for c in range(left, left + n_cols):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                row_empty = False
                break
        if row_empty:
            if stop_on_empty:
                break
            # Lookahead: if next 2 rows are also empty, consider it done
            lookahead_empty = True
            for lr in range(1, 3):
                for c in range(left, left + n_cols):
                    val = ws.cell(row=r + lr, column=c).value
                    if val is not None and str(val).strip() != "":
                        lookahead_empty = False
                        break
                if not lookahead_empty:
                    break
            if lookahead_empty:
                break
        n_data_rows = r - data_start_row + 1
    return n_data_rows


# ---------------------------------------------------------------------------
# Header / merge extraction
# ---------------------------------------------------------------------------

def _extract_header_merges(
    ws: Worksheet, tables: List[TableBlock]
) -> List[Tuple[int, int, int, int]]:
    """Extract merge regions within the header area of tables (relative coords)."""
    merges: List[Tuple[int, int, int, int]] = []
    for tbl in tables:
        header_bottom = tbl.top + tbl.header_rows - 1
        for mr in ws.merged_cells.ranges:
            if (
                mr.min_row >= tbl.top
                and mr.max_row <= header_bottom
                and mr.min_col >= tbl.left
                and mr.max_col <= tbl.left + tbl.n_cols - 1
            ):
                # Convert to table-relative 1-based coords
                rel = (
                    mr.min_row - tbl.top + 1,
                    mr.min_col - tbl.left + 1,
                    mr.max_row - tbl.top + 1,
                    mr.max_col - tbl.left + 1,
                )
                merges.append(rel)
    return merges


def _extract_header_grid(ws: Worksheet, tbl: TableBlock) -> List[List[str]]:
    """Extract the header grid as a 2D list of strings."""
    grid: List[List[str]] = []
    for r in range(tbl.top, tbl.top + tbl.header_rows):
        row_vals: List[str] = []
        for c in range(tbl.left, tbl.left + tbl.n_cols):
            val = _cell_value(ws, r, c)
            row_vals.append(str(val) if val else "")
        grid.append(row_vals)
    return grid


# ---------------------------------------------------------------------------
# Validation extraction
# ---------------------------------------------------------------------------

def _extract_validations(ws: Worksheet) -> List[DataValidationSpec]:
    """Extract data validations from the worksheet."""
    specs: List[DataValidationSpec] = []
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        formula = str(dv.formula1) if dv.formula1 else ""
        allow_empty = bool(dv.allow_blank) if dv.allow_blank is not None else True
        # Parse cell ranges
        for cell_range in dv.sqref.ranges:
            area = (
                cell_range.min_row,
                cell_range.min_col,
                cell_range.max_row,
                cell_range.max_col,
            )
            specs.append(DataValidationSpec(
                kind="list",
                area=area,
                formula=formula,
                allow_empty=allow_empty,
            ))
    return specs


# ---------------------------------------------------------------------------
# Freeze panes extraction
# ---------------------------------------------------------------------------

def _extract_freeze(ws: Worksheet, sh: SheetIR) -> None:
    """Parse freeze_panes string (e.g. 'A3') into __freeze meta."""
    fp = ws.freeze_panes
    if not fp:
        return
    from openpyxl.utils import column_index_from_string
    # freeze_panes is a cell ref like "A3"
    col_str = ""
    row_str = ""
    for ch in str(fp):
        if ch.isalpha():
            col_str += ch
        else:
            row_str += ch
    if col_str and row_str:
        sh.meta["__freeze"] = {
            "row": int(row_str),
            "col": column_index_from_string(col_str),
        }


# ---------------------------------------------------------------------------
# IR → Frames conversion
# ---------------------------------------------------------------------------

def workbookir_to_frames(ir: WorkbookIR) -> Dict[str, Any]:
    """
    Convert a :class:`WorkbookIR` (from :func:`parse_ir`) into a frames dict
    suitable for the pipeline (``Dict[str, pd.DataFrame | dict]``).

    Each visible sheet's first table becomes a DataFrame.
    Hidden ``_meta`` sheet is returned as ``frames["_meta"]`` (plain dict).
    """
    import pandas as pd

    frames: Dict[str, Any] = {}

    for name, sh in ir.sheets.items():
        if not sh.tables:
            frames[name] = pd.DataFrame()
            continue
        tbl = sh.tables[0]
        data = tbl.data if tbl.data is not None else []

        # Reconstruct columns
        if tbl.header_rows > 1 and tbl.headers and " / " in tbl.headers[0]:
            tuples = [tuple(h.split(" / ")) for h in tbl.headers]
            columns = pd.MultiIndex.from_tuples(tuples)
        else:
            columns = tbl.headers or []

        df = pd.DataFrame(data, columns=columns)
        df = df.where(pd.notnull(df), "")
        frames[name] = df

    # Reconstruct meta from hidden _meta sheet
    meta_sh = ir.hidden_sheets.get("_meta")
    if meta_sh:
        blob = meta_sh.meta.get("workbook_meta_blob", "")
        if blob:
            try:
                meta_dict = json.loads(blob)
                if isinstance(meta_dict, dict):
                    frames["_meta"] = meta_dict
            except (json.JSONDecodeError, TypeError):
                pass
            if "_meta" not in frames:
                try:
                    meta_dict = ast.literal_eval(blob)
                    if isinstance(meta_dict, dict):
                        frames["_meta"] = meta_dict
                except (ValueError, SyntaxError):
                    pass
        if "_meta" not in frames:
            kv = {k: v for k, v in meta_sh.meta.items() if k != "_hidden"}
            if kv:
                frames["_meta"] = kv

    return frames
