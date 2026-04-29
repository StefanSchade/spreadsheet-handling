from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill

from spreadsheet_handling.rendering.plan import (
    RenderPlan,
    DefineSheet,
    SetHeader,
    MergeCells,
    ApplyHeaderStyle,
    ApplyColumnStyle,
    SetAutoFilter,
    SetFreeze,
    AddValidation,
    WriteDataBlock,
    WriteMeta,
    DefineNamedRange,
)
from spreadsheet_handling.rendering.formulas import FormulaSpec, ListLiteralFormulaSpec
from spreadsheet_handling.rendering.formulas import LookupFormulaSpec


# --------------------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------------------

def _area_to_ref(r1: int, c1: int, r2: int, c2: int) -> str:
    c1l = get_column_letter(c1)
    c2l = get_column_letter(c2)
    if r1 == r2 and c1 == c2:
        return f"{c1l}{r1}"
    return f"{c1l}{r1}:{c2l}{r2}"


def _get_ws(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _write(ws: Worksheet, row: int, col: int, value: Any) -> None:
    ws.cell(row=row, column=col, value=value)


def _xlsx_validation_formula(formula: FormulaSpec) -> str:
    if isinstance(formula, ListLiteralFormulaSpec):
        output = io.StringIO()
        writer = csv.writer(output, delimiter=",", quotechar='"', lineterminator="")
        writer.writerow(formula.values)
        return f'"{output.getvalue()}"'
    raise TypeError(f"Unsupported formula spec: {type(formula).__name__}")


def _collect_formula_context(
    plan: RenderPlan,
) -> tuple[dict[str, dict[str, int]], dict[str, tuple[int, int]]]:
    sheet_headers: dict[str, dict[str, int]] = {}
    sheet_data_bounds: dict[str, tuple[int, int]] = {}
    for op in plan.ops:
        if isinstance(op, SetHeader):
            sheet_headers.setdefault(op.sheet, {}).setdefault(op.text, op.col)
        elif isinstance(op, WriteDataBlock) and op.data:
            sheet_data_bounds.setdefault(op.sheet, (op.r1, op.r1 + len(op.data) - 1))
    return sheet_headers, sheet_data_bounds


def _quote_xlsx_sheet_name(sheet: str) -> str:
    return "'" + str(sheet).replace("'", "''") + "'"


def _xlsx_string_literal(value: object) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _header_col(
    sheet_headers: dict[str, dict[str, int]],
    *,
    sheet: str,
    column: str,
) -> int:
    try:
        return sheet_headers[sheet][column]
    except KeyError as exc:
        raise KeyError(f"Formula column {column!r} not found on sheet {sheet!r}") from exc


def _data_bounds(
    sheet_data_bounds: dict[str, tuple[int, int]],
    *,
    sheet: str,
) -> tuple[int, int]:
    try:
        return sheet_data_bounds[sheet]
    except KeyError as exc:
        raise KeyError(f"Formula lookup sheet {sheet!r} has no data block") from exc


def _xlsx_lookup_formula(
    formula: LookupFormulaSpec,
    *,
    current_sheet: str,
    row: int,
    sheet_headers: dict[str, dict[str, int]],
    sheet_data_bounds: dict[str, tuple[int, int]],
) -> str:
    source_col = _header_col(
        sheet_headers,
        sheet=current_sheet,
        column=formula.source_key_column,
    )
    lookup_key_col = _header_col(
        sheet_headers,
        sheet=formula.lookup_sheet,
        column=formula.lookup_key_column,
    )
    lookup_value_col = _header_col(
        sheet_headers,
        sheet=formula.lookup_sheet,
        column=formula.lookup_value_column,
    )
    lookup_start, lookup_end = _data_bounds(sheet_data_bounds, sheet=formula.lookup_sheet)

    source_ref = f"${get_column_letter(source_col)}{row}"
    lookup_sheet = _quote_xlsx_sheet_name(formula.lookup_sheet)
    key_range = (
        f"{lookup_sheet}!${get_column_letter(lookup_key_col)}${lookup_start}:"
        f"${get_column_letter(lookup_key_col)}${lookup_end}"
    )
    value_range = (
        f"{lookup_sheet}!${get_column_letter(lookup_value_col)}${lookup_start}:"
        f"${get_column_letter(lookup_value_col)}${lookup_end}"
    )
    missing = _xlsx_string_literal(formula.missing)
    return f"=XLOOKUP({source_ref},{key_range},{value_range},{missing})"


def _xlsx_cell_value(
    value: Any,
    *,
    current_sheet: str,
    row: int,
    sheet_headers: dict[str, dict[str, int]],
    sheet_data_bounds: dict[str, tuple[int, int]],
) -> Any:
    if isinstance(value, LookupFormulaSpec):
        return _xlsx_lookup_formula(
            value,
            current_sheet=current_sheet,
            row=row,
            sheet_headers=sheet_headers,
            sheet_data_bounds=sheet_data_bounds,
        )
    return value


# --------------------------------------------------------------------------------------
# Typed renderer — dispatches on isinstance checks against plan.py dataclasses
# --------------------------------------------------------------------------------------

def _render_from_plan(plan: RenderPlan, out_path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    defined: set[str] = set()
    sheet_headers, sheet_data_bounds = _collect_formula_context(plan)

    # First pass: create sheets deterministically
    for op in plan.ops:
        if isinstance(op, DefineSheet) and op.sheet not in defined:
            _get_ws(wb, op.sheet)
            defined.add(op.sheet)

    for s in plan.sheet_order:
        if s and s not in defined:
            _get_ws(wb, s)
            defined.add(s)

    if not defined:
        _get_ws(wb, "Sheet1")

    # Second pass: execute operations
    for op in plan.ops:

        if isinstance(op, DefineSheet):
            continue

        if isinstance(op, SetHeader):
            ws = _get_ws(wb, op.sheet)
            _write(ws, op.row, op.col, op.text)
            continue

        if isinstance(op, MergeCells):
            ws = _get_ws(wb, op.sheet)
            ws.merge_cells(
                start_row=op.r1, start_column=op.c1,
                end_row=op.r2, end_column=op.c2,
            )
            continue

        if isinstance(op, ApplyHeaderStyle):
            ws = _get_ws(wb, op.sheet)
            cell = ws.cell(row=op.row, column=op.col)
            if op.bold:
                cell.font = Font(bold=True)
            if op.fill_rgb:
                cell.fill = PatternFill("solid", fgColor=op.fill_rgb.lstrip("#"))
            continue

        if isinstance(op, ApplyColumnStyle):
            if op.fill_rgb:
                ws = _get_ws(wb, op.sheet)
                fill = PatternFill("solid", fgColor=op.fill_rgb.lstrip("#"))
                for r in range(op.from_row, op.to_row + 1):
                    ws.cell(row=r, column=op.col).fill = fill
            continue

        if isinstance(op, WriteDataBlock):
            ws = _get_ws(wb, op.sheet)
            for row_off, row_data in enumerate(op.data):
                for col_off, val in enumerate(row_data):
                    row = op.r1 + row_off
                    ws.cell(
                        row=row,
                        column=op.c1 + col_off,
                        value=_xlsx_cell_value(
                            val,
                            current_sheet=op.sheet,
                            row=row,
                            sheet_headers=sheet_headers,
                            sheet_data_bounds=sheet_data_bounds,
                        ),
                    )
            continue

        if isinstance(op, SetFreeze):
            ws = _get_ws(wb, op.sheet)
            ws.freeze_panes = f"{get_column_letter(op.col)}{op.row}"
            continue

        if isinstance(op, SetAutoFilter):
            ws = _get_ws(wb, op.sheet)
            ref = _area_to_ref(op.r1, op.c1, op.r2, op.c2)
            ws.auto_filter.ref = ref
            continue

        if isinstance(op, AddValidation):
            ws = _get_ws(wb, op.sheet)
            ref = _area_to_ref(op.r1, op.c1, op.r2, op.c2)
            dv = DataValidation(
                type="list",
                formula1=_xlsx_validation_formula(op.formula),
                allow_blank=op.allow_empty,
            )
            dv.add(ref)
            ws.add_data_validation(dv)
            continue

        if isinstance(op, WriteMeta):
            sheet_name = op.sheet or "_meta"
            ws = _get_ws(wb, sheet_name)
            row = 1
            for k, v in op.kv.items():
                ws.cell(row=row, column=1, value=str(k))
                ws.cell(row=row, column=2, value=str(v))
                row += 1
            if op.hidden:
                ws.sheet_state = "hidden"
            continue

        if isinstance(op, DefineNamedRange):
            ref = (
                f"'{op.sheet}'!"
                f"${get_column_letter(op.c1)}${op.r1}:"
                f"${get_column_letter(op.c2)}${op.r2}"
            )
            wb.defined_names.add(DefinedName(op.name, attr_text=ref))
            continue

    wb.save(out_path)


# --------------------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------------------

def render_workbook(plan: RenderPlan, out_path: Path | str) -> None:
    """Render a *RenderPlan* to an XLSX file via openpyxl."""
    _render_from_plan(plan, Path(out_path))


__all__ = ["render_workbook"]
