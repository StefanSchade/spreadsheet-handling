from __future__ import annotations

"""Spreadsheet-semantic worksheet interpretation for the XLSX read path."""

from typing import Any, Mapping

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_handling.rendering.ir import DataValidationSpec, SheetIR, TableBlock


OPTION_HINT_KEYS = (
    "freeze_header",
    "auto_filter",
    "header_fill_rgb",
    "helper_fill_rgb",
    "helper_prefix",
)


def build_sheet_meta_hints(
    workbook_meta: Mapping[str, Any],
    *,
    sheet_name: str,
) -> dict[str, Any]:
    """Merge workbook defaults with sheet-local overrides for one visible sheet."""
    meta_hints = {
        key: workbook_meta[key]
        for key in OPTION_HINT_KEYS
        if key in workbook_meta
    }

    sheet_meta_hints = (workbook_meta.get("sheets") or {}).get(sheet_name, {})
    if isinstance(sheet_meta_hints, dict):
        meta_hints.update(sheet_meta_hints)

    return meta_hints


def build_visible_sheet_ir(
    ws: Worksheet,
    *,
    sheet_name: str,
    meta_hints: Mapping[str, Any],
    validations: list[DataValidationSpec],
    freeze_hint: dict[str, int] | None,
    autofilter_ref: str | None,
    anchors: list[tuple[int, int]] | None,
    stop_on_empty_row: bool,
    stop_on_empty_col: bool,
) -> SheetIR:
    """Interpret a visible worksheet into spreadsheet-neutral ``SheetIR``."""
    sh = SheetIR(name=sheet_name)

    options = {}
    for key in OPTION_HINT_KEYS:
        if key in meta_hints:
            options[key] = meta_hints[key]
    if options:
        sh.meta["options"] = options

    table_starts = anchors or [(1, 1)]
    for top, left in table_starts:
        sh.tables.append(
            _parse_table_block(
                ws,
                frame_name=sheet_name,
                top=top,
                left=left,
                stop_on_empty_row=stop_on_empty_row,
                stop_on_empty_col=stop_on_empty_col,
            )
        )

    header_merges = _extract_header_merges(ws, sh.tables)
    if header_merges:
        sh.meta["__header_merges"] = header_merges

    for tbl in sh.tables:
        grid = _extract_header_grid(ws, tbl)
        if grid and tbl.header_rows > 1:
            sh.meta["__header_grid"] = grid

    sh.validations = list(validations)

    if freeze_hint:
        sh.meta["__freeze"] = dict(freeze_hint)

    if autofilter_ref:
        sh.meta["__autofilter_ref"] = autofilter_ref

    return sh


def _parse_table_block(
    ws: Worksheet,
    *,
    frame_name: str,
    top: int,
    left: int,
    stop_on_empty_row: bool,
    stop_on_empty_col: bool,
) -> TableBlock:
    """Discover a single table starting at ``(top, left)``."""
    header_rows = _detect_header_rows(ws, top, left)
    n_cols = _find_col_extent(ws, top, left, stop_on_empty_col)
    data_start_row = top + header_rows
    n_data_rows = _find_row_extent(ws, data_start_row, left, n_cols, stop_on_empty_row)
    n_rows = header_rows + n_data_rows

    headers: list[str] = []
    leaf_row = top + header_rows - 1
    for c in range(left, left + n_cols):
        val = _cell_value(ws, leaf_row, c)
        headers.append(str(val) if val else "")

    if header_rows > 1:
        flat_headers = []
        for c in range(left, left + n_cols):
            parts = []
            for r in range(top, top + header_rows):
                val = _cell_value(ws, r, c)
                parts.append(str(val) if val else "")
            flat_headers.append(" / ".join(p for p in parts if p))
        headers = flat_headers

    header_map = {h: idx + 1 for idx, h in enumerate(headers)}

    data: list[list[Any]] = []
    for r in range(data_start_row, data_start_row + n_data_rows):
        row_vals: list[Any] = []
        for c in range(left, left + n_cols):
            val = ws.cell(row=r, column=c).value
            row_vals.append(str(val) if val is not None else "")
        data.append(row_vals)

    return TableBlock(
        frame_name=frame_name,
        top=top,
        left=left,
        header_rows=header_rows,
        header_cols=1,
        n_rows=n_rows,
        n_cols=n_cols,
        headers=headers,
        header_map=header_map,
        data=data,
    )


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Get a cell value, resolving merged cells to their master value."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if (row, col) in [
                (r, c)
                for r in range(mr.min_row, mr.max_row + 1)
                for c in range(mr.min_col, mr.max_col + 1)
            ]:
                return ws.cell(row=mr.min_row, column=mr.min_col).value
        return None
    return cell.value


def _detect_header_rows(ws: Worksheet, top: int, left: int) -> int:
    """Detect header depth by examining merge regions near the table anchor."""
    max_header_row = top
    has_horizontal_merge_at_top = False

    for mr in ws.merged_cells.ranges:
        if mr.min_row < top or mr.min_col < left:
            continue
        if mr.max_row > mr.min_row and mr.min_row >= top and mr.max_row > max_header_row:
            max_header_row = mr.max_row
        if mr.min_row == top and mr.max_col > mr.min_col and mr.min_row == mr.max_row:
            has_horizontal_merge_at_top = True

    if max_header_row > top:
        return min(max_header_row - top + 1, 10)
    if has_horizontal_merge_at_top:
        return 2
    return 1


def _find_col_extent(ws: Worksheet, top: int, left: int, stop_on_empty: bool) -> int:
    """Find the number of columns by scanning the first header row."""
    n_cols = 0
    for c in range(left, left + 16384):
        val = _cell_value(ws, top, c)
        if val is None or (isinstance(val, str) and val.strip() == ""):
            if stop_on_empty:
                break
            has_more = False
            for lookahead in range(1, 4):
                v2 = _cell_value(ws, top, c + lookahead)
                if v2 is not None and str(v2).strip():
                    has_more = True
                    break
            if not has_more:
                break
        n_cols = c - left + 1
    return max(n_cols, 1)


def _find_row_extent(
    ws: Worksheet,
    data_start_row: int,
    left: int,
    n_cols: int,
    stop_on_empty: bool,
) -> int:
    """Find the number of data rows by scanning down from the first data row."""
    n_data_rows = 0
    for r in range(data_start_row, data_start_row + 1048576):
        row_empty = True
        for c in range(left, left + n_cols):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                row_empty = False
                break
        if row_empty:
            if stop_on_empty:
                break
            lookahead_empty = True
            for lr in range(1, 3):
                for c in range(left, left + n_cols):
                    val = ws.cell(row=r + lr, column=c).value
                    if val is not None and str(val).strip() != "":
                        lookahead_empty = False
                        break
                if not lookahead_empty:
                    break
            if lookahead_empty:
                break
        n_data_rows = r - data_start_row + 1
    return n_data_rows


def _extract_header_merges(
    ws: Worksheet,
    tables: list[TableBlock],
) -> list[tuple[int, int, int, int]]:
    """Extract merge regions within the header area of tables in relative coordinates."""
    merges: list[tuple[int, int, int, int]] = []
    for tbl in tables:
        header_bottom = tbl.top + tbl.header_rows - 1
        for mr in ws.merged_cells.ranges:
            if (
                mr.min_row >= tbl.top
                and mr.max_row <= header_bottom
                and mr.min_col >= tbl.left
                and mr.max_col <= tbl.left + tbl.n_cols - 1
            ):
                merges.append(
                    (
                        mr.min_row - tbl.top + 1,
                        mr.min_col - tbl.left + 1,
                        mr.max_row - tbl.top + 1,
                        mr.max_col - tbl.left + 1,
                    )
                )
    return merges


def _extract_header_grid(ws: Worksheet, tbl: TableBlock) -> list[list[str]]:
    """Extract the header grid as a 2D list of strings."""
    grid: list[list[str]] = []
    for r in range(tbl.top, tbl.top + tbl.header_rows):
        row_vals: list[str] = []
        for c in range(tbl.left, tbl.left + tbl.n_cols):
            val = _cell_value(ws, r, c)
            row_vals.append(str(val) if val else "")
        grid.append(row_vals)
    return grid


__all__ = [
    "OPTION_HINT_KEYS",
    "build_sheet_meta_hints",
    "build_visible_sheet_ir",
]
