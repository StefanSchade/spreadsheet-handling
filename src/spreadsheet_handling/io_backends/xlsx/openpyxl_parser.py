from __future__ import annotations

import ast
import csv
import io
import json
from pathlib import Path
import re
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_handling.io_backends.xlsx.parser_interpretation import (
    build_sheet_meta_hints,
    build_visible_sheet_ir,
)
from spreadsheet_handling.rendering.ir import (
    DataValidationSpec,
    NamedRange,
    SheetIR,
    WorkbookIR,
)
from spreadsheet_handling.rendering.formulas import ListLiteralFormulaSpec


def parse_workbook(path: str | Path) -> WorkbookIR:
    """
    Parse an XLSX workbook into ``WorkbookIR`` via openpyxl.

    Discovery priority for table structure:
      1. Embedded hidden `_meta` payload
      2. Explicit anchors passed by a future caller extension
      3. Heuristic fallback from sheet contents
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    try:
        ir = WorkbookIR()

        embedded_meta = _read_meta_sheet(wb)

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            if ws.sheet_state == "hidden":
                ir.hidden_sheets[ws_name] = _parse_hidden_sheet(ws)
                continue

            legend_hints = _legend_table_hints(embedded_meta, sheet_name=ws_name)
            legend_anchors = [
                (hint["top"], hint["left"])
                for hint in legend_hints
                if isinstance(hint.get("top"), int) and isinstance(hint.get("left"), int)
            ]
            anchors = [(1, 1), *legend_anchors] if legend_anchors else None
            meta_hints = build_sheet_meta_hints(embedded_meta, sheet_name=ws_name)
            ir.sheets[ws_name] = _parse_visible_sheet(
                ws,
                sheet_name=ws_name,
                anchors=anchors,
                meta_hints=meta_hints,
                stop_on_empty_row=True,
                stop_on_empty_col=bool(legend_anchors),
            )
            _apply_legend_table_hints(ir.sheets[ws_name], legend_hints)

        _extract_named_ranges(wb, ir)
        return ir
    finally:
        wb.close()


def _read_meta_sheet(wb: openpyxl.Workbook) -> dict[str, Any]:
    """Read the hidden `_meta` sheet if present and return a dict."""
    if "_meta" not in wb.sheetnames:
        return {}

    ws = wb["_meta"]
    kv: dict[str, str] = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        key = row[0]
        val = row[1] if len(row) > 1 else ""
        if key is not None:
            kv[str(key)] = str(val) if val is not None else ""

    blob_str = kv.get("workbook_meta_blob", "")
    if blob_str:
        try:
            return dict(json.loads(str(blob_str)))
        except (json.JSONDecodeError, TypeError):
            pass
        try:
            result = ast.literal_eval(blob_str)
            if isinstance(result, dict):
                return result
        except (ValueError, SyntaxError):
            pass

    return kv


def _parse_hidden_sheet(ws: Worksheet) -> SheetIR:
    """Parse a hidden sheet (for example `_meta`) into a ``SheetIR``."""
    sh = SheetIR(name=ws.title)
    sh.meta["_hidden"] = True
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        key = row[0]
        val = row[1] if len(row) > 1 else ""
        if key is not None:
            sh.meta[str(key)] = str(val) if val is not None else ""
    return sh


def _legend_table_hints(
    workbook_meta: dict[str, Any],
    *,
    sheet_name: str,
) -> list[dict[str, Any]]:
    raw = workbook_meta.get("legend_blocks") if isinstance(workbook_meta, dict) else None
    if not isinstance(raw, dict):
        return []

    hints: list[dict[str, Any]] = []
    for legend_name, spec in raw.items():
        if not isinstance(spec, dict):
            continue
        resolved = spec.get("resolved")
        if not isinstance(resolved, dict):
            continue
        if str(resolved.get("sheet")) != sheet_name:
            continue
        try:
            hints.append({
                "name": str(legend_name),
                "frame_name": str(resolved.get("frame_name") or f"legend_{legend_name}"),
                "top": int(resolved["top"]),
                "left": int(resolved["left"]),
            })
        except (KeyError, TypeError, ValueError):
            continue
    return hints


def _apply_legend_table_hints(sheet: SheetIR, hints: list[dict[str, Any]]) -> None:
    if not hints:
        return
    by_position = {
        (int(hint["top"]), int(hint["left"])): hint
        for hint in hints
    }
    sheet.meta["__legend_blocks"] = list(hints)
    for table in sheet.tables:
        hint = by_position.get((table.top, table.left))
        if not hint:
            continue
        table.kind = "legend"
        table.frame_name = str(hint["frame_name"])


def _parse_visible_sheet(
    ws: Worksheet,
    *,
    sheet_name: str,
    anchors: list[tuple[int, int]] | None,
    meta_hints: dict[str, Any],
    stop_on_empty_row: bool,
    stop_on_empty_col: bool,
) -> SheetIR:
    """Parse a visible worksheet into ``SheetIR`` by composing extraction and interpretation."""
    return build_visible_sheet_ir(
        ws,
        sheet_name=sheet_name,
        meta_hints=meta_hints,
        validations=_extract_validations(ws),
        freeze_hint=_extract_freeze(ws),
        autofilter_ref=ws.auto_filter.ref if ws.auto_filter and ws.auto_filter.ref else None,
        anchors=anchors,
        stop_on_empty_row=stop_on_empty_row,
        stop_on_empty_col=stop_on_empty_col,
    )


def _extract_validations(ws: Worksheet) -> list[DataValidationSpec]:
    """Extract list validations from the worksheet into ``DataValidationSpec`` values."""
    specs: list[DataValidationSpec] = []
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        formula = _parse_xlsx_list_literal_formula(str(dv.formula1) if dv.formula1 else "")
        if formula is None:
            continue
        allow_empty = bool(dv.allow_blank) if dv.allow_blank is not None else True
        for cell_range in dv.sqref.ranges:
            specs.append(
                DataValidationSpec(
                    kind="list",
                    area=(
                        cell_range.min_row,
                        cell_range.min_col,
                        cell_range.max_row,
                        cell_range.max_col,
                    ),
                    formula=formula,
                    allow_empty=allow_empty,
                )
            )
    return specs


def _parse_xlsx_list_literal_formula(formula: str) -> ListLiteralFormulaSpec | None:
    text = str(formula or "")
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        text = text[1:-1]
    elif text:
        return None
    if not text:
        return ListLiteralFormulaSpec(())
    try:
        values = next(csv.reader(io.StringIO(text), delimiter=",", quotechar='"'))
    except csv.Error:
        return None
    return ListLiteralFormulaSpec(tuple(values))


def _extract_freeze(ws: Worksheet) -> dict[str, int] | None:
    """Parse an openpyxl freeze-pane ref such as ``A3`` into a carrier hint."""
    fp = ws.freeze_panes
    if not fp:
        return None

    from openpyxl.utils import column_index_from_string

    col_str = ""
    row_str = ""
    for ch in str(fp):
        if ch.isalpha():
            col_str += ch
        else:
            row_str += ch
    if col_str and row_str:
        return {
            "row": int(row_str),
            "col": column_index_from_string(col_str),
        }
    return None


def _extract_named_ranges(wb: openpyxl.Workbook, ir: WorkbookIR) -> None:
    """Read workbook defined names and attach them to the matching visible sheet IR."""
    from openpyxl.utils import column_index_from_string as cifs

    for _, dn in wb.defined_names.items():
        for title, coord in dn.destinations:
            if title not in ir.sheets:
                continue
            match = re.match(r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", coord)
            if not match:
                continue
            c1 = cifs(match.group(1))
            r1 = int(match.group(2))
            c2 = cifs(match.group(3))
            r2 = int(match.group(4))
            ir.sheets[title].named_ranges.append(
                NamedRange(name=dn.name, sheet=title, area=(r1, c1, r2, c2))
            )


__all__ = ["parse_workbook"]
