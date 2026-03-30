from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional, Final
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

from spreadsheet_handling.io_backends.base import BackendBase, BackendOptions

import logging

# IR + composer + passes + renderer
from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
from spreadsheet_handling.rendering.passes import apply_all as apply_render_passes
from spreadsheet_handling.rendering.flow import build_render_plan
from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook



log = logging.getLogger("sheets.xlsx")

_RESERVED_FRAME_KEYS: Final[set[str]] = {"_meta"}   # extend here if we add more internals

# ======================================================================================
# Styling helpers
# ======================================================================================

def _decorate_workbook(
        workbook_path: Path,
        *,
        auto_filter: bool = True,
        header_fill_rgb: str = "DDDDDD",
        freeze_header: bool = False,
) -> None:
    """
    Post-process a written XLSX file:
    - apply AutoFilter across the used range
    - color header row (row 1) with a light gray fill & bold font
    - optionally freeze the first row (pane below header)
    """
    wb = load_workbook(workbook_path)

    header_fill = PatternFill("solid", fgColor=header_fill_rgb)
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        # AutoFilter across used range
        if auto_filter and ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions  # e.g. "A1:D100"

        # Header styling (row 1)
        if ws.max_column:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font

        # Freeze pane below header
        if freeze_header:
            ws.freeze_panes = "A2"

    wb.save(workbook_path)


# ======================================================================================
# Validation helpers (worksheet-header based)
# ======================================================================================

def _find_col_index_by_header(ws: Worksheet, header: str) -> Optional[int]:
    """
    Return 1-based column index for the given header text found in row 1.
    Match tolerant compare that ignores whitespace / case  differences.
    """
    want = str(header).strip().lower()
    for j, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip().lower() == want:
            return j
    return None


def _apply_xlsx_validations(
        xlsx_path: Path,
        frames: Dict[str, pd.DataFrame] | dict | Any,
) -> None:
    """
    Apply Excel data validations derived from frames meta.

    - Reads meta from frames.meta (attribute) OR frames["_meta"] (dict key).
    - Resolves the target column by inspecting the worksheet header (row 1),
      not by accessing DataFrame columns.
    - MVP supports rule.type == "in_list" on a named column.
    """
    meta = (
            getattr(frames, "meta", None)
            or (frames.get("_meta") if isinstance(frames, dict) else None)
            or {}
    )
    constraints = list(meta.get("constraints") or [])
    log.info("[xlsx] start validations: %d constraint(s)", len(constraints))
    if not constraints:
        return

    wb = load_workbook(xlsx_path)
    added = 0
    for c in constraints:
        rule = (c or {}).get("rule") or {}
        if rule.get("type") != "in_list":
            continue
        sheet = c.get("sheet")
        col = c.get("column")
        if not sheet or not col or sheet not in wb.sheetnames:
            log.warning("[xlsx] skip: bad sheet/col %s", c)
            continue
        ws = wb[sheet]
        col_ix = _find_col_index_by_header(ws, str(col))
        if not col_ix:
            log.warning("[xlsx] header not found: %s.%s", sheet, col)
            continue
        start_row = 2
        end_row = max(ws.max_row or 1, 2)
        rng = f"{get_column_letter(col_ix)}{start_row}:{get_column_letter(col_ix)}{end_row}"
        values = list(rule.get("values") or [])
        if not values:
            log.warning("[xlsx] no values for %s.%s", sheet, col)
            continue
        dv = DataValidation(type="list", formula1=f'"{",".join(map(str, values))}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)
        added += 1
        log.info("[xlsx] added: sheet=%s col=%s range=%s values=%s", sheet, col, rng, values)
    wb.save(xlsx_path)
    log.info("[xlsx] done: %d validations added", added)


# ======================================================================================
# Header flattening for writer
# ======================================================================================

def _flatten_header_to_level0(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure we write a single header row:
    - if columns is a MultiIndex, take level 0
    - otherwise keep as-is
    """
    if isinstance(df.columns, pd.MultiIndex):
        out = df.copy()
        out.columns = [t[0] for t in df.columns.to_list()]
        return out
    return df


def _flatten_cols_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure single-level string headers for Excel output.
    If already flat, stringify columns; if MultiIndex, choose the first non-empty
    component per tuple.
    """
    if not isinstance(df.columns, pd.MultiIndex):
        df = df.copy()
        df.columns = [str(c) for c in df.columns]
        return df
    def first_nonempty(tup) -> str:
        for x in tup:
            s = str(x)
            if s:
                return s
        return ""
    flat = [first_nonempty(t) for t in df.columns.tolist()]
    new_df = df.copy()
    new_df.columns = flat
    return new_df


# ======================================================================================
# Excel backend (multi-sheet only)
# ======================================================================================

class ExcelBackend(BackendBase):
    """
    XLSX adapter backed by pandas + openpyxl.

    Public API (used by our router/tests):
      - write_multi(frames, path, options=None)
      - read_multi(path, header_levels, options=None)
    """

    def write_multi(
            self,
            frames: dict[str, pd.DataFrame],
            path: str,
            options=None,
    ) -> None:
        """
        Write multiple DataFrames to an XLSX using the new IR pipeline.
        """
        use_ir = (
                (options and getattr(options, "use_ir_backend", None) is True)
                or os.getenv("SH_XLSX_BACKEND", "").lower() in {"ir","new","1","true"}
        )
        if not use_ir:
            # ---- LEGACY PATH (what you had before) ----
            out = Path(path).with_suffix(".xlsx")
            out.parent.mkdir(parents=True, exist_ok=True)
            with pd.ExcelWriter(out, engine="openpyxl") as xw:
                for sheet_name, df in frames.items():
                    name = str(sheet_name)
                    if name in _RESERVED_FRAME_KEYS:
                        continue
                    sheet = (name or "Sheet")[:31]
                    df0 = _ensure_dataframe(df)
                    df_out = _flatten_header_to_level0(df0)
                    df_out.to_excel(xw, sheet_name=sheet, index=False)
            _decorate_workbook(out, auto_filter=True, header_fill_rgb="DDDDDD", freeze_header=False)
            try:
                _apply_xlsx_validations(out, frames)
            except Exception:
                pass
            return

        # ---- NEW IR PATH ----
        out_path = Path(path).with_suffix(".xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Build IR first to know table geometry (header depth, offsets, etc.)
        meta = (frames.get("_meta") if isinstance(frames, dict) else {}) or getattr(frames, "meta", {}) or {}
        ir = compose_workbook(frames, meta)

        # Build IR → passes → render plan (now includes data cells)
        apply_render_passes(ir, meta)
        plan = build_render_plan(ir)
        render_workbook(plan, out_path)

    def read_multi(
            self,
            path: str,
            header_levels: int,
            options: BackendOptions | None = None,
    ) -> Dict[str, pd.DataFrame]:
        """
        Read all sheets assuming a single header row (header=0).
        If header_levels > 1, lift columns to a MultiIndex (level-0 data,
        remaining levels empty strings).

        Hidden sheets (e.g. ``_meta``) are excluded from data frames.
        If a ``_meta`` sheet is present its key/value pairs are extracted
        and returned as ``frames["_meta"]`` (a plain dict, not a DataFrame).

        When the IR backend is active (``SH_XLSX_BACKEND=ir``), the read
        path uses :func:`parse_ir` for geometry-aware extraction.
        """
        use_ir = (
                (options and getattr(options, "use_ir_backend", None) is True)
                or os.getenv("SH_XLSX_BACKEND", "").lower() in {"ir", "new", "1", "true"}
        )
        p = Path(path)

        if use_ir:
            return self._read_multi_ir(p)

        return self._read_multi_legacy(p, header_levels)

    # ------------------------------------------------------------------
    # IR read path
    # ------------------------------------------------------------------

    def _read_multi_ir(self, p: Path) -> Dict[str, pd.DataFrame]:
        from spreadsheet_handling.rendering.parse_ir import parse_ir, workbookir_to_frames
        ir = parse_ir(p)
        return workbookir_to_frames(ir)

    # ------------------------------------------------------------------
    # Legacy (pandas) read path
    # ------------------------------------------------------------------

    def _read_multi_legacy(self, p: Path, header_levels: int) -> Dict[str, pd.DataFrame]:
        # --- detect hidden sheets via openpyxl before pandas reads -----------
        wb = load_workbook(p, read_only=True, data_only=True)
        hidden_names: set[str] = set()
        try:
            for ws in wb.worksheets:
                if ws.sheet_state == "hidden":
                    hidden_names.add(ws.title)
        finally:
            wb.close()

        # --- extract meta from hidden _meta sheet (if present) ---------------
        meta_dict: Dict[str, Any] = {}
        if "_meta" in hidden_names:
            from spreadsheet_handling.rendering.parse_ir import _read_meta_sheet
            wb2 = load_workbook(p, data_only=True)
            try:
                meta_dict = _read_meta_sheet(wb2)
            finally:
                wb2.close()

        # --- read data sheets via pandas -------------------------------------
        sheets = pd.read_excel(p, sheet_name=None, header=0, engine="openpyxl", dtype=str)
        out: Dict[str, pd.DataFrame] = {}
        levels = header_levels if (header_levels and header_levels > 1) else 1

        for name, df in sheets.items():
            if name in hidden_names:
                continue
            df = df.where(pd.notnull(df), "")  # normalize NaNs to empty strings
            if not isinstance(df.columns, pd.MultiIndex) and levels > 1:
                tuples = [(c,) + ("",) * (levels - 1) for c in list(df.columns)]
                df = df.copy()
                df.columns = pd.MultiIndex.from_tuples(tuples)
            out[name] = df

        if meta_dict:
            out["_meta"] = meta_dict  # type: ignore[assignment]

        return out


# ======================================================================================
# Router-facing convenience API
# ======================================================================================

def save_xlsx(
        frames: Dict[str, Any],
        path: str,
        options: BackendOptions | None = None,
) -> None:
    """
    Preserve meta under '_meta' (either from frames['_meta'] or frames.meta)
    so the writer can apply data validations; coerce all sheets to DataFrames.
    """
    sanitized: Dict[str, Any] = {}

    # carry over meta (either attribute or dict key)
    meta_attr = getattr(frames, "meta", None)
    if isinstance(frames, dict) and "_meta" in frames:
        sanitized["_meta"] = frames["_meta"]
    elif meta_attr is not None:
        sanitized["_meta"] = meta_attr

    for name, df in frames.items():
        name_str = str(name)
        if name_str in _RESERVED_FRAME_KEYS:
            # keep meta entry in mapping (writer will skip writing it as a sheet)
            sanitized[name_str] = df
            continue
        df0 = _ensure_dataframe(df)
        sanitized[name_str] = _flatten_cols_for_excel(df0)

    ExcelBackend().write_multi(sanitized, path, options=options)


def load_xlsx(
        path: str,
        options: BackendOptions | None = None,
) -> Dict[str, pd.DataFrame]:
    """
    Router-facing reader: read all sheets, assume single header row.
    (Lifts to MultiIndex of length 1 → effectively stays flat.)
    """
    return ExcelBackend().read_multi(path, header_levels=1, options=options)


def _ensure_dataframe(obj: Any) -> pd.DataFrame:
    """
    Coerce various tabular shapes to a DataFrame:
    - DataFrame -> itself
    - list[dict] / list[list] -> DataFrame(obj)
    - dict[str, list] / dict[str, scalar] -> DataFrame(obj)
    """
    if isinstance(obj, pd.DataFrame):
        return obj
    return pd.DataFrame(obj)


# keep near the end of xlsx_backend.py (module-level helper for tests)
def write_xlsx(
        path: str,
        frames: Dict[str, pd.DataFrame],
        meta: Any,  # kept for signature compatibility; not used here
        ctx: Any,   # may carry styling opts; we ignore for now (ExcelBackend styles by default)
) -> None:
    """
    Test-facing convenience used by unit tests.
    Delegates to ExcelBackend().write_multi with default styling and validations.
    """
    ExcelBackend().write_multi(frames, path, options=None)

__all__ = ["ExcelBackend", "save_xlsx", "load_xlsx"]
