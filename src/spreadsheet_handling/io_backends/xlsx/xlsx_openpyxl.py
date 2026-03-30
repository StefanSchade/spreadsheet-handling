from __future__ import annotations
from pathlib import Path
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation

# New plan-based API
from spreadsheet_handling.rendering.plan import (
    RenderPlan,
    RenderOp,
    DefineSheet,
    SetHeader,
    ApplyHeaderStyle,
    ApplyColumnStyle,
    SetAutoFilter,
    SetFreeze,
    AddValidation,
    WriteMeta,
)

# --- Plan-based renderer (preferred) ---

def render_plan(plan: RenderPlan, out_path: Path | str) -> None:
    """
    Render a backend-agnostic RenderPlan to XLSX using OpenPyXL.
    This is the new, preferred code path.
    """
    out_p = Path(out_path)
    out_p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Pre-create sheets in order
    sheets: Dict[str, object] = {}
    for op in plan.ops:
        if isinstance(op, DefineSheet):
            ws = wb.create_sheet(title=op.sheet)
            sheets[op.sheet] = ws

    # Apply remaining ops
    for op in plan.ops:
        if isinstance(op, DefineSheet):
            continue

        if isinstance(op, SetHeader):
            ws = sheets[op.sheet]
            ws.cell(row=op.row, column=op.col, value=op.text)

        elif isinstance(op, ApplyHeaderStyle):
            ws = sheets[op.sheet]
            cell = ws.cell(row=op.row, column=op.col)
            if op.bold:
                cell.font = Font(bold=True)
            if op.fill_rgb:
                cell.fill = PatternFill("solid", fgColor=op.fill_rgb.lstrip("#"))

        elif isinstance(op, ApplyColumnStyle):
            ws = sheets[op.sheet]
            fill_rgb = op.fill_rgb
            if fill_rgb:
                fill = PatternFill("solid", fgColor=fill_rgb.lstrip("#"))
                for r in range(op.from_row, op.to_row + 1):
                    ws.cell(row=r, column=op.col).fill = fill

        elif isinstance(op, SetAutoFilter):
            ws = sheets[op.sheet]
            tl = ws.cell(row=op.r1, column=op.c1).coordinate
            br = ws.cell(row=op.r2, column=op.c2).coordinate
            ws.auto_filter.ref = f"{tl}:{br}"

        elif isinstance(op, SetFreeze):
            ws = sheets[op.sheet]
            ws.freeze_panes = ws.cell(row=op.row, column=op.col)

        elif isinstance(op, AddValidation):
            ws = sheets[op.sheet]
            if op.kind == "list":
                dv = DataValidation(type="list", formula1=op.formula, allow_blank=op.allow_empty)
                ws.add_data_validation(dv)
                tl = ws.cell(row=op.r1, column=op.c1).coordinate
                br = ws.cell(row=op.r2, column=op.c2).coordinate
                dv.add(f"{tl}:{br}")
            else:
                # Other kinds can be implemented later
                pass

        elif isinstance(op, WriteMeta):
            ws = wb.create_sheet(title=op.sheet)
            row = 1
            for k, v in op.kv.items():
                ws.cell(row=row, column=1, value=str(k))
                ws.cell(row=row, column=2, value=str(v))
                row += 1
            if op.hidden:
                ws.sheet_state = "hidden"

    wb.save(out_p)


