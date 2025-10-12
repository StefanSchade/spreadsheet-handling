from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

from .base import BackendBase, BackendOptions


# ======================================================================================
# Styling helpers
# ======================================================================================

def _decorate_workbook(
        workbook_path: Path,
        *,
        auto_filter: bool = True,
        header_fill_rgb: str = "DDDDDD",
        freeze_header: bool = False,
) -> None:
    """
    Post-process a written XLSX file:
    - apply AutoFilter across the used range
    - color header row (row 1) with a light gray fill & bold font
    - optionally freeze the first row (pane below header)
    """
    wb = load_workbook(workbook_path)

    header_fill = PatternFill("solid", fgColor=header_fill_rgb)
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        # AutoFilter across used range
        if auto_filter and ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions  # e.g. "A1:D100"

        # Header styling (row 1)
        if ws.max_column:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font

        # Freeze pane below header
        if freeze_header:
            ws.freeze_panes = "A2"

    wb.save(workbook_path)


# ======================================================================================
# Validation helpers (worksheet-header based)
# ======================================================================================

def _find_col_index_by_header(ws: Worksheet, header: str) -> Optional[int]:
    """
    Return 1-based column index for the given header text found in row 1.
    Strict string comparison.
    """
    for j, cell in enumerate(ws[1], start=1):
        if str(cell.value) == str(header):
            return j
    return None


def _apply_xlsx_validations(
        xlsx_path: Path,
        frames: Dict[str, pd.DataFrame] | dict | Any,
) -> None:
    """
    Apply Excel data validations derived from frames meta.

    - Reads meta from frames.meta (attribute) OR frames["_meta"] (dict key).
    - Resolves the target column by inspecting the worksheet header (row 1),
      not by accessing DataFrame columns. This works regardless of how frames are
      represented in memory (DataFrame, list-of-dicts, etc.).
    - MVP supports rule.type == "in_list" on a named column.
    """
    meta = (
            getattr(frames, "meta", None)
            or (frames.get("_meta") if isinstance(frames, dict) else None)
            or {}
    )
    constraints: list[dict] = meta.get("constraints") or []
    if not constraints:
        return

    wb = load_workbook(xlsx_path)

    for c in constraints:
        rule = (c or {}).get("rule") or {}
        if rule.get("type") != "in_list":
            continue

        sheet = c.get("sheet")
        column_name = c.get("column")
        if not sheet or not column_name or sheet not in wb.sheetnames:
            continue

        ws: Worksheet = wb[sheet]
        col_ix = _find_col_index_by_header(ws, str(column_name))
        if not col_ix:
            continue

        start_row = 2                       # assume row 1 = header
        end_row = max(ws.max_row or 1, 2)   # at least two rows
        col_letter = get_column_letter(col_ix)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

        values = list(rule.get("values") or [])
        if not values:
            continue

        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(map(str, values))}"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)

    wb.save(xlsx_path)


# ======================================================================================
# Header flattening for writer
# ======================================================================================

def _flatten_header_to_level0(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure we write a single header row:
    - if columns is a MultiIndex, take level 0
    - otherwise keep as-is
    """
    if isinstance(df.columns, pd.MultiIndex):
        out = df.copy()
        out.columns = [t[0] for t in df.columns.to_list()]
        return out
    return df


def _flatten_cols_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure single-level string headers for Excel output.
    If already flat, stringify columns; if MultiIndex, choose the first non-empty
    component per tuple.
    """
    if not isinstance(df.columns, pd.MultiIndex):
        out = df.copy()
        out.columns = [str(c) for c in out.columns]
        return out

    def first_nonempty(tup: tuple[Any, ...]) -> str:
        for x in tup:
            s = "" if x is None else str(x)
            if s:
                return s
        return ""

    flat = [first_nonempty(t) for t in df.columns.tolist()]
    new_df = df.copy()
    new_df.columns = flat
    return new_df


# ======================================================================================
# Excel backend (multi-sheet only)
# ======================================================================================

class ExcelBackend(BackendBase):
    """
    XLSX adapter backed by pandas + openpyxl.

    Public API (used by our router/tests):
      - write_multi(frames, path, options=None)
      - read_multi(path, header_levels, options=None)
    """

    def write_multi(
            self,
            frames: Dict[str, pd.DataFrame],
            path: str,
            options: BackendOptions | None = None,
    ) -> None:
        """
        Write multiple DataFrames to an XLSX:
        - one sheet per dict key
        - flatten MultiIndex headers to level 0 (single header row)
        - apply header styling (autofilter + gray + bold), optionally freeze header
        - apply data validations from frames meta (if present)
        """
        out = Path(path).with_suffix(".xlsx")
        out.parent.mkdir(parents=True, exist_ok=True)

        # 1) write data
        with pd.ExcelWriter(out, engine="openpyxl") as xw:
            for sheet_name, df in frames.items():
                sheet = (sheet_name or "Sheet")[:31]  # Excel sheet name limit
                df0 = _ensure_dataframe(df)
                df_out = _flatten_header_to_level0(df0)
                df_out.to_excel(xw, sheet_name=sheet, index=False)

        # 2) style
        _decorate_workbook(out, auto_filter=True, header_fill_rgb="DDDDDD", freeze_header=False)

        # 3) validations (never fail the write because of presentation)
        try:
            _apply_xlsx_validations(out, frames)
        except Exception:
            pass

    def read_multi(
            self,
            path: str,
            header_levels: int,
            options: BackendOptions | None = None,
    ) -> Dict[str, pd.DataFrame]:
        """
        Read all sheets assuming a single header row (header=0).
        If header_levels > 1, lift columns to a MultiIndex (level-0 data,
        remaining levels empty strings).
        """
        p = Path(path)
        sheets = pd.read_excel(p, sheet_name=None, header=0, engine="openpyxl", dtype=str)
        out: Dict[str, pd.DataFrame] = {}
        levels = header_levels if (header_levels and header_levels > 1) else 1

        for name, df in sheets.items():
            df = df.where(pd.notnull(df), "")  # normalize NaNs to empty strings
            if not isinstance(df.columns, pd.MultiIndex) and levels > 1:
                tuples = [(c,) + ("",) * (levels - 1) for c in list(df.columns)]
                df = df.copy()
                df.columns = pd.MultiIndex.from_tuples(tuples)
            out[name] = df

        return out


# ======================================================================================
# Router-facing convenience API
# ======================================================================================

def save_xlsx(
        frames: Dict[str, pd.DataFrame],
        path: str,
        options: BackendOptions | None = None,
) -> None:
    """Router-facing saver that guarantees flat string headers."""
    sanitized: Dict[str, pd.DataFrame] = {}
    for name, df in frames.items():
        df0 = _ensure_dataframe(df)
        sanitized[name] = _flatten_cols_for_excel(df0)
    ExcelBackend().write_multi(sanitized, path, options=options)


def load_xlsx(
        path: str,
        options: BackendOptions | None = None,
) -> Dict[str, pd.DataFrame]:
    """
    Router-facing reader: read all sheets, assume single header row.
    (Lifts to MultiIndex of length 1 → effectively stays flat.)
    """
    return ExcelBackend().read_multi(path, header_levels=1, options=options)


def _ensure_dataframe(obj: Any) -> pd.DataFrame:
    """
    Coerce various tabular shapes to a DataFrame:
    - DataFrame -> itself
    - list[dict] / list[list] -> DataFrame(obj)
    - dict[str, list] / dict[str, scalar] -> DataFrame(obj)
    """
    if isinstance(obj, pd.DataFrame):
        return obj
    return pd.DataFrame(obj)


# keep near the end of xlsx_backend.py (module-level helper for tests)
def write_xlsx(
        path: str,
        frames: Dict[str, pd.DataFrame],
        meta: Any,  # kept for signature compatibility; not used here
        ctx: Any,   # may carry styling opts; we ignore for now (ExcelBackend styles by default)
) -> None:
    """
    Test-facing convenience used by unit tests.
    Delegates to ExcelBackend().write_multi with default styling and validations.
    """
    ExcelBackend().write_multi(frames, path, options=None)

__all__ = ["ExcelBackend", "save_xlsx", "load_xlsx"]
