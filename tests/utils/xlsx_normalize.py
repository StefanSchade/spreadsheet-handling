from __future__ import annotations
from openpyxl import load_workbook

def normalize_xlsx(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    out = {"sheets": [], "styles": {}, "filters": {}, "freeze": {}, "validations": {}, "meta": {}}

    for ws in wb.worksheets:
        out["sheets"].append(ws.title)

        header = {}
        for cell in ws[1]:
            if cell.value is None:
                continue
            rgb = None
            if cell.fill and getattr(cell.fill, "fgColor", None) and cell.fill.fgColor.type == "rgb":
                rgb = "#" + cell.fill.fgColor.rgb[-6:]
            header[cell.coordinate] = {"value": cell.value, "bold": bool(cell.font and cell.font.b), "fill": rgb}
        if header:
            out["styles"][ws.title] = {"header": header}

        if ws.auto_filter and ws.auto_filter.ref:
            out["filters"][ws.title] = ws.auto_filter.ref

        # ws.freeze_panes is a string reference, not a cell object. the test has to convert to strings for comparison
        # using actual cells (coerce ws.freeze_panes to a cell ws["A2"]) would make our code dependent on OpenPyXL internals

        if ws.freeze_panes:
            # OpenPyXL >=3.1 returns a string like "A2"
            coord = ws.freeze_panes if isinstance(ws.freeze_panes, str) else getattr(ws.freeze_panes, "coordinate", None)
            out["freeze"][ws.title] = coord

        dvs = []
        if ws.data_validations:
            for dv in ws.data_validations.dataValidation:
                dvs.append({
                    "type": dv.type,
                    "allow_blank": bool(dv.allowBlank),
                    "sqref": " ".join(str(r) for r in dv.sqref),
                    "formula1": dv.formula1,
                })
        if dvs:
            out["validations"][ws.title] = dvs

        if ws.title == "_meta":
            kv = {}
            for row in ws.iter_rows(min_row=1, max_col=2):
                k = row[0].value
                v = row[1].value if len(row) > 1 else None
                if k:
                    kv[str(k)] = str(v) if v is not None else ""
            out["meta"] = kv

    return out
