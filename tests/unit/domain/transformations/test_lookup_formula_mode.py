"""Tests for FTR-LOOKUP-HELPER-FORMULA-MODE-P4A.

Covers:
- enrich_lookup with helper_value_mode='formula' emits LookupFormulaSpec cells
- Formula sheet-name resolution in select_render_frames
- XLSX end-to-end rendering of lookup formulas
- ODS rendering parity
- Backward compatibility: value mode remains default
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pytest

from spreadsheet_handling.domain.transformations.enrich_lookup import enrich_lookup
from spreadsheet_handling.core.formulas import LookupFormulaSpec

pytestmark = pytest.mark.ftr("FTR-LOOKUP-HELPER-FORMULA-MODE-P4A")


def _frames() -> dict[str, Any]:
    variables = pd.DataFrame({
        "variable_id": ["v1", "v2", "v3"],
        "label_de": ["Eins", "Zwei", "Drei"],
        "data_type": ["string", "int", "bool"],
    })
    matrix = pd.DataFrame({
        "variable_id": ["v1", "v2"],
        "P-001": ["E", "A"],
    })
    return {
        "variables": variables,
        "matrix_raw": matrix,
        "_meta": {},
    }


class TestFormulaMode:

    def test_formula_mode_emits_lookup_formula_specs(self) -> None:
        frames = _frames()
        out = enrich_lookup(
            frames,
            source="matrix_raw",
            lookup="variables",
            output="matrix",
            key="variable_id",
            helpers={"fields": ["label_de"]},
            missing="empty",
            helper_value_mode="formula",
        )
        df = out["matrix"]
        assert "label_de" in df.columns
        cell = df["label_de"].iloc[0]
        assert isinstance(cell, LookupFormulaSpec)
        assert cell.source_key_column == "variable_id"
        assert cell.lookup_sheet == "variables"
        assert cell.lookup_key_column == "variable_id"
        assert cell.lookup_value_column == "label_de"

    def test_formula_mode_multiple_fields(self) -> None:
        frames = _frames()
        out = enrich_lookup(
            frames,
            source="matrix_raw",
            lookup="variables",
            output="matrix",
            key="variable_id",
            helpers={"fields": ["label_de", "data_type"]},
            missing="empty",
            helper_value_mode="formula",
        )
        df = out["matrix"]
        for field in ("label_de", "data_type"):
            cell = df[field].iloc[0]
            assert isinstance(cell, LookupFormulaSpec)
            assert cell.lookup_value_column == field

    @pytest.mark.ftr("FTR-REVIEW-001-FORMULAS-CORE-MOVE-P3")
    def test_formula_mode_rejects_fail_missing_mode(self) -> None:
        frames = _frames()

        with pytest.raises(ValueError, match="missing='fail'"):
            enrich_lookup(
                frames,
                source="matrix_raw",
                lookup="variables",
                output="matrix",
                key="variable_id",
                helpers={"fields": ["label_de"]},
                missing="fail",
                helper_value_mode="formula",
            )

    def test_value_mode_remains_default(self) -> None:
        frames = _frames()
        out = enrich_lookup(
            frames,
            source="matrix_raw",
            lookup="variables",
            output="matrix",
            key="variable_id",
            helpers={"fields": ["label_de"]},
            missing="empty",
        )
        df = out["matrix"]
        assert df["label_de"].iloc[0] == "Eins"
        assert not isinstance(df["label_de"].iloc[0], LookupFormulaSpec)

    def test_explicit_values_mode_still_merges(self) -> None:
        frames = _frames()
        out = enrich_lookup(
            frames,
            source="matrix_raw",
            lookup="variables",
            output="matrix",
            key="variable_id",
            helpers={"fields": ["label_de"]},
            missing="empty",
            helper_value_mode="values",
        )
        df = out["matrix"]
        assert df["label_de"].iloc[0] == "Eins"

    def test_formula_mode_writes_provenance(self) -> None:
        frames = _frames()
        out = enrich_lookup(
            frames,
            source="matrix_raw",
            lookup="variables",
            output="matrix",
            key="variable_id",
            helpers={"fields": ["label_de"]},
            missing="empty",
            helper_value_mode="formula",
        )
        prov = out["_meta"]["derived"]["sheets"]["matrix"]["enrich_lookup"]
        assert prov["helper_columns"] == ["label_de"]
        assert prov["lookup"] == "variables"

    def test_invalid_value_mode_raises(self) -> None:
        frames = _frames()
        with pytest.raises(ValueError, match="helper_value_mode"):
            enrich_lookup(
                frames,
                source="matrix_raw",
                lookup="variables",
                output="matrix",
                key="variable_id",
                helpers={"fields": ["label_de"]},
                missing="empty",
                helper_value_mode="invalid",
            )


class TestFormulaSheetNameResolution:

    def test_formulas_resolve_to_physical_sheet_names(self) -> None:
        from spreadsheet_handling.rendering.frame_selection import select_render_frames

        formula = LookupFormulaSpec(
            source_key_column="variable_id",
            lookup_sheet="variables",
            lookup_key_column="variable_id",
            lookup_value_column="label_de",
            missing="",
        )
        frames = {
            "variables": pd.DataFrame({
                "variable_id": ["v1"],
                "label_de": ["Eins"],
            }),
            "matrix": pd.DataFrame({
                "variable_id": ["v1"],
                "label_de": [formula],
            }),
            "_meta": {
                "workbook_view": {
                    "sheets": [
                        {"frame": "variables", "sheet": "Entities"},
                        {"frame": "matrix", "sheet": "Matrix"},
                    ],
                }
            },
        }
        selected = select_render_frames(frames, frames["_meta"])

        cell = selected["Matrix"]["label_de"].iloc[0]
        assert isinstance(cell, LookupFormulaSpec)
        assert cell.lookup_sheet == "Entities"

    def test_no_rename_preserves_frame_name(self) -> None:
        from spreadsheet_handling.rendering.frame_selection import select_render_frames

        formula = LookupFormulaSpec(
            source_key_column="variable_id",
            lookup_sheet="variables",
            lookup_key_column="variable_id",
            lookup_value_column="label_de",
            missing="",
        )
        frames = {
            "variables": pd.DataFrame({
                "variable_id": ["v1"],
                "label_de": ["Eins"],
            }),
            "matrix": pd.DataFrame({
                "variable_id": ["v1"],
                "label_de": [formula],
            }),
            "_meta": {
                "workbook_view": {
                    "sheets": [
                        {"frame": "variables", "sheet": "variables"},
                        {"frame": "matrix", "sheet": "matrix"},
                    ],
                }
            },
        }
        selected = select_render_frames(frames, frames["_meta"])

        cell = selected["matrix"]["label_de"].iloc[0]
        assert isinstance(cell, LookupFormulaSpec)
        assert cell.lookup_sheet == "variables"


class TestXlsxLookupFormulaRendering:

    def test_xlsx_renders_xlookup_formulas(self, tmp_path: Path) -> None:
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.rendering.flow import (
            apply_ir_passes,
            build_render_plan,
            default_p1_passes,
        )
        from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook
        from openpyxl import load_workbook

        formula = LookupFormulaSpec(
            source_key_column="variable_id",
            lookup_sheet="variables",
            lookup_key_column="variable_id",
            lookup_value_column="label_de",
            missing="",
        )
        frames = {
            "variables": pd.DataFrame({
                "variable_id": ["v1", "v2"],
                "label_de": ["Eins", "Zwei"],
            }),
            "matrix": pd.DataFrame({
                "variable_id": ["v1", "v2"],
                "label_de": [formula, formula],
            }),
        }
        ir = compose_workbook(frames, None)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "formula.xlsx"
        render_workbook(plan, out)

        wb = load_workbook(out)
        ws = wb["matrix"]
        cell = ws.cell(row=2, column=2)
        assert cell.value is not None
        assert "XLOOKUP" in str(cell.value)
        assert "variables" in str(cell.value)


class TestOdsLookupFormulaRendering:

    def test_ods_renders_xlookup_formulas(self, tmp_path: Path) -> None:
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.rendering.flow import (
            apply_ir_passes,
            build_render_plan,
            default_p1_passes,
        )
        from spreadsheet_handling.io_backends.ods.odf_renderer import render_workbook

        formula = LookupFormulaSpec(
            source_key_column="variable_id",
            lookup_sheet="variables",
            lookup_key_column="variable_id",
            lookup_value_column="label_de",
            missing="",
        )
        frames = {
            "variables": pd.DataFrame({
                "variable_id": ["v1", "v2"],
                "label_de": ["Eins", "Zwei"],
            }),
            "matrix": pd.DataFrame({
                "variable_id": ["v1", "v2"],
                "label_de": [formula, formula],
            }),
        }
        ir = compose_workbook(frames, None)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "formula.ods"
        render_workbook(plan, out)
        assert out.exists()


class TestEndToEndFormulaEnrichment:

    def test_full_pipeline_enrich_render_xlsx(self, tmp_path: Path) -> None:
        from spreadsheet_handling.domain.workbook_views import configure_workbook_view
        from spreadsheet_handling.io_backends.spreadsheet_contract import (
            build_spreadsheet_render_plan,
        )
        from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook
        from openpyxl import load_workbook

        variables = pd.DataFrame({
            "variable_id": ["v1", "v2"],
            "label_de": ["Eins", "Zwei"],
        })
        matrix_raw = pd.DataFrame({
            "variable_id": ["v1", "v2"],
            "P-001": ["E", "A"],
        })
        frames: dict[str, Any] = {
            "variables": variables,
            "matrix_raw": matrix_raw,
            "_meta": {},
        }

        enriched = enrich_lookup(
            frames,
            source="matrix_raw",
            lookup="variables",
            output="matrix",
            key="variable_id",
            helpers={"fields": ["label_de"]},
            missing="empty",
            helper_value_mode="formula",
        )

        viewed = configure_workbook_view(
            enriched,
            sheets=[
                {"frame": "variables", "sheet": "Entities"},
                {"frame": "matrix", "sheet": "Matrix"},
            ],
        )

        plan = build_spreadsheet_render_plan(viewed, viewed.get("_meta"))
        out = tmp_path / "e2e.xlsx"
        render_workbook(plan, out)

        wb = load_workbook(out)
        ws = wb["Matrix"]
        # Matrix columns: variable_id (1), P-001 (2), label_de helper (3)
        cell = ws.cell(row=2, column=3)
        formula_text = str(cell.value)
        assert "XLOOKUP" in formula_text
        # Formula should reference 'Entities' (physical sheet name), not 'variables'
        assert "Entities" in formula_text
