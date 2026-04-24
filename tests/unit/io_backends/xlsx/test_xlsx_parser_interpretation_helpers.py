from __future__ import annotations

from openpyxl import Workbook

import pytest

from spreadsheet_handling.io_backends.xlsx.parser_interpretation import (
    build_sheet_meta_hints,
    build_visible_sheet_ir,
)
from spreadsheet_handling.rendering.ir import DataValidationSpec


pytestmark = pytest.mark.ftr("FTR-XLSX-PARSER-MODULARIZATION-P3I")


def test_build_sheet_meta_hints_merges_workbook_defaults_and_sheet_overrides():
    workbook_meta = {
        "freeze_header": True,
        "auto_filter": True,
        "helper_prefix": "_",
        "sheets": {
            "Products": {
                "freeze_header": False,
                "header_fill_rgb": "#CCE5FF",
            }
        },
    }

    hints = build_sheet_meta_hints(workbook_meta, sheet_name="Products")

    assert hints == {
        "freeze_header": False,
        "auto_filter": True,
        "helper_prefix": "_",
        "header_fill_rgb": "#CCE5FF",
    }


def test_build_visible_sheet_ir_combines_interpretation_inputs():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="title")
    ws.cell(row=2, column=1, value="P-001")
    ws.cell(row=2, column=2, value="Alpha")

    validations = [
        DataValidationSpec(
            kind="list",
            area=(2, 2, 2, 2),
            formula='"new,done"',
            allow_empty=True,
        )
    ]

    sheet = build_visible_sheet_ir(
        ws,
        sheet_name="Products",
        meta_hints={"freeze_header": True, "auto_filter": True},
        validations=validations,
        freeze_hint={"row": 2, "col": 1},
        autofilter_ref="A1:B2",
        anchors=None,
        stop_on_empty_row=True,
        stop_on_empty_col=False,
    )

    assert sheet.meta["options"] == {"freeze_header": True, "auto_filter": True}
    assert sheet.meta["__freeze"] == {"row": 2, "col": 1}
    assert sheet.meta["__autofilter_ref"] == "A1:B2"
    assert sheet.validations == validations
    assert sheet.tables[0].headers == ["id", "title"]
    assert sheet.tables[0].data == [["P-001", "Alpha"]]
