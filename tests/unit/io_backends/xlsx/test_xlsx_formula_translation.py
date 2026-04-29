from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import pytest

from spreadsheet_handling.io_backends.xlsx.openpyxl_parser import _parse_xlsx_list_literal_formula
from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import (
    _xlsx_lookup_formula,
    _xlsx_validation_formula,
    render_workbook,
)
from spreadsheet_handling.rendering.formulas import ListLiteralFormulaSpec, lookup_formula
from spreadsheet_handling.rendering.plan import DefineSheet, RenderPlan, SetHeader, WriteDataBlock


pytestmark = pytest.mark.ftr("FTR-FORMULA-PROVIDERS")


def test_xlsx_translates_list_literal_formula_spec() -> None:
    formula = ListLiteralFormulaSpec(("new", "done"))

    assert _xlsx_validation_formula(formula) == '"new,done"'


def test_xlsx_parser_recovers_list_literal_formula_spec() -> None:
    assert _parse_xlsx_list_literal_formula('"new,done"') == ListLiteralFormulaSpec(
        ("new", "done")
    )


def test_xlsx_translator_and_parser_preserve_commas_and_quotes() -> None:
    formula = ListLiteralFormulaSpec(("needs,review", 'he said "yes"'))

    rendered = _xlsx_validation_formula(formula)

    assert _parse_xlsx_list_literal_formula(rendered) == formula


@pytest.mark.ftr("FTR-FORMULA-FK-HELPER-PROVIDERS-P4A")
def test_xlsx_translates_lookup_formula_spec() -> None:
    formula = lookup_formula(
        source_key_column="id_(Customers)",
        lookup_sheet="Customers",
        lookup_key_column="id",
        lookup_value_column="name",
    )

    rendered = _xlsx_lookup_formula(
        formula,
        current_sheet="Orders",
        row=5,
        sheet_headers={
            "Orders": {"id": 1, "id_(Customers)": 2, "_Customers_name": 3},
            "Customers": {"id": 1, "name": 2},
        },
        sheet_data_bounds={"Orders": (2, 5), "Customers": (2, 4)},
    )

    assert rendered == (
        "=XLOOKUP($B5,'Customers'!$A$2:$A$4,'Customers'!$B$2:$B$4,\"\")"
    )


@pytest.mark.ftr("FTR-FORMULA-FK-HELPER-PROVIDERS-P4A")
def test_xlsx_renderer_emits_lookup_formula_cells(tmp_path: Path) -> None:
    formula = lookup_formula(
        source_key_column="id_(Customers)",
        lookup_sheet="Customers",
        lookup_key_column="id",
        lookup_value_column="name",
    )
    plan = RenderPlan()
    plan.add(DefineSheet("Orders", 0))
    plan.add(SetHeader("Orders", 1, 1, "id"))
    plan.add(SetHeader("Orders", 1, 2, "id_(Customers)"))
    plan.add(SetHeader("Orders", 1, 3, "_Customers_name"))
    plan.add(WriteDataBlock("Orders", 2, 1, (("O-1", "C-1", formula),)))
    plan.add(DefineSheet("Customers", 1))
    plan.add(SetHeader("Customers", 1, 1, "id"))
    plan.add(SetHeader("Customers", 1, 2, "name"))
    plan.add(WriteDataBlock("Customers", 2, 1, (("C-1", "Ada"), ("C-2", "Bob"))))

    out = tmp_path / "lookup_formula.xlsx"
    render_workbook(plan, out)

    wb = load_workbook(out, data_only=False)
    try:
        assert wb["Orders"]["C2"].value == (
            "=XLOOKUP($B2,'Customers'!$A$2:$A$3,'Customers'!$B$2:$B$3,\"\")"
        )
    finally:
        wb.close()


@pytest.mark.ftr("FTR-FORMULA-FK-HELPER-PROVIDERS-P4A")
def test_xlsx_renderer_lookup_range_spans_multiple_data_blocks(tmp_path: Path) -> None:
    formula = lookup_formula(
        source_key_column="id_(Customers)",
        lookup_sheet="Customers",
        lookup_key_column="id",
        lookup_value_column="name",
    )
    plan = RenderPlan()
    plan.add(DefineSheet("Orders", 0))
    plan.add(SetHeader("Orders", 1, 1, "id"))
    plan.add(SetHeader("Orders", 1, 2, "id_(Customers)"))
    plan.add(SetHeader("Orders", 1, 3, "_Customers_name"))
    plan.add(WriteDataBlock("Orders", 2, 1, (("O-1", "C-3", formula),)))
    plan.add(DefineSheet("Customers", 1))
    plan.add(SetHeader("Customers", 1, 1, "id"))
    plan.add(SetHeader("Customers", 1, 2, "name"))
    plan.add(WriteDataBlock("Customers", 2, 1, (("C-1", "Ada"), ("C-2", "Bob"))))
    plan.add(WriteDataBlock("Customers", 4, 1, (("C-3", "Cy"), ("C-4", "Dee"))))

    out = tmp_path / "lookup_formula_multiblock.xlsx"
    render_workbook(plan, out)

    wb = load_workbook(out, data_only=False)
    try:
        assert wb["Orders"]["C2"].value == (
            "=XLOOKUP($B2,'Customers'!$A$2:$A$5,'Customers'!$B$2:$B$5,\"\")"
        )
    finally:
        wb.close()
