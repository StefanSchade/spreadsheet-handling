"""
Tests for FTR-IR-DATA-CELLS — data cell emission via WriteDataBlock.

Verifies that the IR render plan emits WriteDataBlock ops and the
openpyxl renderer writes cell data without a pandas pre-write step.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
from spreadsheet_handling.rendering.flow import build_render_plan, default_p1_passes, apply_ir_passes
from spreadsheet_handling.rendering.plan import WriteDataBlock, DefineSheet, SetHeader
from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook

pytestmark = pytest.mark.ftr("FTR-IR-DATA-CELLS")


@pytest.fixture
def simple_frames():
    return {"Products": pd.DataFrame({"id": ["P1", "P2"], "name": ["Widget", "Gadget"]})}


@pytest.fixture
def multi_frames():
    return {
        "Products": pd.DataFrame({"id": ["P1", "P2"], "name": ["Widget", "Gadget"]}),
        "Orders": pd.DataFrame({"order_id": [1, 2, 3], "amount": [10.0, 20.0, 30.0]}),
    }


# ------------------------------------------------------------------
# Plan emission tests
# ------------------------------------------------------------------

def test_plan_contains_write_data_block(simple_frames):
    ir = compose_workbook(simple_frames, {})
    for p in default_p1_passes():
        ir = p.apply(ir)
    plan = build_render_plan(ir)

    data_ops = [op for op in plan.ops if isinstance(op, WriteDataBlock)]
    assert len(data_ops) == 1
    op = data_ops[0]
    assert op.sheet == "Products"
    assert op.r1 == 2  # data starts after 1 header row
    assert op.c1 == 1
    assert len(op.data) == 2  # 2 data rows
    assert op.data[0] == ("P1", "Widget")
    assert op.data[1] == ("P2", "Gadget")


def test_plan_data_block_multi_sheet(multi_frames):
    ir = compose_workbook(multi_frames, {})
    for p in default_p1_passes():
        ir = p.apply(ir)
    plan = build_render_plan(ir)

    data_ops = [op for op in plan.ops if isinstance(op, WriteDataBlock)]
    sheets = {op.sheet for op in data_ops}
    assert sheets == {"Products", "Orders"}


def test_plan_data_block_with_multiindex():
    cols = pd.MultiIndex.from_tuples([("order", "id"), ("order", "name"), ("customer", "city")])
    df = pd.DataFrame([["O1", "First", "Berlin"], ["O2", "Second", "Munich"]], columns=cols)
    ir = compose_workbook({"Sheet1": df}, {})
    for p in default_p1_passes():
        ir = p.apply(ir)
    plan = build_render_plan(ir)

    data_ops = [op for op in plan.ops if isinstance(op, WriteDataBlock)]
    assert len(data_ops) == 1
    op = data_ops[0]
    # MultiIndex has 2 header rows, so data starts at row 3
    assert op.r1 == 3
    assert len(op.data) == 2


def test_plan_empty_dataframe():
    ir = compose_workbook({"Empty": pd.DataFrame({"a": pd.Series([], dtype="object")})}, {})
    for p in default_p1_passes():
        ir = p.apply(ir)
    plan = build_render_plan(ir)

    data_ops = [op for op in plan.ops if isinstance(op, WriteDataBlock)]
    assert len(data_ops) == 1
    assert len(data_ops[0].data) == 0


# ------------------------------------------------------------------
# Roundtrip tests (write XLSX via IR, read back with openpyxl)
# ------------------------------------------------------------------

def test_ir_write_data_cells_roundtrip(simple_frames, tmp_path):
    ir = compose_workbook(simple_frames, {})
    for p in default_p1_passes():
        ir = p.apply(ir)
    plan = build_render_plan(ir)

    out = tmp_path / "out.xlsx"
    render_workbook(plan, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Products"]
    # Header row
    assert ws.cell(1, 1).value == "id"
    assert ws.cell(1, 2).value == "name"
    # Data rows
    assert ws.cell(2, 1).value == "P1"
    assert ws.cell(2, 2).value == "Widget"
    assert ws.cell(3, 1).value == "P2"
    assert ws.cell(3, 2).value == "Gadget"


def test_ir_write_numeric_data_roundtrip(tmp_path):
    frames = {"Numbers": pd.DataFrame({"x": [1, 2, 3], "y": [4.5, 5.5, 6.5]})}
    ir = compose_workbook(frames, {})
    for p in default_p1_passes():
        ir = p.apply(ir)
    plan = build_render_plan(ir)

    out = tmp_path / "out.xlsx"
    render_workbook(plan, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Numbers"]
    assert ws.cell(2, 1).value == 1
    assert ws.cell(4, 2).value == 6.5


def test_ir_backend_write_multi_produces_data(tmp_path):
    """Integration: ExcelBackend.write_multi with IR path writes data cells."""
    import os
    os.environ["SH_XLSX_BACKEND"] = "ir"
    try:
        from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend
        frames = {"Items": pd.DataFrame({"code": ["A", "B"], "qty": [10, 20]})}
        out = tmp_path / "backend.xlsx"
        ExcelBackend().write_multi(frames, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Items"]
        assert ws.cell(1, 1).value == "code"
        assert ws.cell(2, 1).value == "A"
        assert ws.cell(3, 2).value == 20
    finally:
        os.environ.pop("SH_XLSX_BACKEND", None)
