from __future__ import annotations

import ast
from pathlib import Path

import pytest
from openpyxl import Workbook

import spreadsheet_handling.io_backends.xlsx.openpyxl_parser as xp
from spreadsheet_handling.io_backends.xlsx.parser_interpretation import (
    build_sheet_meta_hints,
    build_visible_sheet_ir,
)
from spreadsheet_handling.rendering.ir import DataValidationSpec


pytestmark = pytest.mark.ftr("FTR-XLSX-PARSER-MODULARIZATION-P3I")


def test_build_sheet_meta_hints_merges_workbook_defaults_and_sheet_overrides():
    workbook_meta = {
        "freeze_header": True,
        "auto_filter": True,
        "helper_prefix": "_",
        "sheets": {
            "Products": {
                "freeze_header": False,
                "header_fill_rgb": "#CCE5FF",
            }
        },
    }

    hints = build_sheet_meta_hints(workbook_meta, sheet_name="Products")

    assert hints == {
        "freeze_header": False,
        "auto_filter": True,
        "helper_prefix": "_",
        "header_fill_rgb": "#CCE5FF",
    }


def test_build_visible_sheet_ir_combines_interpretation_inputs():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="title")
    ws.cell(row=2, column=1, value="P-001")
    ws.cell(row=2, column=2, value="Alpha")

    validations = [
        DataValidationSpec(
            kind="list",
            area=(2, 2, 2, 2),
            formula='"new,done"',
            allow_empty=True,
        )
    ]

    sh = build_visible_sheet_ir(
        ws,
        sheet_name="Products",
        meta_hints={"freeze_header": True, "auto_filter": True},
        validations=validations,
        freeze_hint={"row": 2, "col": 1},
        autofilter_ref="A1:B2",
        anchors=None,
        stop_on_empty_row=True,
        stop_on_empty_col=False,
    )

    assert sh.meta["options"] == {"freeze_header": True, "auto_filter": True}
    assert sh.meta["__freeze"] == {"row": 2, "col": 1}
    assert sh.meta["__autofilter_ref"] == "A1:B2"
    assert sh.validations == validations
    assert sh.tables[0].headers == ["id", "title"]
    assert sh.tables[0].data == [["P-001", "Alpha"]]


def test_openpyxl_parser_delegates_visible_sheet_interpretation_to_helper_module():
    module_path = Path(xp.__file__).resolve()
    tree = ast.parse(module_path.read_text(encoding="utf-8"), filename=str(module_path))

    imported_modules: list[str] = []
    defined_functions: list[str] = []
    called_functions: list[str] = []

    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.extend(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom):
            imported_modules.append(node.module or "")
        elif isinstance(node, ast.FunctionDef):
            defined_functions.append(node.name)
        elif isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            called_functions.append(node.func.id)

    assert "spreadsheet_handling.io_backends.xlsx.parser_interpretation" in imported_modules
    assert "_parse_visible_sheet" in defined_functions
    assert "build_visible_sheet_ir" in called_functions
