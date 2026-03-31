from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend
import pytest

pytestmark = pytest.mark.ftr("FTR-MULTIHEADER-P2")



def test_ir_backend_writes_multirow_headers_with_merges(tmp_path, monkeypatch):

    cols = pd.MultiIndex.from_tuples([
        ("order", "id"),
        ("order", "date"),
        ("customer", "name"),
    ])
    frames = {
        "Orders": pd.DataFrame(
            [["O1", "2026-01-01", "Alice"], ["O2", "2026-01-02", "Bob"]],
            columns=cols,
        )
    }
    out = tmp_path / "multiheader.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    ws = wb["Orders"]

    # Two header rows from MultiIndex
    assert ws.cell(row=1, column=1).value == "order"
    assert ws.cell(row=2, column=1).value == "id"
    assert ws.cell(row=2, column=2).value == "date"
    assert ws.cell(row=1, column=3).value == "customer"
    assert ws.cell(row=2, column=3).value == "name"

    # Top-level repeated label merged across A1:B1
    merged = {str(rng) for rng in ws.merged_cells.ranges}
    assert "A1:B1" in merged

    # Data starts below both header rows
    assert ws.cell(row=3, column=1).value == "O1"
    assert ws.cell(row=3, column=2).value == "2026-01-01"
    assert ws.cell(row=4, column=1).value == "O2"
