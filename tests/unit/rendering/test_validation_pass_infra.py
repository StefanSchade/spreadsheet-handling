"""Tests for FTR-VALIDATION-PASS-INFRA: neutral constraint propagation through IR."""
from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from spreadsheet_handling.domain.validations.validate_columns import add_validations
from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
from spreadsheet_handling.rendering.passes.core import ValidationPass
from spreadsheet_handling.rendering.passes import apply_all
from spreadsheet_handling.rendering.flow import build_render_plan
from spreadsheet_handling.rendering.formulas import formula_list_values
from spreadsheet_handling.rendering.ir import SheetIR, WorkbookIR

pytestmark = pytest.mark.ftr("FTR-VALIDATION-PASS-INFRA")



# ---------------------------------------------------------------------------
# Domain step: add_validations writes constraints into meta
# ---------------------------------------------------------------------------

class TestAddValidationsWritesMeta:
    def test_constraints_stored_in_meta_dict_key(self):
        frames = {"Kunden": pd.DataFrame({"status": ["A", "B"]})}
        frames["_meta"] = {}
        add_validations(frames, rules=[{
            "sheet": "Kunden",
            "column": "status",
            "rule": {"type": "in_list", "values": ["A", "B", "C"]},
        }])
        constraints = frames["_meta"]["constraints"]
        assert len(constraints) == 1
        assert constraints[0]["sheet"] == "Kunden"
        assert constraints[0]["rule"]["values"] == ["A", "B", "C"]

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_constraints_stored_in_meta_dict_key(self):
        frames = {"Kunden": pd.DataFrame({"status": ["E", "R"]})}
        frames["_meta"] = {}
        add_validations(frames, rules=[{
            "sheet": "Kunden",
            "column": "status",
            "rule": {
                "type": "from_legend",
                "legend": "status_codes",
                "include_empty": True,
            },
        }])

        constraints = frames["_meta"]["constraints"]
        assert constraints == [
            {
                "sheet": "Kunden",
                "column": "status",
                "rule": {
                    "type": "from_legend",
                    "legend": "status_codes",
                    "include_empty": True,
                },
                "on_violation": "error",
            }
        ]

    def test_multiple_constraints_accumulate(self):
        frames = {"Kunden": pd.DataFrame({"status": ["A"], "typ": ["X"]})}
        frames["_meta"] = {}
        add_validations(frames, rules=[
            {"sheet": "Kunden", "column": "status", "rule": {"type": "in_list", "values": ["A", "B"]}},
            {"sheet": "Kunden", "column": "typ", "rule": {"type": "in_list", "values": ["X", "Y"]}},
        ])
        assert len(frames["_meta"]["constraints"]) == 2


# ---------------------------------------------------------------------------
# ValidationPass: workbook-level constraints → SheetIR.validations
# ---------------------------------------------------------------------------

class TestValidationPassFromConstraints:
    def _make_ir_with_constraints(self, constraints, extra_meta=None):
        """Build a WorkbookIR with a Products sheet and stash constraints in _meta."""
        frames = {"Products": pd.DataFrame({"id": [1, 2], "category": ["A", "B"]})}
        meta = {"constraints": constraints, **dict(extra_meta or {})}
        ir = compose_workbook(frames, meta)
        return ir

    def test_in_list_constraint_creates_validation_spec(self):
        ir = self._make_ir_with_constraints([{
            "sheet": "Products",
            "column": "category",
            "rule": {"type": "in_list", "values": ["A", "B", "C"]},
        }])
        vp = ValidationPass()
        ir = vp.apply(ir)
        assert len(ir.sheets["Products"].validations) == 1
        dv = ir.sheets["Products"].validations[0]
        assert dv.kind == "list"
        assert formula_list_values(dv.formula) == ("A", "B", "C")

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_constraint_creates_validation_spec(self):
        ir = self._make_ir_with_constraints(
            [{
                "sheet": "Products",
                "column": "category",
                "rule": {"type": "from_legend", "legend": "status_codes"},
            }],
            extra_meta={
                "legend_blocks": {
                    "status_codes": {
                        "entries": [
                            {"token": "E", "label": "Editable"},
                            {"token": "R", "label": "Read-only"},
                            {"token": "K", "label": "Composite"},
                        ],
                    }
                }
            },
        )

        ir = ValidationPass().apply(ir)

        dv = ir.sheets["Products"].validations[0]
        assert dv.kind == "list"
        assert formula_list_values(dv.formula) == ("E", "R", "K")

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_supports_list_style_legend_blocks(self):
        ir = self._make_ir_with_constraints(
            [{
                "sheet": "Products",
                "column": "category",
                "rule": {"type": "from_legend", "legend": "status_codes"},
            }],
            extra_meta={
                "legend_blocks": [
                    {
                        "name": "status_codes",
                        "entries": [
                            {"token": "E", "label": "Editable"},
                            {"token": "R", "label": "Read-only"},
                        ],
                    }
                ]
            },
        )

        ir = ValidationPass().apply(ir)

        dv = ir.sheets["Products"].validations[0]
        assert dv.kind == "list"
        assert formula_list_values(dv.formula) == ("E", "R")

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_include_empty_adds_blank_choice(self):
        ir = self._make_ir_with_constraints(
            [{
                "sheet": "Products",
                "column": "category",
                "rule": {
                    "type": "from_legend",
                    "legend": "status_codes",
                    "include_empty": True,
                },
            }],
            extra_meta={
                "legend_blocks": {
                    "status_codes": {
                        "entries": [
                            {"token": "E", "label": "Editable"},
                            {"token": "R", "label": "Read-only"},
                        ],
                    }
                }
            },
        )

        ir = ValidationPass().apply(ir)

        assert formula_list_values(ir.sheets["Products"].validations[0].formula) == ("", "E", "R")

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_column_wildcard_targets_all_data_columns(self):
        ir = self._make_ir_with_constraints(
            [{
                "sheet": "Products",
                "column": "*",
                "rule": {"type": "from_legend", "legend": "status_codes"},
            }],
            extra_meta={
                "legend_blocks": {
                    "status_codes": {
                        "entries": [
                            {"token": "E", "label": "Editable"},
                            {"token": "R", "label": "Read-only"},
                        ],
                    }
                }
            },
        )

        ir = ValidationPass().apply(ir)

        assert [dv.area for dv in ir.sheets["Products"].validations] == [
            (2, 1, 3, 1),
            (2, 2, 3, 2),
        ]

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_warns_for_large_dropdowns(self, caplog):
        entries = [
            {"token": f"T{index}", "label": f"Token {index}"}
            for index in range(51)
        ]
        ir = self._make_ir_with_constraints(
            [{
                "sheet": "Products",
                "column": "category",
                "rule": {"type": "from_legend", "legend": "status_codes"},
            }],
            extra_meta={"legend_blocks": {"status_codes": {"entries": entries}}},
        )

        with caplog.at_level("WARNING", logger="sheets.validation"):
            ValidationPass().apply(ir)

        assert "Excel dropdown UX may degrade" in caplog.text

    def test_unknown_sheet_silently_skipped(self):
        ir = self._make_ir_with_constraints([{
            "sheet": "NonExistent",
            "column": "x",
            "rule": {"type": "in_list", "values": ["a"]},
        }])
        vp = ValidationPass()
        ir = vp.apply(ir)
        assert len(ir.sheets["Products"].validations) == 0

    def test_unknown_column_silently_skipped(self):
        ir = self._make_ir_with_constraints([{
            "sheet": "Products",
            "column": "nonexistent_col",
            "rule": {"type": "in_list", "values": ["a"]},
        }])
        vp = ValidationPass()
        ir = vp.apply(ir)
        assert len(ir.sheets["Products"].validations) == 0

    def test_unsupported_rule_type_skipped(self):
        ir = self._make_ir_with_constraints([{
            "sheet": "Products",
            "column": "category",
            "rule": {"type": "whole_number", "min": 0},
        }])
        vp = ValidationPass()
        ir = vp.apply(ir)
        assert len(ir.sheets["Products"].validations) == 0

    def test_validation_area_covers_data_rows(self):
        ir = self._make_ir_with_constraints([{
            "sheet": "Products",
            "column": "category",
            "rule": {"type": "in_list", "values": ["X"]},
        }])
        vp = ValidationPass()
        ir = vp.apply(ir)
        dv = ir.sheets["Products"].validations[0]
        r1, c1, r2, c2 = dv.area
        assert r1 == 2  # data starts at row 2 (after header)
        assert r2 == 3  # 2 data rows + 1 header = n_rows=3, data rows 2..3


# ---------------------------------------------------------------------------
# Legacy path: _p1_validations still works
# ---------------------------------------------------------------------------

class TestValidationPassLegacy:
    def test_p1_validations_still_processed(self):
        ir = WorkbookIR()
        sh = SheetIR(name="Sheet1", meta={
            "_p1_validations": [
                {"kind": "list", "col": 2, "from_row": 2, "to_row": 10, "values": ["X", "Y"]},
            ]
        })
        ir.sheets["Sheet1"] = sh
        vp = ValidationPass()
        ir = vp.apply(ir)
        assert len(ir.sheets["Sheet1"].validations) == 1
        assert ir.sheets["Sheet1"].validations[0].kind == "list"


# ---------------------------------------------------------------------------
# apply_all: unified path (production)
# ---------------------------------------------------------------------------

class TestApplyAll:
    def test_constraints_via_apply_all(self):
        frames = {"Products": pd.DataFrame({"id": [1], "cat": ["A"]})}
        meta = {"constraints": [{
            "sheet": "Products",
            "column": "cat",
            "rule": {"type": "in_list", "values": ["A", "B"]},
        }]}
        ir = compose_workbook(frames, meta)
        ir = apply_all(ir, meta)
        assert len(ir.sheets["Products"].validations) >= 1


# ---------------------------------------------------------------------------
# End-to-end: domain → compose → passes → render plan → XLSX ops
# ---------------------------------------------------------------------------

class TestEndToEndConstraintPropagation:
    def test_domain_constraints_reach_render_plan(self):
        frames = {"Orders": pd.DataFrame({"id": [1], "status": ["open"]})}
        frames["_meta"] = {}
        add_validations(frames, rules=[{
            "sheet": "Orders",
            "column": "status",
            "rule": {"type": "in_list", "values": ["open", "closed", "pending"]},
        }])
        meta = frames["_meta"]
        ir = compose_workbook(frames, meta)
        ir = apply_all(ir, meta)
        plan = build_render_plan(ir)

        # The plan should contain an AddValidation op for Orders
        val_ops = [op for op in plan.ops
                   if type(op).__name__ == "AddValidation" and getattr(op, "sheet", None) == "Orders"]
        assert len(val_ops) >= 1
        assert val_ops[0].kind == "list"
        assert "open" in formula_list_values(val_ops[0].formula)


class TestEndToEndXlsx:
    def test_constraints_appear_in_xlsx(self, tmp_path):
        from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook
        from tests.utils.xlsx_normalize import normalize_xlsx

        frames = {"Data": pd.DataFrame({"id": [1, 2], "color": ["red", "blue"]})}
        frames["_meta"] = {}
        add_validations(frames, rules=[{
            "sheet": "Data",
            "column": "color",
            "rule": {"type": "in_list", "values": ["red", "green", "blue"]},
        }])
        meta = frames["_meta"]
        ir = compose_workbook(frames, meta)
        ir = apply_all(ir, meta)
        plan = build_render_plan(ir)

        out = tmp_path / "test.xlsx"
        render_workbook(plan, str(out))

        shape = normalize_xlsx(str(out))
        assert "Data" in shape["validations"]
        dvs = shape["validations"]["Data"]
        assert any(dv["type"] == "list" for dv in dvs)

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_constraint_appears_in_xlsx(self, tmp_path):
        from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook
        from tests.utils.xlsx_normalize import normalize_xlsx

        frames = {"Data": pd.DataFrame({"id": [1, 2], "code": ["E", "R"]})}
        meta = {
            "legend_blocks": {
                "status_codes": {
                    "entries": [
                        {"token": "E", "label": "Editable"},
                        {"token": "R", "label": "Read-only"},
                        {"token": "K", "label": "Composite"},
                    ],
                }
            },
            "constraints": [
                {
                    "sheet": "Data",
                    "column": "code",
                    "rule": {"type": "from_legend", "legend": "status_codes"},
                }
            ],
        }
        ir = compose_workbook(frames, meta)
        ir = apply_all(ir, meta)
        plan = build_render_plan(ir)

        out = tmp_path / "legend_validation.xlsx"
        render_workbook(plan, str(out))

        shape = normalize_xlsx(str(out))
        dvs = shape["validations"]["Data"]
        assert any(
            dv["type"] == "list"
            and dv["sqref"] == "B2:B3"
            and dv["formula1"] == '"E,R,K"'
            for dv in dvs
        )

        workbook = load_workbook(out, data_only=True)
        try:
            ws = workbook["Data"]
            assert [ws["B2"].value, ws["B3"].value] == ["E", "R"]
        finally:
            workbook.close()

    @pytest.mark.ftr("FTR-LEGEND-VALIDATION-LISTS-P4A")
    def test_from_legend_xlsx_reimport_preserves_data_cells(self, tmp_path):
        from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend

        meta = {
            "legend_blocks": {
                "status_codes": {
                    "entries": [
                        {"token": "E", "label": "Editable"},
                        {"token": "R", "label": "Read-only"},
                        {"token": "K", "label": "Composite"},
                    ],
                }
            },
            "constraints": [
                {
                    "sheet": "Data",
                    "column": "code",
                    "rule": {"type": "from_legend", "legend": "status_codes"},
                }
            ],
        }
        frames = {
            "Data": pd.DataFrame({"id": [1, 2], "code": ["E", "R"]}),
            "_meta": meta,
        }
        out = tmp_path / "legend_validation_reimport.xlsx"

        ExcelBackend().write_multi(frames, str(out))
        back = ExcelBackend().read_multi(str(out), header_levels=1)

        assert back["Data"].to_dict(orient="records") == [
            {"id": "1", "code": "E"},
            {"id": "2", "code": "R"},
        ]
