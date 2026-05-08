"""Tests for FTR-EDITABLE-COLUMNS-AND-PROTECTION-P4A."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from spreadsheet_handling.rendering.ir import WorkbookIR, SheetIR, TableBlock
from spreadsheet_handling.rendering.passes.core import ProtectionPass
from spreadsheet_handling.rendering.flow import (
    apply_ir_passes,
    build_render_plan,
    default_p1_passes,
)
from spreadsheet_handling.rendering.plan import ApplyCellLock, SetSheetProtection

pytestmark = pytest.mark.ftr("FTR-EDITABLE-COLUMNS-AND-PROTECTION-P4A")


def _make_ir(
    headers: list[str],
    n_data_rows: int = 3,
    options: dict | None = None,
) -> WorkbookIR:
    header_map = {h: i + 1 for i, h in enumerate(headers)}
    tbl = TableBlock(
        frame_name="matrix",
        top=1,
        left=1,
        header_rows=1,
        header_cols=1,
        n_rows=n_data_rows + 1,
        n_cols=len(headers),
        headers=headers,
        header_map=header_map,
    )
    sh = SheetIR(name="matrix", tables=[tbl])
    if options:
        sh.meta["options"] = options
    wb = WorkbookIR(sheets={"matrix": sh})
    return wb


class TestProtectionPass:

    def test_explicit_editable_columns_marks_protection_metadata(self) -> None:
        ir = _make_ir(
            ["variable_id", "P-001", "P-002", "label"],
            options={
                "protection": {
                    "editable_columns": ["P-001", "P-002"],
                }
            },
        )
        ProtectionPass().apply(ir)

        prot = ir.sheets["matrix"].meta["__protection"]
        assert sorted(prot["unlocked_cols"]) == [2, 3]
        assert sorted(prot["locked_cols"]) == [1, 4]
        assert prot["password"] is None

    def test_non_helper_editable_mode_uses_helper_metadata(self) -> None:
        ir = _make_ir(
            ["variable_id", "label", "P-001", "P-002"],
            options={
                "helper_columns": ["label"],
                "protection": {"editable": "non_helper"},
            },
        )
        ProtectionPass().apply(ir)

        prot = ir.sheets["matrix"].meta["__protection"]
        # label is helper → locked; variable_id, P-001, P-002 are editable
        assert 2 in prot["locked_cols"]  # label is col 2
        assert sorted(prot["unlocked_cols"]) == [1, 3, 4]

    def test_no_protection_without_option(self) -> None:
        ir = _make_ir(["ID", "value"])
        ProtectionPass().apply(ir)

        assert "__protection" not in ir.sheets["matrix"].meta

    def test_protection_with_password(self) -> None:
        ir = _make_ir(
            ["ID", "value"],
            options={
                "protection": {
                    "editable_columns": ["value"],
                    "password": "test123",
                }
            },
        )
        ProtectionPass().apply(ir)

        prot = ir.sheets["matrix"].meta["__protection"]
        assert prot["password"] == "test123"


class TestProtectionRenderPlan:

    def test_build_render_plan_emits_protection_ops(self) -> None:
        ir = _make_ir(
            ["ID", "editable_value", "helper_col"],
            n_data_rows=5,
            options={
                "protection": {
                    "editable_columns": ["editable_value"],
                }
            },
        )
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        lock_ops = [op for op in plan.ops if isinstance(op, ApplyCellLock)]
        prot_ops = [op for op in plan.ops if isinstance(op, SetSheetProtection)]

        assert len(prot_ops) == 1
        assert prot_ops[0].sheet == "matrix"
        assert prot_ops[0].password is None

        # editable_value (col 2) should be unlocked
        unlocked = [op for op in lock_ops if op.locked is False]
        assert len(unlocked) == 1
        assert unlocked[0].col == 2
        assert unlocked[0].from_row == 1  # includes header
        assert unlocked[0].to_row == 6  # 5 data rows + 1 header


class TestXlsxProtectionOutput:

    def test_xlsx_protection_locks_helper_unlocks_editable(self, tmp_path: Path) -> None:
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook
        from openpyxl import load_workbook

        frames = {
            "matrix": pd.DataFrame(
                [
                    {"variable_id": "v1", "P-001": "output", "label": "Rate"},
                    {"variable_id": "v2", "P-001": "input", "label": "Amount"},
                ]
            )
        }
        meta = {
            "sheets": {
                "matrix": {
                    "protection": {
                        "editable_columns": ["P-001"],
                    }
                }
            }
        }
        ir = compose_workbook(frames, meta)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "protected.xlsx"
        render_workbook(plan, out)

        wb = load_workbook(out)
        ws = wb.active
        # Sheet protection should be enabled
        assert ws.protection.sheet is True
        # P-001 is column 2 → should be unlocked
        editable_cell = ws.cell(row=2, column=2)
        assert editable_cell.protection.locked is False
        # variable_id is column 1 → should remain locked (default)
        locked_cell = ws.cell(row=2, column=1)
        # In openpyxl, default locked is True when sheet protection is on
        assert locked_cell.protection.locked in (True, None)

    def test_reimport_reads_data_correctly_from_protected_workbook(self, tmp_path: Path) -> None:
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import render_workbook
        from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend

        frames = {
            "matrix": pd.DataFrame(
                [
                    {"variable_id": "v1", "P-001": "output"},
                ]
            )
        }
        meta = {
            "sheets": {
                "matrix": {
                    "protection": {
                        "editable_columns": ["P-001"],
                    }
                }
            }
        }
        ir = compose_workbook(frames, meta)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "reimport-protected.xlsx"
        render_workbook(plan, out)

        # Reimport should read the data correctly
        backend = ExcelBackend()
        reimported = backend.read_multi(str(out), header_levels=1)
        assert "matrix" in reimported
        assert reimported["matrix"].to_dict(orient="records") == [
            {"variable_id": "v1", "P-001": "output"},
        ]


class TestConfigureWorkbookViewProtection:

    def test_editable_columns_flows_into_sheet_protection_options(self) -> None:
        from spreadsheet_handling.domain.workbook_views import configure_workbook_view

        frames = {
            "matrix_view": pd.DataFrame(
                [{"variable_id": "v1", "P-001": "output", "label": "Rate"}]
            )
        }

        out = configure_workbook_view(
            frames,
            sheets=[
                {
                    "frame": "matrix_view",
                    "sheet": "Matrix",
                    "editable_columns": ["P-001"],
                    "helper_columns": ["label"],
                }
            ],
        )

        sheet_opts = out["_meta"]["sheets"]["Matrix"]
        assert sheet_opts["protection"] == {"editable_columns": ["P-001"]}
        assert sheet_opts["helper_columns"] == ["label"]

    def test_protection_mapping_flows_directly(self) -> None:
        from spreadsheet_handling.domain.workbook_views import configure_workbook_view

        frames = {
            "matrix_view": pd.DataFrame(
                [{"variable_id": "v1", "P-001": "output"}]
            )
        }

        out = configure_workbook_view(
            frames,
            sheets=[
                {
                    "frame": "matrix_view",
                    "sheet": "Matrix",
                    "protection": {"editable": "non_helper", "password": "x"},
                }
            ],
        )

        sheet_opts = out["_meta"]["sheets"]["Matrix"]
        assert sheet_opts["protection"] == {"editable": "non_helper", "password": "x"}
