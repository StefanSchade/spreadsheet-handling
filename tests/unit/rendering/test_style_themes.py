"""Tests for FTR-STYLE-THEMES: configurable styling via IR passes.

Covers:
- StylePass produces correct __style and __helper_cols annotations
- build_render_plan emits ApplyHeaderStyle and ApplyColumnStyle ops
- Full IR path renders XLSX with configured header fill, autofilter, freeze, helper highlight
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from spreadsheet_handling.rendering.ir import WorkbookIR, SheetIR, TableBlock
from spreadsheet_handling.rendering.passes.core import (
    StylePass,
    FilterPass,
    FreezePass,
)
from spreadsheet_handling.rendering.flow import (
    apply_ir_passes,
    build_render_plan,
    default_p1_passes,
)
from spreadsheet_handling.rendering.plan import (
    ApplyHeaderStyle,
    ApplyColumnStyle,
    SetAutoFilter,
    SetFreeze,
)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _make_ir(
    headers: list[str],
    n_data_rows: int = 3,
    options: dict | None = None,
) -> WorkbookIR:
    """Build a minimal WorkbookIR with one sheet and one table."""
    header_map = {h: i + 1 for i, h in enumerate(headers)}
    tbl = TableBlock(
        frame_name="data",
        top=1,
        left=1,
        header_rows=1,
        header_cols=1,
        n_rows=n_data_rows + 1,  # +1 for header
        n_cols=len(headers),
        headers=headers,
        header_map=header_map,
    )
    sh = SheetIR(name="data", tables=[tbl])
    if options:
        sh.meta["options"] = options
    wb = WorkbookIR(sheets={"data": sh})
    return wb


# ---------------------------------------------------------------------------
# StylePass unit tests
# ---------------------------------------------------------------------------

class TestStylePass:

    def test_default_header_fill(self) -> None:
        ir = _make_ir(["id", "name"])
        StylePass().apply(ir)

        style = ir.sheets["data"].meta["__style"]
        assert style["header"]["bold"] is True
        assert style["header"]["fill"] == "#F2F2F2"

    def test_custom_header_fill_via_options(self) -> None:
        ir = _make_ir(["id", "name"], options={"header_fill_rgb": "#AABBCC"})
        StylePass().apply(ir)

        style = ir.sheets["data"].meta["__style"]
        assert style["header"]["fill"] == "#AABBCC"

    def test_custom_header_fill_via_constructor(self) -> None:
        ir = _make_ir(["id", "name"])
        StylePass(default_header_fill_rgb="#112233").apply(ir)

        assert ir.sheets["data"].meta["__style"]["header"]["fill"] == "#112233"

    def test_helper_cols_detected(self) -> None:
        ir = _make_ir(["id", "name", "_helper_name"])
        StylePass().apply(ir)

        hc = ir.sheets["data"].meta["__helper_cols"]
        assert hc["cols"] == [3]  # _helper_name is column 3
        assert hc["fill"] == "#E8F0FE"  # default

    def test_helper_fill_from_options(self) -> None:
        ir = _make_ir(["id", "_h1", "_h2"], options={"helper_fill_rgb": "#FFCC00"})
        StylePass().apply(ir)

        hc = ir.sheets["data"].meta["__helper_cols"]
        assert set(hc["cols"]) == {2, 3}
        assert hc["fill"] == "#FFCC00"

    def test_no_helper_cols_when_none_present(self) -> None:
        ir = _make_ir(["id", "name", "value"])
        StylePass().apply(ir)

        assert "__helper_cols" not in ir.sheets["data"].meta

    def test_helper_fill_disabled(self) -> None:
        ir = _make_ir(["id", "_helper"], options={"helper_fill_rgb": None})
        StylePass(default_helper_fill_rgb=None).apply(ir)

        assert "__helper_cols" not in ir.sheets["data"].meta


# ---------------------------------------------------------------------------
# FilterPass and FreezePass
# ---------------------------------------------------------------------------

class TestFilterAndFreezePass:

    def test_autofilter_on_by_default(self) -> None:
        ir = _make_ir(["id", "name"])
        FilterPass().apply(ir)

        af = ir.sheets["data"].meta["__autofilter"]
        assert af["top_left"] == (1, 1)
        assert af["bottom_right"] == (4, 2)  # 3 data rows + 1 header = 4 total

    def test_autofilter_disabled(self) -> None:
        ir = _make_ir(["id"], options={"auto_filter": False})
        FilterPass().apply(ir)

        assert "__autofilter" not in ir.sheets["data"].meta

    def test_freeze_header_off_by_default(self) -> None:
        ir = _make_ir(["id"])
        FreezePass().apply(ir)

        assert "__freeze" not in ir.sheets["data"].meta

    def test_freeze_header_on(self) -> None:
        ir = _make_ir(["id"], options={"freeze_header": True})
        FreezePass().apply(ir)

        fz = ir.sheets["data"].meta["__freeze"]
        assert fz == {"row": 2, "col": 1}


# ---------------------------------------------------------------------------
# RenderPlan generation
# ---------------------------------------------------------------------------

class TestBuildRenderPlan:

    def test_header_style_ops_emitted(self) -> None:
        ir = _make_ir(["id", "name"])
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        header_ops = [op for op in plan.ops if isinstance(op, ApplyHeaderStyle)]
        assert len(header_ops) == 2
        assert all(op.bold for op in header_ops)
        assert all(op.fill_rgb == "#F2F2F2" for op in header_ops)

    def test_helper_column_ops_emitted(self) -> None:
        ir = _make_ir(["id", "_helper_name"], n_data_rows=5)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        col_ops = [op for op in plan.ops if isinstance(op, ApplyColumnStyle)]
        assert len(col_ops) == 1
        assert col_ops[0].col == 2         # _helper_name is column 2
        assert col_ops[0].from_row == 2    # data starts at row 2
        assert col_ops[0].to_row == 6      # 5 data rows → row 2..6
        assert col_ops[0].fill_rgb == "#E8F0FE"

    def test_autofilter_op_emitted(self) -> None:
        ir = _make_ir(["a", "b"])
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        af_ops = [op for op in plan.ops if isinstance(op, SetAutoFilter)]
        assert len(af_ops) == 1

    def test_freeze_op_emitted_when_enabled(self) -> None:
        ir = _make_ir(["a"], options={"freeze_header": True})
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        fz_ops = [op for op in plan.ops if isinstance(op, SetFreeze)]
        assert len(fz_ops) == 1
        assert fz_ops[0].row == 2

    def test_no_helper_ops_when_no_helper_cols(self) -> None:
        ir = _make_ir(["id", "name", "value"])
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        col_ops = [op for op in plan.ops if isinstance(op, ApplyColumnStyle)]
        assert len(col_ops) == 0


# ---------------------------------------------------------------------------
# XLSX output verification (end-to-end through IR path)
# ---------------------------------------------------------------------------

class TestXlsxStyleOutput:

    @pytest.mark.xlsx_ir
    def test_header_bold_and_fill_in_xlsx(self, tmp_path: Path) -> None:
        """Verify that header cells in the output workbook have bold font and fill."""
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.io_backends.xlsx.xlsx_openpyxl import render_plan
        from openpyxl import load_workbook

        frames = {"products": pd.DataFrame([{"id": "a", "name": "Alpha"}])}
        ir = compose_workbook(frames, None)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "out.xlsx"
        render_plan(plan, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=1).fill.fgColor.rgb is not None

    @pytest.mark.xlsx_ir
    def test_autofilter_present(self, tmp_path: Path) -> None:
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.io_backends.xlsx.xlsx_openpyxl import render_plan
        from openpyxl import load_workbook

        frames = {"products": pd.DataFrame([{"id": "a", "name": "Alpha"}])}
        ir = compose_workbook(frames, None)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "out.xlsx"
        render_plan(plan, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws.auto_filter.ref is not None

    @pytest.mark.xlsx_ir
    def test_helper_column_fill_in_xlsx(self, tmp_path: Path) -> None:
        """Verify helper columns get a distinct fill color in the output."""
        from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
        from spreadsheet_handling.io_backends.xlsx.xlsx_openpyxl import render_plan
        from openpyxl import load_workbook

        frames = {
            "products": pd.DataFrame([
                {"id": "a", "name": "Alpha", "_helper": "h1"},
                {"id": "b", "name": "Bravo", "_helper": "h2"},
            ])
        }
        ir = compose_workbook(frames, None)
        apply_ir_passes(ir, default_p1_passes())
        plan = build_render_plan(ir)

        out = tmp_path / "out.xlsx"
        render_plan(plan, out)

        wb = load_workbook(out)
        ws = wb.active
        # _helper is column 3; data rows are 2 and 3
        helper_cell = ws.cell(row=2, column=3)
        normal_cell = ws.cell(row=2, column=2)
        # helper cell should have fill, normal data cell should not
        assert helper_cell.fill.fgColor.rgb != "00000000"  # not default/empty
        assert normal_cell.fill.fgColor.rgb == "00000000"  # default/empty
