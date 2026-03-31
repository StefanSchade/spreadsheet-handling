# tests/integration/io_backends/xlsx/test_ir_golden.py
"""
FTR-PARITY-IR-P3 / FTR-CLEANUP-IR-P4 — IR-only golden structural tests.

Verifies that the IR renderer produces correct XLSX output:
  - Sheet names & order
  - Data cell values
  - Header styles (bold + fill)
  - AutoFilter presence
  - Freeze panes
  - Data validations
  - Hidden _meta sheet
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Set

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend

pytestmark = pytest.mark.ftr("FTR-PARITY-IR-P3")


# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------

def _visible_sheet_names(path: Path) -> List[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets if ws.sheet_state != "hidden"]
    finally:
        wb.close()


def _hidden_sheet_names(path: Path) -> Set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {ws.title for ws in wb.worksheets if ws.sheet_state == "hidden"}
    finally:
        wb.close()


def _sheet_data(ws: Worksheet) -> List[List[str]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    return rows


def _header_styles(ws: Worksheet, row: int = 1) -> List[Dict[str, Any]]:
    styles = []
    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row=row, column=col)
        fill_rgb = None
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = str(cell.fill.fgColor.rgb)
            if rgb != "00000000":
                fill_rgb = rgb
        styles.append({
            "bold": bool(cell.font and cell.font.bold),
            "has_fill": fill_rgb is not None,
        })
    return styles


def _has_autofilter(ws: Worksheet) -> bool:
    return bool(ws.auto_filter and ws.auto_filter.ref)


def _freeze_cell(ws: Worksheet) -> str | None:
    fp = ws.freeze_panes
    return str(fp) if fp else None


def _validation_summaries(ws: Worksheet) -> List[Dict[str, Any]]:
    out = []
    for dv in ws.data_validations.dataValidation:
        out.append({"type": dv.type, "formula1": dv.formula1})
    return sorted(out, key=lambda d: (d["type"] or "", d["formula1"] or ""))


# ---------------------------------------------------------------------------
# Write helper
# ---------------------------------------------------------------------------

def _write(frames: dict, path: Path) -> Path:
    ExcelBackend().write_multi(frames, str(path))
    return Path(str(path).replace(".xlsx", "") + ".xlsx")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def flat_frames():
    return {
        "products": pd.DataFrame([
            {"id": "P-001", "name": "Alpha", "branch_id": "B-001"},
            {"id": "P-002", "name": "Beta", "branch_id": "B-002"},
        ]),
        "branches": pd.DataFrame([
            {"branch_id": "B-001", "manager": "Alice"},
            {"branch_id": "B-002", "manager": "Bob"},
        ]),
    }


@pytest.fixture
def frames_with_validation():
    return {
        "products": pd.DataFrame([
            {"id": "P-001", "name": "Alpha", "status": "active"},
            {"id": "P-002", "name": "Beta", "status": "draft"},
        ]),
        "_meta": {
            "constraints": [
                {
                    "sheet": "products",
                    "column": "status",
                    "rule": {"type": "in_list", "values": ["active", "draft", "archived"]},
                }
            ],
        },
    }


# ===========================================================================
# Golden tests — IR-only
# ===========================================================================

class TestSheetStructure:
    def test_visible_sheet_names(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        assert _visible_sheet_names(out) == ["products", "branches"]

    def test_sheet_count(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        assert len(_visible_sheet_names(out)) == 2


class TestDataEquivalence:
    def test_data_values(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        wb = load_workbook(out, data_only=True)
        try:
            data = _sheet_data(wb["products"])
            assert data[0] == ["id", "name", "branch_id"]  # header
            assert data[1] == ["P-001", "Alpha", "B-001"]
            assert data[2] == ["P-002", "Beta", "B-002"]
        finally:
            wb.close()


class TestHeaderStyles:
    def test_headers_are_bold(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        wb = load_workbook(out)
        try:
            for name in _visible_sheet_names(out):
                styles = _header_styles(wb[name])
                assert all(s["bold"] for s in styles), f"Not all headers bold in {name}"
        finally:
            wb.close()

    def test_headers_have_fill(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        wb = load_workbook(out)
        try:
            for name in _visible_sheet_names(out):
                styles = _header_styles(wb[name])
                assert all(s["has_fill"] for s in styles), f"Not all headers filled in {name}"
        finally:
            wb.close()


class TestAutoFilter:
    def test_autofilter_present(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        wb = load_workbook(out)
        try:
            for name in _visible_sheet_names(out):
                assert _has_autofilter(wb[name]), f"AutoFilter missing on {name}"
        finally:
            wb.close()


class TestFreeze:
    def test_freeze_when_enabled(self, tmp_path):
        frames = {
            "data": pd.DataFrame([{"a": "1", "b": "2"}]),
            "_meta": {"sheets": {"data": {"freeze_header": True}}},
        }
        out = _write(frames, tmp_path / "out")
        wb = load_workbook(out)
        try:
            assert _freeze_cell(wb["data"]) == "A2"
        finally:
            wb.close()


class TestValidations:
    def test_validation_applied(self, tmp_path, frames_with_validation):
        out = _write(frames_with_validation, tmp_path / "out")
        wb = load_workbook(out)
        try:
            dvs = _validation_summaries(wb["products"])
            assert len(dvs) >= 1, "No validations found"
            assert dvs[0]["type"] == "list"
            assert "active" in (dvs[0]["formula1"] or "")
        finally:
            wb.close()


class TestMetaSheet:
    def test_has_hidden_meta(self, tmp_path, flat_frames):
        out = _write(flat_frames, tmp_path / "out")
        hidden = _hidden_sheet_names(out)
        assert "_meta" in hidden
