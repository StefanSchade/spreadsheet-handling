"""XLSX horizontal alignment roundtrip + composition integration tests."""

from __future__ import annotations

import json

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend


pytestmark = pytest.mark.ftr("FTR-CELL-ALIGNMENT-ROUNDTRIP-P5")


def _write_source_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["B1"] = "Value"
    ws["C1"] = "Notes"
    ws["A2"] = "alpha"
    ws["B2"] = 1
    ws["C2"] = "ok"
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["B1"].alignment = Alignment(horizontal="center")
    ws["C1"].alignment = Alignment(horizontal="right")
    wb.save(path)
    wb.close()


def _horizontal_alignments(frames, sheet_name="Data"):
    return frames["_meta"]["sheets"][sheet_name]["horizontal_alignments"]


def test_xlsx_horizontal_alignments_are_parsed_into_sheet_metadata(tmp_path):
    source = tmp_path / "source_alignment.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = _horizontal_alignments(frames)
    assert aligns["A1"] == {"horizontal": "left", "source": "workbook"}
    assert aligns["B1"] == {"horizontal": "center", "source": "workbook"}
    assert aligns["C1"] == {"horizontal": "right", "source": "workbook"}


def test_xlsx_renderer_applies_horizontal_alignment_metadata(tmp_path):
    frames = {
        "Data": pd.DataFrame({"Category": ["alpha"], "Value": [1], "Notes": ["ok"]}),
        "_meta": {
            "sheets": {
                "Data": {
                    "horizontal_alignments": {
                        "A1": {"horizontal": "left", "source": "workbook"},
                        "B1": {"horizontal": "center", "source": "workbook"},
                        "C1": {"horizontal": "right", "source": "workbook"},
                    }
                }
            }
        },
    }
    out = tmp_path / "rendered_alignment.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert ws["A1"].alignment.horizontal == "left"
        assert ws["B1"].alignment.horizontal == "center"
        assert ws["C1"].alignment.horizontal == "right"
    finally:
        wb.close()


def test_xlsx_horizontal_alignments_survive_parse_write_parse_roundtrip(tmp_path):
    source = tmp_path / "source_alignment.xlsx"
    roundtripped = tmp_path / "roundtripped_alignment.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    ExcelBackend().write_multi(frames, str(roundtripped))
    back = ExcelBackend().read_multi(str(roundtripped), header_levels=1)

    aligns = _horizontal_alignments(back)
    assert aligns["A1"]["horizontal"] == "left"
    assert aligns["B1"]["horizontal"] == "center"
    assert aligns["C1"]["horizontal"] == "right"
    assert back["Data"]["Category"].tolist() == ["alpha"]


def test_xlsx_renderer_keeps_defaults_when_alignment_metadata_is_absent(tmp_path):
    out = tmp_path / "defaults_alignment.xlsx"

    ExcelBackend().write_multi(
        {"Data": pd.DataFrame({"Category": ["alpha"], "Value": [1]})},
        str(out),
    )

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        # openpyxl default is None / 'general'; either is accepted as "no metadata".
        assert ws["A1"].alignment.horizontal in (None, "general")
        assert ws["B1"].alignment.horizontal in (None, "general")
    finally:
        wb.close()


def test_xlsx_default_general_alignment_is_not_emitted_into_metadata(tmp_path):
    # An XLSX whose cells leave alignment.horizontal at the default (general /
    # None) must produce no horizontal_alignments entries on read. Otherwise
    # the canonical metadata bloats by one entry per cell openpyxl iterates.
    source = tmp_path / "defaults_source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["A2"] = "alpha"
    # Explicitly set 'general' to make sure it is dropped, not just absence.
    ws["A1"].alignment = Alignment(horizontal="general")
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    sheets_meta = frames.get("_meta", {}).get("sheets", {})
    assert "horizontal_alignments" not in sheets_meta.get("Data", {})


def test_xlsx_unsupported_horizontal_values_are_dropped_on_read(tmp_path):
    # `justify`, `fill`, `distributed`, `centerContinuous` are out-of-slice;
    # extracting them would force the renderer to emit values it does not
    # know how to translate to ODS, so they are dropped.
    source = tmp_path / "unsupported_alignment.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["A2"] = "alpha"
    ws["A1"].alignment = Alignment(horizontal="justify")
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    sheets_meta = frames.get("_meta", {}).get("sheets", {})
    assert "horizontal_alignments" not in sheets_meta.get("Data", {})


def test_set_horizontal_alignment_op_rejects_out_of_vocabulary_value():
    # Regression for M3: direct RenderPlan construction must not bypass the
    # canonical-vocabulary filter applied by rendering.flow. Passing an
    # out-of-slice value to the render op raises at construction time, so
    # invalid intent never reaches an XLSX or ODS renderer.
    from spreadsheet_handling.rendering.plan import SetHorizontalAlignment

    for invalid in ("justify", "fill", "distributed", "centerContinuous", "general", ""):
        with pytest.raises(ValueError, match="must be one of"):
            SetHorizontalAlignment(sheet="Data", row=1, col=1, horizontal=invalid)

    # Canonical values still construct without error.
    for valid in ("left", "center", "right"):
        SetHorizontalAlignment(sheet="Data", row=1, col=1, horizontal=valid)


def test_xlsx_stale_embedded_alignment_is_dropped_when_carrier_is_clear(tmp_path):
    # Regression for B1: a workbook that already carries embedded
    # ``horizontal_alignments`` metadata (e.g. written by a prior roundtrip)
    # but whose visible cells no longer carry the alignment attribute must be
    # read back without the stale metadata. Otherwise the next write would
    # silently reapply formatting the user has explicitly removed.
    source = tmp_path / "stale_alignment.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["A2"] = "alpha"
    # Intentionally leave A1.alignment at the default (no horizontal carrier).

    meta_ws = wb.create_sheet("_meta")
    blob = json.dumps(
        {
            "sheets": {
                "Data": {
                    "horizontal_alignments": {
                        "A1": {"horizontal": "left", "source": "workbook"},
                    }
                }
            }
        }
    )
    meta_ws["A1"] = "workbook_meta_blob"
    meta_ws["B1"] = blob
    meta_ws.sheet_state = "hidden"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    data_sheet_meta = frames.get("_meta", {}).get("sheets", {}).get("Data", {})
    assert "horizontal_alignments" not in data_sheet_meta, (
        "Stale embedded horizontal_alignments must be cleared when the "
        f"visible carrier has no alignment; got: {data_sheet_meta}"
    )


def test_xlsx_rotation_and_horizontal_alignment_coexist_on_same_cell(tmp_path):
    # Composition guard against the openpyxl Alignment-overwrite trap:
    # SetTextOrientation and SetHorizontalAlignment must both survive on the
    # same cell, regardless of which render op the flow emits first.
    frames = {
        "Data": pd.DataFrame({"Category": ["alpha"]}),
        "_meta": {
            "sheets": {
                "Data": {
                    "text_orientations": {
                        "A1": {"rotation": 90, "source": "workbook"},
                    },
                    "horizontal_alignments": {
                        "A1": {"horizontal": "center", "source": "workbook"},
                    },
                }
            }
        },
    }
    out = tmp_path / "composition.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    try:
        cell = wb["Data"]["A1"]
        assert cell.alignment.text_rotation == 90
        assert cell.alignment.horizontal == "center"
    finally:
        wb.close()

    # And the full parse-write-parse roundtrip carries both attributes.
    back = ExcelBackend().read_multi(str(out), header_levels=1)
    sheets_meta = back["_meta"]["sheets"]["Data"]
    assert sheets_meta["text_orientations"]["A1"]["rotation"] == 90
    assert sheets_meta["horizontal_alignments"]["A1"]["horizontal"] == "center"
