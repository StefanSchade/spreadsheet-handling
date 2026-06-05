"""XLSX vertical alignment roundtrip + composition integration tests."""

from __future__ import annotations

import json

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend


pytestmark = pytest.mark.ftr("FTR-VERTICAL-ALIGNMENT-ROUNDTRIP-P5")


def _write_source_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "T"
    ws["B1"] = "C"
    ws["C1"] = "B"
    ws["A2"] = "alpha"
    ws["B2"] = 1
    ws["C2"] = "ok"
    ws["A1"].alignment = Alignment(vertical="top")
    ws["B1"].alignment = Alignment(vertical="center")
    ws["C1"].alignment = Alignment(vertical="bottom")
    wb.save(path)
    wb.close()


def _vertical_alignments(frames, sheet_name="Data"):
    return frames["_meta"]["sheets"][sheet_name]["vertical_alignments"]


def test_xlsx_vertical_alignments_are_parsed_into_sheet_metadata(tmp_path):
    source = tmp_path / "source_vertical.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = _vertical_alignments(frames)
    assert aligns["A1"] == {"vertical": "top", "source": "workbook"}
    assert aligns["B1"] == {"vertical": "center", "source": "workbook"}
    assert aligns["C1"] == {"vertical": "bottom", "source": "workbook"}


def test_xlsx_renderer_applies_vertical_alignment_metadata(tmp_path):
    frames = {
        "Data": pd.DataFrame({"T": ["alpha"], "C": [1], "B": ["ok"]}),
        "_meta": {
            "sheets": {
                "Data": {
                    "vertical_alignments": {
                        "A1": {"vertical": "top", "source": "workbook"},
                        "B1": {"vertical": "center", "source": "workbook"},
                        "C1": {"vertical": "bottom", "source": "workbook"},
                    }
                }
            }
        },
    }
    out = tmp_path / "rendered_vertical.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert ws["A1"].alignment.vertical == "top"
        assert ws["B1"].alignment.vertical == "center"
        assert ws["C1"].alignment.vertical == "bottom"
    finally:
        wb.close()


def test_xlsx_vertical_alignments_survive_parse_write_parse_roundtrip(tmp_path):
    source = tmp_path / "source_vertical.xlsx"
    roundtripped = tmp_path / "roundtripped_vertical.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    ExcelBackend().write_multi(frames, str(roundtripped))
    back = ExcelBackend().read_multi(str(roundtripped), header_levels=1)

    aligns = _vertical_alignments(back)
    assert aligns["A1"]["vertical"] == "top"
    assert aligns["B1"]["vertical"] == "center"
    assert aligns["C1"]["vertical"] == "bottom"
    assert back["Data"]["T"].tolist() == ["alpha"]


def test_xlsx_renderer_keeps_defaults_when_vertical_metadata_is_absent(tmp_path):
    out = tmp_path / "defaults_vertical.xlsx"

    ExcelBackend().write_multi(
        {"Data": pd.DataFrame({"T": ["alpha"], "C": [1]})},
        str(out),
    )

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert ws["A1"].alignment.vertical is None
        assert ws["B1"].alignment.vertical is None
    finally:
        wb.close()


def test_xlsx_default_vertical_alignment_is_not_emitted_into_metadata(tmp_path):
    # A workbook whose cells leave alignment.vertical at the default (None)
    # must produce no vertical_alignments entries on read.
    source = tmp_path / "defaults_source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "T"
    ws["A2"] = "alpha"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    sheets_meta = frames.get("_meta", {}).get("sheets", {})
    assert "vertical_alignments" not in sheets_meta.get("Data", {})


@pytest.mark.parametrize("unsupported", ["justify", "distributed"])
def test_xlsx_unsupported_vertical_values_are_dropped_on_read(tmp_path, unsupported):
    # XLSX-only ``justify`` / ``distributed`` are dropped on read; the
    # renderer has no canonical contract for them across backends.
    source = tmp_path / f"unsupported_{unsupported}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "T"
    ws["A2"] = "alpha"
    ws["A1"].alignment = Alignment(vertical=unsupported)
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    sheets_meta = frames.get("_meta", {}).get("sheets", {})
    assert "vertical_alignments" not in sheets_meta.get("Data", {})


def test_xlsx_rotation_horizontal_and_vertical_coexist_on_same_cell(tmp_path):
    # Composition guard for three-attribute alignment: SetTextOrientation,
    # SetHorizontalAlignment, and SetVerticalAlignment must all survive on
    # the same cell regardless of dispatch order. The _merge_xlsx_alignment
    # helper preserves unrelated attributes by construction; this test pins
    # the intended composition behaviour for future regressions.
    frames = {
        "Data": pd.DataFrame({"Category": ["alpha"]}),
        "_meta": {
            "sheets": {
                "Data": {
                    "text_orientations": {
                        "A1": {"rotation": 90, "source": "workbook"},
                    },
                    "horizontal_alignments": {
                        "A1": {"horizontal": "center", "source": "workbook"},
                    },
                    "vertical_alignments": {
                        "A1": {"vertical": "top", "source": "workbook"},
                    },
                }
            }
        },
    }
    out = tmp_path / "composition_vertical.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    try:
        cell = wb["Data"]["A1"]
        assert cell.alignment.text_rotation == 90
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "top"
    finally:
        wb.close()

    back = ExcelBackend().read_multi(str(out), header_levels=1)
    sheets_meta = back["_meta"]["sheets"]["Data"]
    assert sheets_meta["text_orientations"]["A1"]["rotation"] == 90
    assert sheets_meta["horizontal_alignments"]["A1"]["horizontal"] == "center"
    assert sheets_meta["vertical_alignments"]["A1"]["vertical"] == "top"


def test_xlsx_stale_embedded_vertical_alignment_is_dropped_when_carrier_is_clear(tmp_path):
    # Stale-meta regression for vertical_alignments on XLSX: a workbook
    # carrying embedded ``vertical_alignments`` metadata but with cleared
    # visible carriers must read back without the stale entry. Validates the
    # carrier-authority call site is wired through
    # ``apply_cell_addressed_presentation_meta`` with
    # ``family_key="vertical_alignments"``.
    source = tmp_path / "stale_vertical.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["A2"] = "alpha"
    # Intentionally leave A1.alignment at the default (no vertical carrier).

    meta_ws = wb.create_sheet("_meta")
    blob = json.dumps(
        {
            "sheets": {
                "Data": {
                    "vertical_alignments": {
                        "A1": {"vertical": "top", "source": "workbook"},
                    }
                }
            }
        }
    )
    meta_ws["A1"] = "workbook_meta_blob"
    meta_ws["B1"] = blob
    meta_ws.sheet_state = "hidden"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    data_sheet_meta = frames.get("_meta", {}).get("sheets", {}).get("Data", {})
    assert "vertical_alignments" not in data_sheet_meta, (
        "Stale embedded vertical_alignments must be cleared when the visible "
        f"carrier has no vertical alignment; got: {data_sheet_meta}"
    )


def test_set_vertical_alignment_op_rejects_out_of_vocabulary_value():
    # Direct RenderPlan construction must not bypass the canonical-vocabulary
    # filter. Out-of-slice values (ODS-only ``middle`` / ``automatic``,
    # XLSX-only ``justify`` / ``distributed``, empty string) raise at
    # construction time.
    from spreadsheet_handling.rendering.plan import SetVerticalAlignment

    for invalid in ("middle", "automatic", "justify", "distributed", "", "Top"):
        with pytest.raises(ValueError, match="must be one of"):
            SetVerticalAlignment(sheet="Data", row=1, col=1, vertical=invalid)

    for valid in ("top", "center", "bottom"):
        SetVerticalAlignment(sheet="Data", row=1, col=1, vertical=valid)
