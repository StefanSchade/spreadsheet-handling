from __future__ import annotations

import json

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend


pytestmark = pytest.mark.ftr("FTR-TEXT-ORIENTATION-ROUNDTRIP-P5")


def _write_source_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["B1"] = "Value"
    ws["A2"] = "alpha"
    ws["B2"] = 1
    ws["A1"].alignment = Alignment(text_rotation=90)
    ws["B1"].alignment = Alignment(text_rotation=45)
    wb.save(path)
    wb.close()


def _text_orientations(frames, sheet_name="Data"):
    return frames["_meta"]["sheets"][sheet_name]["text_orientations"]


def test_xlsx_text_orientations_are_parsed_into_sheet_metadata(tmp_path):
    source = tmp_path / "source_rotation.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    orients = _text_orientations(frames)
    assert orients["A1"] == {"rotation": 90, "source": "workbook"}
    assert orients["B1"] == {"rotation": 45, "source": "workbook"}


def test_xlsx_renderer_applies_text_orientation_metadata(tmp_path):
    frames = {
        "Data": pd.DataFrame({"Category": ["alpha"], "Value": [1]}),
        "_meta": {
            "sheets": {
                "Data": {
                    "text_orientations": {
                        "A1": {"rotation": 90, "source": "workbook"},
                        "B1": {"rotation": 45, "source": "workbook"},
                    }
                }
            }
        },
    }
    out = tmp_path / "rendered_rotation.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert ws["A1"].alignment.text_rotation == 90
        assert ws["B1"].alignment.text_rotation == 45
    finally:
        wb.close()


def test_xlsx_text_orientations_survive_parse_write_parse_roundtrip(tmp_path):
    source = tmp_path / "source_rotation.xlsx"
    roundtripped = tmp_path / "roundtripped_rotation.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    ExcelBackend().write_multi(frames, str(roundtripped))
    back = ExcelBackend().read_multi(str(roundtripped), header_levels=1)

    orients = _text_orientations(back)
    assert orients["A1"]["rotation"] == 90
    assert orients["B1"]["rotation"] == 45
    assert back["Data"]["Category"].tolist() == ["alpha"]


def test_xlsx_renderer_keeps_defaults_when_orientation_metadata_is_absent(tmp_path):
    out = tmp_path / "defaults_rotation.xlsx"

    ExcelBackend().write_multi(
        {"Data": pd.DataFrame({"Category": ["alpha"], "Value": [1]})},
        str(out),
    )

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert ws["A1"].alignment.text_rotation in (None, 0)
        assert ws["B1"].alignment.text_rotation in (None, 0)
    finally:
        wb.close()


@pytest.mark.ftr("FTR-PRESENTATION-META-CARRIER-AUTHORITY-P5")
def test_xlsx_stale_embedded_text_orientation_is_dropped_when_carrier_is_clear(tmp_path):
    # Regression for the separate-backlog stale-meta defect closed by
    # FTR-PRESENTATION-META-CARRIER-AUTHORITY-P5: a workbook that already
    # carries embedded ``text_orientations`` metadata (e.g. written by a prior
    # roundtrip) but whose visible cells no longer carry a rotation must be
    # read back without the stale metadata. Otherwise the next write would
    # silently reapply rotation the user has explicitly removed.
    source = tmp_path / "stale_orientation.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Category"
    ws["A2"] = "alpha"
    # Intentionally leave A1.alignment at the default (no rotation carrier).

    meta_ws = wb.create_sheet("_meta")
    blob = json.dumps(
        {
            "sheets": {
                "Data": {
                    "text_orientations": {
                        "A1": {"rotation": 90, "source": "workbook"},
                    }
                }
            }
        }
    )
    meta_ws["A1"] = "workbook_meta_blob"
    meta_ws["B1"] = blob
    meta_ws.sheet_state = "hidden"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    data_sheet_meta = frames.get("_meta", {}).get("sheets", {}).get("Data", {})
    assert "text_orientations" not in data_sheet_meta, (
        "Stale embedded text_orientations must be cleared when the visible "
        f"carrier has no rotation; got: {data_sheet_meta}"
    )
