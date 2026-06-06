"""XLSX column-level presentation metadata readback tests.

Guards BUG-COLUMN-LEVEL-PRESENTATION-META-READBACK-P5: formatting applied to
an entire column (via ``ws.column_dimensions``) must be read back into
cell-addressed metadata just like formatting applied cell-by-cell.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend


pytestmark = pytest.mark.ftr("BUG-COLUMN-LEVEL-PRESENTATION-META-READBACK-P5")


def _make_workbook_with_column_alignment(path, *, horizontal=None, vertical=None, rotation=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Header"
    ws["A2"] = "value"
    ws["B1"] = "Other"
    ws["B2"] = "data"
    align_kwargs = {}
    if horizontal is not None:
        align_kwargs["horizontal"] = horizontal
    if vertical is not None:
        align_kwargs["vertical"] = vertical
    if rotation is not None:
        align_kwargs["text_rotation"] = rotation
    ws.column_dimensions["A"].alignment = Alignment(**align_kwargs)
    wb.save(path)
    wb.close()


def test_xlsx_column_level_horizontal_alignment_is_read(tmp_path):
    source = tmp_path / "col_horiz.xlsx"
    _make_workbook_with_column_alignment(source, horizontal="center")

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = frames["_meta"]["sheets"]["Data"]["horizontal_alignments"]
    assert aligns["A1"]["horizontal"] == "center"
    assert aligns["A2"]["horizontal"] == "center"
    assert "B1" not in aligns
    assert "B2" not in aligns


def test_xlsx_column_level_vertical_alignment_is_read(tmp_path):
    source = tmp_path / "col_vert.xlsx"
    _make_workbook_with_column_alignment(source, vertical="top")

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = frames["_meta"]["sheets"]["Data"]["vertical_alignments"]
    assert aligns["A1"]["vertical"] == "top"
    assert aligns["A2"]["vertical"] == "top"
    assert "B1" not in aligns
    assert "B2" not in aligns


def test_xlsx_column_level_text_orientation_is_read(tmp_path):
    source = tmp_path / "col_rot.xlsx"
    _make_workbook_with_column_alignment(source, rotation=90)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    orients = frames["_meta"]["sheets"]["Data"]["text_orientations"]
    assert orients["A1"]["rotation"] == 90
    assert orients["A2"]["rotation"] == 90
    assert "B1" not in orients
    assert "B2" not in orients


def test_xlsx_cell_level_style_overrides_column_default(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Override"
    ws["A2"] = "Default"
    ws["A3"] = "Also default"
    ws.column_dimensions["A"].alignment = Alignment(horizontal="center")
    ws["A1"].alignment = Alignment(horizontal="right")
    source = tmp_path / "col_override.xlsx"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = frames["_meta"]["sheets"]["Data"]["horizontal_alignments"]
    assert aligns["A1"]["horizontal"] == "right"
    assert aligns["A2"]["horizontal"] == "center"
    assert aligns["A3"]["horizontal"] == "center"


def test_xlsx_explicit_default_alignment_suppresses_column_fallback(tmp_path):
    # A cell with an explicit alignment whose value is non-canonical ("general")
    # must NOT inherit the column default.  The explicit xf carrier is present, so
    # the column fallback must be suppressed even though "general" is dropped from
    # canonical metadata.  A2 has no explicit style and must receive the column default.
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "explicit-general"
    ws["A2"] = "no-style"
    ws.column_dimensions["A"].alignment = Alignment(horizontal="center")
    ws["A1"].alignment = Alignment(horizontal="general")
    source = tmp_path / "explicit_general.xlsx"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = frames["_meta"]["sheets"]["Data"].get("horizontal_alignments", {})
    assert "A1" not in aligns, "explicit 'general' must suppress the column fallback"
    assert aligns["A2"]["horizontal"] == "center"


def test_xlsx_explicit_rotation_zero_suppresses_column_fallback(tmp_path):
    # A cell with an explicit alignment carrier where textRotation=0 must NOT
    # inherit the column's rotation of 90.
    # openpyxl only writes an <alignment> element when bool(Alignment(...)) is
    # True.  Alignment(text_rotation=0) alone is falsy (0 is the default) so it
    # is never written.  Adding horizontal="left" makes it truthy, forcing both
    # horizontal and textRotation=0 into the xf's <alignment> element.
    # After roundtrip: _get_cell_xf_alignment returns Alignment(horizontal="left",
    # text_rotation=0); text_rotation is 0 (not None) → explicit carrier present
    # → column fallback suppressed.
    # A2 has no explicit style and must still inherit rotation=90.
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "explicit-zero"
    ws["A2"] = "no-style"
    ws.column_dimensions["A"].alignment = Alignment(text_rotation=90)
    ws["A1"].alignment = Alignment(text_rotation=0, horizontal="left")
    source = tmp_path / "explicit_zero_rot.xlsx"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    orients = frames["_meta"]["sheets"]["Data"].get("text_orientations", {})
    assert "A1" not in orients, "explicit alignment carrier with text_rotation=0 must suppress column fallback"
    assert orients["A2"]["rotation"] == 90


def test_xlsx_cell_with_unrelated_style_inherits_column_rotation(tmp_path):
    # A cell with a font-only xf (has_style=True but no <alignment> element)
    # must still inherit the column-level text rotation.
    # _get_cell_xf_alignment returns None for a font-only xf, so the fallback
    # path applies and the cell receives the column default.
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "font-only"
    ws["A2"] = "no-style"
    ws.column_dimensions["A"].alignment = Alignment(text_rotation=90)
    ws["A1"].font = Font(bold=True)
    source = tmp_path / "font_only_inherits_rot.xlsx"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    orients = frames["_meta"]["sheets"]["Data"].get("text_orientations", {})
    assert orients["A1"]["rotation"] == 90, "font-only cell must inherit column rotation"
    assert orients["A2"]["rotation"] == 90


def test_xlsx_column_dimension_span_provides_fallback(tmp_path):
    # A <col min="N" max="M"> span must provide fallback metadata for every
    # column in the span, not only the representative key column.
    # openpyxl groups adjacent identical column dimensions into a single span
    # on save, so after a roundtrip the loaded dim for "B" has min=2, max=3
    # and there is no separate "C" entry.  Both B and C cells must get metadata.
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "skip"
    ws["B1"] = "col_b"
    ws["C1"] = "col_c"
    ws["B2"] = "val_b"
    ws["C2"] = "val_c"
    align = Alignment(horizontal="center")
    ws.column_dimensions["B"].alignment = align
    ws.column_dimensions["C"].alignment = align
    source = tmp_path / "col_span.xlsx"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    aligns = frames["_meta"]["sheets"]["Data"]["horizontal_alignments"]
    assert aligns["B1"]["horizontal"] == "center"
    assert aligns["B2"]["horizontal"] == "center"
    assert aligns["C1"]["horizontal"] == "center"
    assert aligns["C2"]["horizontal"] == "center"
    assert "A1" not in aligns
    assert "A2" not in aligns
