# tests/integration/io_backends/xlsx/test_parity_ir_p3.py
"""
FTR-PARITY-IR-P3 — Structural parity between legacy and IR XLSX backends.

Compares normalized structural equivalence (not byte equality):
  - Sheet names & order
  - Data cell values
  - Header styles (bold + fill)
  - AutoFilter presence
  - Freeze panes
  - Data validations
  - Hidden _meta sheet (IR-only, verified separately)
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend

pytestmark = pytest.mark.ftr("FTR-PARITY-IR-P3")


# ---------------------------------------------------------------------------
# Extraction helpers — pull normalised structure from an XLSX file
# ---------------------------------------------------------------------------

def _visible_sheet_names(path: Path) -> List[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets if ws.sheet_state != "hidden"]
    finally:
        wb.close()


def _hidden_sheet_names(path: Path) -> Set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {ws.title for ws in wb.worksheets if ws.sheet_state == "hidden"}
    finally:
        wb.close()


def _sheet_data(ws: Worksheet) -> List[List[str]]:
    """Return all cell values as string grid (normalised)."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    return rows


def _header_styles(ws: Worksheet, row: int = 1) -> List[Dict[str, Any]]:
    """Extract bold + fill for each cell in a header row."""
    styles = []
    for col in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(row=row, column=col)
        fill_rgb = None
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = str(cell.fill.fgColor.rgb)
            if rgb != "00000000":  # default / no fill
                fill_rgb = rgb
        styles.append({
            "bold": bool(cell.font and cell.font.bold),
            "has_fill": fill_rgb is not None,
        })
    return styles


def _has_autofilter(ws: Worksheet) -> bool:
    return bool(ws.auto_filter and ws.auto_filter.ref)


def _freeze_cell(ws: Worksheet) -> str | None:
    fp = ws.freeze_panes
    return str(fp) if fp else None


def _validation_summaries(ws: Worksheet) -> List[Dict[str, Any]]:
    """Return simplified validation descriptors (type + formula)."""
    out = []
    for dv in ws.data_validations.dataValidation:
        out.append({
            "type": dv.type,
            "formula1": dv.formula1,
        })
    return sorted(out, key=lambda d: (d["type"] or "", d["formula1"] or ""))


# ---------------------------------------------------------------------------
# Write helper — produce XLSX with a specific backend
# ---------------------------------------------------------------------------

def _write_with_backend(
    frames: dict,
    path: Path,
    backend: str,
    monkeypatch: pytest.MonkeyPatch,
) -> Path:
    if backend == "ir":
        monkeypatch.setenv("SH_XLSX_BACKEND", "ir")
    else:
        monkeypatch.delenv("SH_XLSX_BACKEND", raising=False)

    ExcelBackend().write_multi(frames, str(path))
    return Path(str(path).replace(".xlsx", "") + ".xlsx")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def flat_frames():
    """Flat single-header frames — the baseline both backends must handle."""
    return {
        "products": pd.DataFrame([
            {"id": "P-001", "name": "Alpha", "branch_id": "B-001"},
            {"id": "P-002", "name": "Beta", "branch_id": "B-002"},
        ]),
        "branches": pd.DataFrame([
            {"branch_id": "B-001", "manager": "Alice"},
            {"branch_id": "B-002", "manager": "Bob"},
        ]),
    }


@pytest.fixture
def frames_with_validation():
    """Frames with in-list validation constraint."""
    return {
        "products": pd.DataFrame([
            {"id": "P-001", "name": "Alpha", "status": "active"},
            {"id": "P-002", "name": "Beta", "status": "draft"},
        ]),
        "_meta": {
            "constraints": [
                {
                    "sheet": "products",
                    "column": "status",
                    "rule": {"type": "in_list", "values": ["active", "draft", "archived"]},
                }
            ],
        },
    }


# ===========================================================================
# Parity tests — parametrised over both backends
# ===========================================================================

class TestSheetStructure:
    """Sheet names and order must match (excluding hidden sheets)."""

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_visible_sheet_names(self, tmp_path, flat_frames, backend, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", backend, monkeypatch)
        names = _visible_sheet_names(out)
        assert names == ["products", "branches"]

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_sheet_count(self, tmp_path, flat_frames, backend, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", backend, monkeypatch)
        names = _visible_sheet_names(out)
        assert len(names) == 2


class TestDataEquivalence:
    """Cell values in data rows must be structurally equivalent."""

    def test_data_matches_across_backends(self, tmp_path, flat_frames, monkeypatch):
        legacy_path = _write_with_backend(flat_frames, tmp_path / "legacy", "legacy", monkeypatch)
        ir_path = _write_with_backend(flat_frames, tmp_path / "ir", "ir", monkeypatch)

        for sheet_name in ["products", "branches"]:
            wb_l = load_workbook(legacy_path, data_only=True)
            wb_i = load_workbook(ir_path, data_only=True)
            try:
                data_l = _sheet_data(wb_l[sheet_name])
                data_i = _sheet_data(wb_i[sheet_name])
                assert data_l == data_i, f"Data mismatch in sheet '{sheet_name}'"
            finally:
                wb_l.close()
                wb_i.close()


class TestHeaderStyles:
    """Both backends must apply bold + fill to header row."""

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_headers_are_bold(self, tmp_path, flat_frames, backend, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", backend, monkeypatch)
        wb = load_workbook(out)
        try:
            for name in _visible_sheet_names(out):
                styles = _header_styles(wb[name])
                assert all(s["bold"] for s in styles), (
                    f"Not all headers bold in {name} ({backend})"
                )
        finally:
            wb.close()

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_headers_have_fill(self, tmp_path, flat_frames, backend, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", backend, monkeypatch)
        wb = load_workbook(out)
        try:
            for name in _visible_sheet_names(out):
                styles = _header_styles(wb[name])
                assert all(s["has_fill"] for s in styles), (
                    f"Not all headers filled in {name} ({backend})"
                )
        finally:
            wb.close()


class TestAutoFilter:
    """Both backends must set AutoFilter on data sheets."""

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_autofilter_present(self, tmp_path, flat_frames, backend, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", backend, monkeypatch)
        wb = load_workbook(out)
        try:
            for name in _visible_sheet_names(out):
                assert _has_autofilter(wb[name]), (
                    f"AutoFilter missing on {name} ({backend})"
                )
        finally:
            wb.close()


class TestFreeze:
    """Freeze pane at A2 when freeze_header is enabled."""

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_freeze_when_enabled(self, tmp_path, backend, monkeypatch):
        frames = {
            "data": pd.DataFrame([{"a": "1", "b": "2"}]),
            "_meta": {"sheets": {"data": {"freeze_header": True}}},
        }
        if backend == "legacy":
            # Legacy freeze is hardcoded off in default _decorate_workbook;
            # we skip the legacy freeze test as it's not configurable from meta.
            pytest.skip("Legacy freeze not meta-driven")

        out = _write_with_backend(frames, tmp_path / "out", backend, monkeypatch)
        wb = load_workbook(out)
        try:
            assert _freeze_cell(wb["data"]) == "A2"
        finally:
            wb.close()


class TestValidations:
    """Both backends must produce equivalent data validations from constraints."""

    @pytest.mark.parametrize("backend", ["legacy", "ir"])
    def test_validation_applied(self, tmp_path, frames_with_validation, backend, monkeypatch):
        out = _write_with_backend(frames_with_validation, tmp_path / "out", backend, monkeypatch)
        wb = load_workbook(out)
        try:
            dvs = _validation_summaries(wb["products"])
            assert len(dvs) >= 1, f"No validations found ({backend})"
            assert dvs[0]["type"] == "list"
            # formula should contain the allowed values
            assert "active" in (dvs[0]["formula1"] or "")
        finally:
            wb.close()


class TestMetaSheet:
    """IR backend produces hidden _meta sheet; legacy does not."""

    def test_ir_has_hidden_meta(self, tmp_path, flat_frames, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", "ir", monkeypatch)
        hidden = _hidden_sheet_names(out)
        assert "_meta" in hidden

    def test_legacy_has_no_meta_sheet(self, tmp_path, flat_frames, monkeypatch):
        out = _write_with_backend(flat_frames, tmp_path / "out", "legacy", monkeypatch)
        hidden = _hidden_sheet_names(out)
        assert "_meta" not in hidden


# ===========================================================================
# Side-by-side structural comparison (comprehensive)
# ===========================================================================

class TestSideBySideParity:
    """Direct comparison of legacy vs IR output for flat single-header frames."""

    def test_visible_sheets_identical(self, tmp_path, flat_frames, monkeypatch):
        lp = _write_with_backend(flat_frames, tmp_path / "l", "legacy", monkeypatch)
        ip = _write_with_backend(flat_frames, tmp_path / "i", "ir", monkeypatch)
        assert _visible_sheet_names(lp) == _visible_sheet_names(ip)

    def test_header_bold_parity(self, tmp_path, flat_frames, monkeypatch):
        lp = _write_with_backend(flat_frames, tmp_path / "l", "legacy", monkeypatch)
        ip = _write_with_backend(flat_frames, tmp_path / "i", "ir", monkeypatch)
        wb_l = load_workbook(lp)
        wb_i = load_workbook(ip)
        try:
            for name in _visible_sheet_names(lp):
                bold_l = [s["bold"] for s in _header_styles(wb_l[name])]
                bold_i = [s["bold"] for s in _header_styles(wb_i[name])]
                assert bold_l == bold_i, f"Bold mismatch in {name}"
        finally:
            wb_l.close()
            wb_i.close()

    def test_header_fill_parity(self, tmp_path, flat_frames, monkeypatch):
        """Both backends apply a header fill (exact color may differ)."""
        lp = _write_with_backend(flat_frames, tmp_path / "l", "legacy", monkeypatch)
        ip = _write_with_backend(flat_frames, tmp_path / "i", "ir", monkeypatch)
        wb_l = load_workbook(lp)
        wb_i = load_workbook(ip)
        try:
            for name in _visible_sheet_names(lp):
                fill_l = [s["has_fill"] for s in _header_styles(wb_l[name])]
                fill_i = [s["has_fill"] for s in _header_styles(wb_i[name])]
                assert fill_l == fill_i, f"Fill presence mismatch in {name}"
        finally:
            wb_l.close()
            wb_i.close()

    def test_autofilter_parity(self, tmp_path, flat_frames, monkeypatch):
        lp = _write_with_backend(flat_frames, tmp_path / "l", "legacy", monkeypatch)
        ip = _write_with_backend(flat_frames, tmp_path / "i", "ir", monkeypatch)
        wb_l = load_workbook(lp)
        wb_i = load_workbook(ip)
        try:
            for name in _visible_sheet_names(lp):
                assert _has_autofilter(wb_l[name]) == _has_autofilter(wb_i[name]), (
                    f"AutoFilter mismatch in {name}"
                )
        finally:
            wb_l.close()
            wb_i.close()

    def test_validation_parity(self, tmp_path, frames_with_validation, monkeypatch):
        lp = _write_with_backend(frames_with_validation, tmp_path / "l", "legacy", monkeypatch)
        ip = _write_with_backend(frames_with_validation, tmp_path / "i", "ir", monkeypatch)
        wb_l = load_workbook(lp)
        wb_i = load_workbook(ip)
        try:
            dvs_l = _validation_summaries(wb_l["products"])
            dvs_i = _validation_summaries(wb_i["products"])
            assert len(dvs_l) == len(dvs_i), "Validation count mismatch"
            for vl, vi in zip(dvs_l, dvs_i):
                assert vl["type"] == vi["type"]
                assert vl["formula1"] == vi["formula1"]
        finally:
            wb_l.close()
            wb_i.close()
