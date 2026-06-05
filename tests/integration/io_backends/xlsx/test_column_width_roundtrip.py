from __future__ import annotations

import json

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend


pytestmark = pytest.mark.ftr("FTR-COLUMN-WIDTH-ROUNDTRIP-P4")


def _write_source_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "id"
    ws["B1"] = "title"
    ws["A2"] = "P-1"
    ws["B2"] = "Compact matrix label"
    ws.column_dimensions["A"].width = 18.0
    ws.column_dimensions["B"].width = 42.5
    wb.save(path)
    wb.close()


def _column_widths(frames, sheet_name="Data"):
    return frames["_meta"]["sheets"][sheet_name]["column_widths"]


def test_xlsx_column_widths_are_parsed_into_sheet_metadata(tmp_path):
    source = tmp_path / "source_widths.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    widths = _column_widths(frames)
    assert widths["A"] == {"width": pytest.approx(18.0), "source": "workbook"}
    assert widths["B"] == {"width": pytest.approx(42.5), "source": "workbook"}


def test_xlsx_renderer_applies_column_width_metadata(tmp_path):
    frames = {
        "Data": pd.DataFrame({"id": ["P-1"], "title": ["Compact matrix label"]}),
        "_meta": {
            "sheets": {
                "Data": {
                    "column_widths": {
                        "A": {"width": 18.0, "source": "workbook"},
                        "B": {"width": 42.5, "source": "workbook"},
                    }
                }
            }
        },
    }
    out = tmp_path / "rendered_widths.xlsx"

    ExcelBackend().write_multi(frames, str(out))

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert ws.column_dimensions["A"].width == pytest.approx(18.0)
        assert ws.column_dimensions["B"].width == pytest.approx(42.5)
    finally:
        wb.close()


def test_xlsx_column_widths_survive_parse_write_parse_roundtrip(tmp_path):
    source = tmp_path / "source_widths.xlsx"
    roundtripped = tmp_path / "roundtripped_widths.xlsx"
    _write_source_workbook(source)

    frames = ExcelBackend().read_multi(str(source), header_levels=1)
    ExcelBackend().write_multi(frames, str(roundtripped))
    back = ExcelBackend().read_multi(str(roundtripped), header_levels=1)

    widths = _column_widths(back)
    assert widths["A"]["width"] == pytest.approx(18.0)
    assert widths["B"]["width"] == pytest.approx(42.5)
    assert back["Data"].to_dict(orient="records") == [
        {"id": "P-1", "title": "Compact matrix label"}
    ]


def test_xlsx_renderer_keeps_defaults_when_width_metadata_is_absent(tmp_path):
    out = tmp_path / "defaults.xlsx"

    ExcelBackend().write_multi(
        {"Data": pd.DataFrame({"id": ["P-1"], "title": ["Alpha"]})},
        str(out),
    )

    wb = load_workbook(out)
    try:
        ws = wb["Data"]
        assert "A" not in ws.column_dimensions
        assert "B" not in ws.column_dimensions
    finally:
        wb.close()


@pytest.mark.ftr("FTR-PRESENTATION-META-CARRIER-AUTHORITY-P5")
def test_xlsx_stale_embedded_column_widths_dropped_from_sheet_ir_and_render_plan(
    tmp_path,
):
    # Regression for FTR-PRESENTATION-META-CARRIER-AUTHORITY-P5 review
    # finding B1: the XLSX parser previously seeded
    # ``sheet_ir.meta["__column_widths"]`` from embedded ``_meta`` hints
    # *before* the carrier-authority helper ran. The helper cleared the
    # embedded blob but the in-memory cache survived, so
    # ``build_render_plan(ir)`` still emitted ``SetColumnWidth`` for widths
    # the user had cleared from the visible carrier. This test pins the
    # parser/render-plan boundary, not just the public readback.
    from spreadsheet_handling.io_backends.xlsx.openpyxl_parser import parse_workbook
    from spreadsheet_handling.rendering.flow import build_render_plan
    from spreadsheet_handling.rendering.plan import SetColumnWidth

    source = tmp_path / "stale_widths_parser.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "id"
    ws["B1"] = "title"
    ws["A2"] = "P-1"
    ws["B2"] = "label"
    # Default column dimensions: no custom width carrier.

    meta_ws = wb.create_sheet("_meta")
    blob = json.dumps(
        {
            "sheets": {
                "Data": {
                    "column_widths": {
                        "A": {"width": 18.0, "source": "workbook"},
                        "B": {"width": 42.5, "source": "workbook"},
                    }
                }
            }
        }
    )
    meta_ws["A1"] = "workbook_meta_blob"
    meta_ws["B1"] = blob
    meta_ws.sheet_state = "hidden"
    wb.save(source)
    wb.close()

    ir = parse_workbook(source)

    assert "__column_widths" not in ir.sheets["Data"].meta, (
        "Stale embedded column_widths must not survive in SheetIR.meta when "
        "the visible carrier has no custom widths; got: "
        f"{ir.sheets['Data'].meta.get('__column_widths')!r}"
    )

    plan = build_render_plan(ir)
    column_width_ops = [op for op in plan.ops if isinstance(op, SetColumnWidth)]
    assert column_width_ops == [], (
        "build_render_plan must not emit SetColumnWidth ops for cleared "
        f"carrier widths; got: {column_width_ops!r}"
    )


@pytest.mark.ftr("FTR-PRESENTATION-META-CARRIER-AUTHORITY-P5")
def test_xlsx_stale_embedded_column_widths_dropped_when_carrier_is_clear(tmp_path):
    # Regression for the separate-backlog stale-meta defect closed by
    # FTR-PRESENTATION-META-CARRIER-AUTHORITY-P5: a workbook that already
    # carries embedded ``column_widths`` metadata (e.g. written by a prior
    # roundtrip) but whose visible columns no longer carry an explicit width
    # must be read back without the stale metadata. Otherwise the next write
    # would silently reapply widths the user has explicitly removed.
    source = tmp_path / "stale_widths.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "id"
    ws["B1"] = "title"
    ws["A2"] = "P-1"
    ws["B2"] = "label"
    # Intentionally leave column_dimensions at default (no custom width).

    meta_ws = wb.create_sheet("_meta")
    blob = json.dumps(
        {
            "sheets": {
                "Data": {
                    "column_widths": {
                        "A": {"width": 18.0, "source": "workbook"},
                        "B": {"width": 42.5, "source": "workbook"},
                    }
                }
            }
        }
    )
    meta_ws["A1"] = "workbook_meta_blob"
    meta_ws["B1"] = blob
    meta_ws.sheet_state = "hidden"
    wb.save(source)
    wb.close()

    frames = ExcelBackend().read_multi(str(source), header_levels=1)

    data_sheet_meta = frames.get("_meta", {}).get("sheets", {}).get("Data", {})
    assert "column_widths" not in data_sheet_meta, (
        "Stale embedded column_widths must be cleared when the visible "
        f"columns carry no explicit width; got: {data_sheet_meta}"
    )
