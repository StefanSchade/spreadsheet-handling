"""Regression tests for the XLSX hidden ``_meta`` carrier projection.

The canonical workbook meta contract does **not** include a top-level
``workbook_meta_blob`` key -- that name is reserved for the hidden ``_meta``
sheet's first cell (the carrier). Two regressions on either side of the
carrier boundary previously broke workbook-view readback:

* on **write**, a ``frames["_meta"]`` that already carried a
  ``workbook_meta_blob`` string (e.g. a sheet-level representation fed
  back into the writer) was wholesale dumped as the new blob, nesting
  the old blob and burying ``workbook_view`` one level deep;
* on **read**, a malformed XLSX with that nested shape projected back
  into ``frames["_meta"]`` with ``workbook_meta_blob`` shadowing the
  canonical ``workbook_view`` mapping.

Both sides now canonicalize at the projection boundary, so the
``workbook_view`` mapping is always recoverable as a top-level key in
``frames["_meta"]`` after an XLSX parse.

A follow-up residual symptom is also covered: when the inner blob is
degenerate (empty / unparseable / a non-dict value) but the outer
wrapper still carries the sheet-level export markers (``author``,
``exported_at``, ``version``), the projection used to leak those
markers as if they were canonical workbook meta. The
``carrier-layer-marker strip`` in
``canonicalize_workbook_meta`` removes them on every level so
``frames["_meta"]`` never advertises wrapper noise as canonical content.

See:
* ``docs/backlog/BUG-XLSX-WORKBOOK-VIEW-BLOB-READBACK-P4A.adoc``
* ``docs/backlog/BUG-XLSX-WORKBOOK-VIEW-CANONICAL-META-LOSS-P4A.adoc``
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from spreadsheet_handling.io_backends.xlsx.openpyxl_parser import parse_workbook
from spreadsheet_handling.io_backends.xlsx.xlsx_backend import (
    ExcelBackend,
    load_xlsx,
    save_xlsx,
)
from spreadsheet_handling.rendering.workbook_projection import workbookir_to_frames


pytestmark = pytest.mark.ftr("BUG-XLSX-WORKBOOK-VIEW-BLOB-READBACK-P4A")


def _canonical_workbook_view() -> dict:
    return {
        "mode": "editable",
        "drop_redundant_data": True,
        "unknown_frame_policy": "fail",
        "sheets": [{"frame": "items", "sheet": "items", "order": 0}],
        "sheet_mappings": [{"frame": "items", "sheet": "items"}],
    }


def _canonical_meta_with_workbook_view() -> dict:
    return {
        "auto_filter": True,
        "freeze_header": True,
        "workbook_view": _canonical_workbook_view(),
    }


def _frames_with_workbook_view() -> dict:
    return {
        "items": pd.DataFrame({"id": ["a", "b"], "label": ["one", "two"]}),
        "_meta": _canonical_meta_with_workbook_view(),
    }


def test_xlsx_carrier_roundtrip_recovers_top_level_workbook_view(
    tmp_path: Path,
) -> None:
    """Writing canonical meta with ``workbook_view`` and reading it back
    yields ``frames["_meta"]["workbook_view"]`` as a mapping at the top
    level, without any nested ``workbook_meta_blob`` shadow."""
    out_path = tmp_path / "carrier.xlsx"
    save_xlsx(_frames_with_workbook_view(), str(out_path))

    ir = parse_workbook(out_path)
    frames = workbookir_to_frames(ir)

    meta = frames["_meta"]
    assert isinstance(meta, dict)
    assert "workbook_meta_blob" not in meta, (
        "Carrier-local helper field workbook_meta_blob must not appear at "
        f"the top level of frames['_meta'] after parse; got keys: {sorted(meta)}"
    )
    view = meta.get("workbook_view")
    assert isinstance(view, dict), (
        "_meta.workbook_view must be a mapping after XLSX carrier parse; "
        f"got type={type(view).__name__}, meta keys={sorted(meta)}"
    )
    assert view["sheet_mappings"] == [{"frame": "items", "sheet": "items"}]


def test_xlsx_writer_does_not_nest_stray_workbook_meta_blob(
    tmp_path: Path,
) -> None:
    """If a caller passes ``frames["_meta"]`` that already carries a
    stray ``workbook_meta_blob`` string (a sheet-level representation
    fed back into the writer), the writer must unwrap it instead of
    nesting it inside a new blob. Otherwise the next read would bury
    ``workbook_view`` one level deep."""
    canonical = _canonical_meta_with_workbook_view()
    sheet_level_representation = {
        "author": None,
        "version": None,
        "exported_at": None,
        "workbook_meta_blob": json.dumps(canonical, sort_keys=True),
    }
    frames = {
        "items": pd.DataFrame({"id": ["a"], "label": ["one"]}),
        "_meta": sheet_level_representation,
    }

    out_path = tmp_path / "wrapped.xlsx"
    save_xlsx(frames, str(out_path))

    read_back = load_xlsx(str(out_path))
    meta = read_back["_meta"]

    assert "workbook_meta_blob" not in meta, (
        "Writer must canonicalize away stray workbook_meta_blob from input "
        f"frames['_meta']; got keys: {sorted(meta)}"
    )
    view = meta.get("workbook_view")
    assert isinstance(view, dict), (
        "Inner canonical workbook_view must surface at the top level after "
        f"writer canonicalization; got type={type(view).__name__}"
    )
    assert view["sheet_mappings"] == [{"frame": "items", "sheet": "items"}]


def test_xlsx_reader_unwraps_nested_workbook_meta_blob(tmp_path: Path) -> None:
    """A malformed XLSX whose blob is itself a wrapper containing a
    nested ``workbook_meta_blob`` must still project to a clean
    ``frames["_meta"]`` with ``workbook_view`` at the top level."""
    canonical = _canonical_meta_with_workbook_view()
    inner_blob = json.dumps(canonical, sort_keys=True)
    wrapped = {
        "author": None,
        "version": None,
        "exported_at": None,
        "workbook_meta_blob": inner_blob,
    }
    wrapped_blob = json.dumps(wrapped, sort_keys=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(["id", "label"])
    ws.append(["a", "one"])
    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["workbook_meta_blob", wrapped_blob])
    meta_ws.append(["version", None])
    meta_ws.append(["exported_at", None])
    meta_ws.append(["author", None])

    out_path = tmp_path / "malformed.xlsx"
    wb.save(out_path)
    wb.close()

    ir = parse_workbook(out_path)
    frames = workbookir_to_frames(ir)

    meta = frames["_meta"]
    assert "workbook_meta_blob" not in meta
    view = meta.get("workbook_view")
    assert isinstance(view, dict)
    assert view["sheet_mappings"] == [{"frame": "items", "sheet": "items"}]


def test_xlsx_carrier_end_to_end_via_excel_backend_preserves_workbook_view(
    tmp_path: Path,
) -> None:
    """End-to-end check through ``ExcelBackend.read_multi`` (the public
    routing surface) so the contract holds through the full backend
    facade -- not just the lower-level parser projection helpers."""
    out_path = tmp_path / "backend.xlsx"
    ExcelBackend().write_multi(_frames_with_workbook_view(), str(out_path))

    read_back = ExcelBackend().read_multi(str(out_path), header_levels=1)
    meta = read_back["_meta"]

    assert "workbook_meta_blob" not in meta
    view = meta.get("workbook_view")
    assert isinstance(view, dict)
    assert view["sheet_mappings"] == [{"frame": "items", "sheet": "items"}]


@pytest.mark.ftr("BUG-XLSX-WORKBOOK-VIEW-CANONICAL-META-LOSS-P4A")
def test_xlsx_reader_strips_carrier_layer_markers_from_canonical_meta(
    tmp_path: Path,
) -> None:
    """A hidden ``_meta`` sheet whose carrier blob is itself a wrapper of
    sheet-level export markers (``author``, ``version``, ``exported_at``)
    plus a degenerate inner ``workbook_meta_blob`` must not surface those
    markers as if they were canonical workbook meta. Without the strip,
    canonicalize would return ``{author, exported_at, sheets, version}``
    -- the exact post-fix residual symptom reported against the original
    fix commit (see
    ``docs/backlog/BUG-XLSX-WORKBOOK-VIEW-CANONICAL-META-LOSS-P4A.adoc``).
    """
    # Outer wrapper that contains the four wrapper-shell keys plus a
    # degenerate inner blob. Mirrors what the bug report observed in
    # ``tmp/dino_stories.xlsx`` after the e939623 fix.
    wrapped_blob = json.dumps(
        {
            "author": None,
            "exported_at": None,
            "version": None,
            "sheets": {"items": {}},
            "workbook_meta_blob": "",
        },
        sort_keys=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append(["id", "label"])
    ws.append(["a", "one"])
    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["workbook_meta_blob", wrapped_blob])
    meta_ws.append(["version", None])
    meta_ws.append(["exported_at", None])
    meta_ws.append(["author", None])

    out_path = tmp_path / "wrapper_only.xlsx"
    wb.save(out_path)
    wb.close()

    ir = parse_workbook(out_path)
    frames = workbookir_to_frames(ir)

    meta = frames["_meta"]
    assert isinstance(meta, dict)
    assert "workbook_meta_blob" not in meta
    # The wrapper-shell keys describe the hidden _meta sheet (MetaPass
    # markers) and must not appear as canonical content. Without the
    # carrier-layer strip, canonicalize used to leak them.
    leaked = sorted(set(meta) & {"author", "exported_at", "version"})
    assert not leaked, (
        "Carrier-layer marker keys leaked into canonical frames['_meta'] "
        f"after XLSX parse: {leaked}. They describe the hidden _meta "
        "sheet, not the canonical workbook meta contract."
    )


@pytest.mark.ftr("BUG-XLSX-WORKBOOK-VIEW-CANONICAL-META-LOSS-P4A")
def test_xlsx_writer_strips_carrier_layer_markers_from_input_meta(
    tmp_path: Path,
) -> None:
    """If a caller passes ``frames["_meta"]`` carrying sheet-level export
    markers alongside canonical content, the writer must canonicalize
    them away rather than serialising them into the next blob. Otherwise
    the next read would observe ``author``/``exported_at``/``version``
    as if they were canonical workbook meta."""
    contaminated_meta = {
        "author": None,
        "exported_at": None,
        "version": None,
        **_canonical_meta_with_workbook_view(),
    }
    frames = {
        "items": pd.DataFrame({"id": ["a"], "label": ["one"]}),
        "_meta": contaminated_meta,
    }

    out_path = tmp_path / "contaminated_input.xlsx"
    save_xlsx(frames, str(out_path))

    read_back = load_xlsx(str(out_path))
    meta = read_back["_meta"]

    assert "workbook_meta_blob" not in meta
    leaked = sorted(set(meta) & {"author", "exported_at", "version"})
    assert not leaked, (
        "Writer must strip carrier-layer marker keys from canonical "
        f"meta before serialising; leaked: {leaked}"
    )
    view = meta.get("workbook_view")
    assert isinstance(view, dict), (
        "workbook_view must still surface after the writer canonicalizes "
        f"input meta carrying carrier-layer noise; got meta keys={sorted(meta)}"
    )
    assert view["sheet_mappings"] == [{"frame": "items", "sheet": "items"}]


@pytest.mark.ftr("BUG-XLSX-WORKBOOK-VIEW-CANONICAL-META-LOSS-P4A")
def test_xlsx_carrier_roundtrip_recovers_workbook_view_after_degenerate_inner_blob(
    tmp_path: Path,
) -> None:
    """End-to-end check on the bug's exact reproducer: a caller writes
    ``frames["_meta"]`` shaped as a sheet-level wrapper with an empty
    inner blob -- the worst case of a corrupted carrier surviving a
    previous roundtrip. After the strip, the post-read meta is honest
    about the absence of canonical content rather than disguising it
    behind wrapper noise.
    """
    contaminated = {
        "author": None,
        "exported_at": None,
        "version": None,
        "sheets": {"items": {}},
        "workbook_meta_blob": "",
    }
    frames = {
        "items": pd.DataFrame({"id": ["a"], "label": ["one"]}),
        "_meta": contaminated,
    }

    out_path = tmp_path / "degenerate_carrier.xlsx"
    save_xlsx(frames, str(out_path))

    read_back = load_xlsx(str(out_path))
    meta = read_back["_meta"]

    # No workbook_meta_blob, no wrapper-shell noise.
    assert "workbook_meta_blob" not in meta
    leaked = sorted(set(meta) & {"author", "exported_at", "version"})
    assert not leaked, (
        "Post-fix residual symptom: wrapper-shell keys must not appear "
        f"in frames['_meta']; got {sorted(meta)}"
    )
