"""Cross-adapter spreadsheet parity integration slice.

Runs representative XLSX and ODS product paths and compares visible frames,
canonical metadata, and parser-visible semantics for the currently supported
portable spreadsheet surface.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pytest

from spreadsheet_handling.io_backends.ods.odf_parser import parse_workbook as parse_ods_workbook
from spreadsheet_handling.io_backends.ods.ods_backend import OdsBackend
from spreadsheet_handling.io_backends.xlsx.openpyxl_parser import (
    parse_workbook as parse_xlsx_workbook,
)
from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend
from spreadsheet_handling.core.formulas import FormulaSpec, formula_list_values
from spreadsheet_handling.rendering.ir import SheetIR
from spreadsheet_handling.domain.extractions.frame_extract import extract_frame
from spreadsheet_handling.domain.workbook_views import configure_workbook_view
from spreadsheet_handling.domain.transformations.join_views import join_frames
from spreadsheet_handling.domain.transformations.tabular_views import pivot_frame

pytestmark = [
    pytest.mark.ftr("FTR-ODS-CALC-PARITY-TESTS-P3J"),
]


def _supported_frames() -> dict[str, object]:
    return {
        "Products": pd.DataFrame(
            [
                {"id": "P-001", "status": "new", "_helper": "A"},
                {"id": "P-002", "status": "done", "_helper": "B"},
            ]
        ),
        "Summary": pd.DataFrame(
            [
                {"metric": "count", "value": "2"},
            ]
        ),
        "_meta": {
            "version": "3.4",
            "author": "cross-adapter-parity",
            "freeze_header": True,
            "auto_filter": True,
            "helper_prefix": "_",
            "header_fill_rgb": "#CCE5FF",
            "helper_fill_rgb": "#FFF5CC",
            "constraints": [
                {
                    "sheet": "Products",
                    "column": "status",
                    "rule": {"type": "in_list", "values": ["new", "done", "archived"]},
                }
            ],
            "sheets": {
                "Summary": {
                    "freeze_header": False,
                }
            },
        },
    }


def _multiheader_frames() -> dict[str, object]:
    columns = pd.MultiIndex.from_tuples(
        [
            ("order", "id"),
            ("order", "status"),
            ("audit", "_helper"),
        ]
    )
    return {
        "Orders": pd.DataFrame(
            [
                ["O-001", "new", "A"],
                ["O-002", "done", "B"],
            ],
            columns=columns,
        ),
        "_meta": {
            "version": "3.4",
            "author": "cross-adapter-parity",
            "freeze_header": True,
            "auto_filter": True,
            "helper_prefix": "_",
        },
    }


def _write_both(frames: dict[str, object], tmp_path: Path, stem: str) -> tuple[Path, Path]:
    xlsx_path = tmp_path / f"{stem}.xlsx"
    ods_path = tmp_path / f"{stem}.ods"

    ExcelBackend().write_multi(frames, str(xlsx_path))
    OdsBackend().write_multi(frames, str(ods_path))

    return xlsx_path, ods_path


def _visible_sheet_names(frames: dict[str, object]) -> list[str]:
    return [
        name
        for name, value in frames.items()
        if name != "_meta" and isinstance(value, pd.DataFrame)
    ]


def _assert_roundtrip_matches_expected(
    actual: dict[str, object],
    expected: dict[str, object],
) -> None:
    expected_visible = _visible_sheet_names(expected)

    assert _visible_sheet_names(actual) == expected_visible
    assert actual["_meta"] == expected["_meta"]

    for name in expected_visible:
        pd.testing.assert_frame_equal(
            actual[name],
            expected[name],
            check_dtype=False,
        )


def _named_range_summary(sheet: SheetIR) -> list[tuple[str, str, tuple[int, int, int, int]]]:
    return sorted(
        (named_range.name, named_range.sheet, named_range.area)
        for named_range in sheet.named_ranges
    )


def _validation_summary(
    sheet: SheetIR,
) -> list[tuple[str, tuple[int, int, int, int], tuple[str, ...], bool]]:
    return sorted(
        (
            validation.kind,
            validation.area,
            _normalized_validation_values(validation.formula),
            validation.allow_empty,
        )
        for validation in sheet.validations
    )


def _normalized_validation_values(formula: FormulaSpec) -> tuple[str, ...]:
    return formula_list_values(formula)


def _supported_sheet_summary(sheet: SheetIR) -> dict[str, Any]:
    table = sheet.tables[0]
    return {
        "headers": list(table.headers),
        "data": [list(row) for row in (table.data or [])],
        "header_rows": table.header_rows,
        "n_rows": table.n_rows,
        "n_cols": table.n_cols,
        "options": dict(sheet.meta.get("options", {})),
        "autofilter_ref": sheet.meta.get("__autofilter_ref"),
        "named_ranges": _named_range_summary(sheet),
        "validations": _validation_summary(sheet),
    }


def _supported_workbook_summary(workbook) -> dict[str, Any]:
    return {
        "visible_sheets": list(workbook.sheets),
        "hidden_sheets": sorted(workbook.hidden_sheets),
        "sheets": {
            name: _supported_sheet_summary(sheet) for name, sheet in workbook.sheets.items()
        },
    }


def test_supported_frames_and_meta_roundtrip_portably_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = _supported_frames()
    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="supported")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    _assert_roundtrip_matches_expected(xlsx_back, frames)
    _assert_roundtrip_matches_expected(ods_back, frames)

    for sheet_name in _visible_sheet_names(frames):
        pd.testing.assert_frame_equal(
            xlsx_back[sheet_name],
            ods_back[sheet_name],
            check_dtype=False,
        )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


def test_supported_parser_semantics_match_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = _supported_frames()
    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="semantics")

    xlsx_ir = parse_xlsx_workbook(xlsx_path)
    ods_ir = parse_ods_workbook(ods_path)

    assert _supported_workbook_summary(xlsx_ir) == _supported_workbook_summary(ods_ir)
    assert "_meta" in xlsx_ir.hidden_sheets
    assert "_meta" in ods_ir.hidden_sheets


def test_multiheader_roundtrip_is_portable_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = _multiheader_frames()
    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="multiheader")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    assert isinstance(xlsx_back["Orders"].columns, pd.MultiIndex)
    assert isinstance(ods_back["Orders"].columns, pd.MultiIndex)

    pd.testing.assert_frame_equal(
        xlsx_back["Orders"],
        frames["Orders"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        ods_back["Orders"],
        frames["Orders"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        xlsx_back["Orders"],
        ods_back["Orders"],
        check_dtype=False,
    )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


@pytest.mark.ftr("FTR-GENERIC-FRAME-EXTRACTIONS-P4A")
def test_extracted_frame_renders_portably_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = extract_frame(
        {
            "Variables": pd.DataFrame(
                [
                    {"ID": "v2", "label": "Amount", "active": False},
                    {"ID": "v1", "label": "Rate", "active": True},
                ]
            )
        },
        source="Variables",
        output="VisibleVariables",
        columns=["ID", "label"],
        where={"column": "active", "equals": True},
        sort_by=["ID"],
    )

    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="extracted-frame")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    pd.testing.assert_frame_equal(
        xlsx_back["VisibleVariables"],
        frames["VisibleVariables"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        ods_back["VisibleVariables"],
        frames["VisibleVariables"],
        check_dtype=False,
    )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


@pytest.mark.ftr("FTR-DECLARATIVE-TABULAR-VIEW-OPS-P4A")
def test_pivoted_frame_renders_portably_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = pivot_frame(
        {
            "MappingRows": pd.DataFrame(
                [
                    {"variable_id": "v1", "mapping_name": "request", "display": "amount"},
                    {"variable_id": "v1", "mapping_name": "response", "display": "result"},
                    {"variable_id": "v2", "mapping_name": "request", "display": "term"},
                ]
            )
        },
        source="MappingRows",
        output="MappingView",
        index_columns=["variable_id"],
        column_key="mapping_name",
        value_column="display",
        column_keys=["request", "response"],
    )

    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="pivoted-frame")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    pd.testing.assert_frame_equal(
        xlsx_back["MappingView"],
        frames["MappingView"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        ods_back["MappingView"],
        frames["MappingView"],
        check_dtype=False,
    )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


@pytest.mark.ftr("FTR-SIMPLE-JOIN-VIEWS-P4A")
def test_joined_frame_renders_portably_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = join_frames(
        {
            "Variables": pd.DataFrame(
                [
                    {"variable_id": "v1", "label": "Rate"},
                    {"variable_id": "v2", "label": "Amount"},
                ]
            ),
            "Metadata": pd.DataFrame(
                [
                    {"variable_id": "v1", "component": "cashflow"},
                ]
            ),
        },
        left="Variables",
        right="Metadata",
        output="VariableView",
        key="variable_id",
        how="left",
    )

    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="joined-frame")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    pd.testing.assert_frame_equal(
        xlsx_back["VariableView"],
        frames["VariableView"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        ods_back["VariableView"],
        frames["VariableView"],
        check_dtype=False,
    )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


@pytest.mark.ftr("FTR-DECLARATIVE-WORKBOOK-VIEWS-P4A")
def test_configured_workbook_view_renders_portably_across_xlsx_and_ods(tmp_path: Path) -> None:
    frames = configure_workbook_view(
        {
            "variables_view": pd.DataFrame(
                [
                    {"variable_id": "v1", "label": "Rate"},
                ]
            ),
            "products_view": pd.DataFrame(
                [
                    {"product_id": "P-001", "label": "Annuity"},
                ]
            ),
            "raw_variables": pd.DataFrame(
                [
                    {"variable_id": "v1"},
                ]
            ),
        },
        sheets=[
            {"frame": "products_view", "sheet": "Products"},
            {
                "frame": "variables_view",
                "sheet": "Variables",
                "options": {"freeze_header": True},
            },
        ],
    )

    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="configured-workbook-view")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    assert _visible_sheet_names(xlsx_back) == ["Products", "Variables"]
    assert _visible_sheet_names(ods_back) == ["Products", "Variables"]
    pd.testing.assert_frame_equal(
        xlsx_back["Products"],
        frames["products_view"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        ods_back["Variables"],
        frames["variables_view"],
        check_dtype=False,
    )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


@pytest.mark.ftr("FTR-HELPER-COLUMN-STYLE-METADATA-P4A")
def test_explicit_helper_column_style_metadata_is_portable(tmp_path: Path) -> None:
    frames = {
        "Variables": pd.DataFrame(
            [
                {"ID": "v1", "value_label_de": "Rate", "editable_value": "x"},
            ]
        ),
        "_meta": {
            "sheets": {
                "Variables": {
                    "helper_columns": ["value_label_de"],
                    "helper_fill_rgb": "#FFF2CC",
                }
            }
        },
    }
    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="explicit-helper-style")

    xlsx_back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
    ods_back = OdsBackend().read_multi(str(ods_path), header_levels=1)

    pd.testing.assert_frame_equal(
        xlsx_back["Variables"],
        frames["Variables"],
        check_dtype=False,
    )
    pd.testing.assert_frame_equal(
        ods_back["Variables"],
        frames["Variables"],
        check_dtype=False,
    )
    assert xlsx_back["_meta"] == ods_back["_meta"] == frames["_meta"]


def test_freeze_parse_hint_difference_is_an_explicit_accepted_gap(tmp_path: Path) -> None:
    frames = _supported_frames()
    xlsx_path, ods_path = _write_both(frames, tmp_path, stem="freeze-gap")

    xlsx_ir = parse_xlsx_workbook(xlsx_path)
    ods_ir = parse_ods_workbook(ods_path)

    assert xlsx_ir.sheets["Products"].meta["options"]["freeze_header"] is True
    assert ods_ir.sheets["Products"].meta["options"]["freeze_header"] is True

    assert xlsx_ir.sheets["Products"].meta["__freeze"] == {"row": 2, "col": 1}
    assert "__freeze" not in ods_ir.sheets["Products"].meta

    assert "__freeze" not in xlsx_ir.sheets["Summary"].meta
    assert "__freeze" not in ods_ir.sheets["Summary"].meta


@pytest.mark.ftr("FTR-EDITABLE-COLUMNS-AND-PROTECTION-P4A")
def test_protection_is_xlsx_only_ods_documented_gap(tmp_path: Path) -> None:
    """XLSX applies protection ops; ODS skips them (documented capability gap)."""
    from openpyxl import load_workbook
    from spreadsheet_handling.rendering.composer.layout_composer import compose_workbook
    from spreadsheet_handling.rendering.flow import (
        apply_ir_passes,
        build_render_plan,
        default_p1_passes,
    )
    from spreadsheet_handling.io_backends.xlsx.openpyxl_renderer import (
        render_workbook as render_xlsx,
    )
    from spreadsheet_handling.io_backends.ods.odf_renderer import (
        render_workbook as render_ods,
    )

    frames = {
        "data": pd.DataFrame(
            [{"id": "v1", "value": 42, "helper_note": "auto"}]
        )
    }
    meta = {
        "sheets": {
            "data": {
                "protection": {"editable_columns": ["value"]},
            }
        }
    }
    ir = compose_workbook(frames, meta)
    apply_ir_passes(ir, default_p1_passes())
    plan = build_render_plan(ir)

    xlsx_path = tmp_path / "protected.xlsx"
    ods_path = tmp_path / "protected.ods"
    render_xlsx(plan, xlsx_path)
    render_ods(plan, ods_path)

    # XLSX: protection applied
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.protection.sheet is True
    assert ws.cell(row=2, column=2).protection.locked is False

    # ODS: file produced without error (protection is a documented gap)
    assert ods_path.exists()

