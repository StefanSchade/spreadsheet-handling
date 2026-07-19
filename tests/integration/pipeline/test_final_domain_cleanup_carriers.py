"""Carrier-level integration tests for implicit final domain cleanup.

Proves the FTR-META-ONTOLOGY-REMOVAL-WORKBOOK-PROJECTION-EPIC-P4A first
cleanup slice at the orchestrator boundary:

* structured (json_dir) and spreadsheet (xlsx) adapters observe the same
  cleaned logical frame set before any spreadsheet-specific projection;
* consumed cleanup commands do not roundtrip through the workbook meta blob
  and cannot re-execute against frames a reverse pipeline recreates;
* transformation intent metadata referencing the removed frame survives and
  is sufficient for the inverse transformation to recreate it;
* builder keep mode provides the generic equivalent of the consumer-owned
  ``worldbuilding/plugins/frame_sets.py::keep_frames`` workaround.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from spreadsheet_handling.application.orchestrator import orchestrate
from spreadsheet_handling.pipeline.build import build_steps_from_config


pytestmark = pytest.mark.ftr("FTR-META-ONTOLOGY-REMOVAL-WORKBOOK-PROJECTION-EPIC-P4A")


def _write_json_dir(path: Path, data: dict[str, list[dict]]) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for name, records in data.items():
        (path / f"{name}.json").write_text(
            json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
        )


SAMPLE_DATA = {
    "stories": [
        {"id": "s1", "title": "First"},
        {"id": "s2", "title": "Second"},
    ],
    "story_places": [
        {"story_id": "s1", "place": "cave", "relation": "setting"},
        {"story_id": "s2", "place": "camp", "relation": "journey"},
    ],
}

_XREF_STEP = {
    "step": "contract_xref",
    "relation": "story_places",
    "output": "story_place_matrix",
    "row_keys": "story_id",
    "column_key": "place",
    "value": "relation",
    "drop_source": True,
    "name": "story_place_matrix",
}


def _json_frame_names(out_dir: Path) -> set[str]:
    return {path.stem for path in out_dir.glob("*.json")}


def _meta_sidecar(out_dir: Path) -> dict:
    sidecar = out_dir / "_meta.yaml"
    if not sidecar.exists():
        return {}
    loaded = yaml.safe_load(sidecar.read_text(encoding="utf-8"))
    return loaded if isinstance(loaded, dict) else {}


def test_structured_and_spreadsheet_adapters_observe_same_cleaned_frame_set(
    tmp_path: Path,
) -> None:
    in_dir = tmp_path / "in"
    _write_json_dir(in_dir, SAMPLE_DATA)
    json_out = tmp_path / "out_json"
    xlsx_out = tmp_path / "out.xlsx"

    steps = [_XREF_STEP]
    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "json_dir", "path": str(json_out)},
        steps=build_steps_from_config(steps),
    )
    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "xlsx", "path": str(xlsx_out)},
        steps=build_steps_from_config(steps),
    )

    json_frames = _json_frame_names(json_out)
    workbook = load_workbook(xlsx_out)
    xlsx_frames = {name for name in workbook.sheetnames if name != "_meta"}

    assert json_frames == {"stories", "story_place_matrix"}
    assert xlsx_frames == json_frames


def test_consumed_commands_do_not_survive_into_sidecar_or_workbook_blob(
    tmp_path: Path,
) -> None:
    in_dir = tmp_path / "in"
    _write_json_dir(in_dir, SAMPLE_DATA)
    json_out = tmp_path / "out_json"
    xlsx_out = tmp_path / "out.xlsx"

    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "json_dir", "path": str(json_out)},
        steps=build_steps_from_config([_XREF_STEP]),
    )
    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "xlsx", "path": str(xlsx_out)},
        steps=build_steps_from_config([_XREF_STEP]),
    )

    assert "pipeline_cleanup" not in _meta_sidecar(json_out)

    reread = orchestrate(
        input={"kind": "xlsx", "path": str(xlsx_out)},
        output={"kind": "discard", "path": "-"},
    )
    meta = reread.get("_meta") or {}
    assert "pipeline_cleanup" not in meta


def test_intent_meta_survives_and_inverse_recreates_dropped_frame(
    tmp_path: Path,
) -> None:
    in_dir = tmp_path / "in"
    _write_json_dir(in_dir, SAMPLE_DATA)
    xlsx_out = tmp_path / "out.xlsx"
    reimport_out = tmp_path / "reimported"

    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "xlsx", "path": str(xlsx_out)},
        steps=build_steps_from_config([_XREF_STEP]),
    )

    # The workbook no longer carries the relation frame, but the persisted
    # xref intent still references it: required intent, not a contradiction.
    # (contract_xref keys its intent entry by the relation name.)
    loaded = orchestrate(
        input={"kind": "xlsx", "path": str(xlsx_out)},
        output={"kind": "discard", "path": "-"},
    )
    assert "story_places" not in loaded
    loaded_intent = (loaded.get("_meta") or {}).get("xref_crosstable") or {}
    assert loaded_intent["story_places"]["relation"] == "story_places"
    assert loaded_intent["story_places"]["matrix"] == "story_place_matrix"

    orchestrate(
        input={"kind": "xlsx", "path": str(xlsx_out)},
        output={"kind": "json_dir", "path": str(reimport_out)},
        steps=build_steps_from_config(
            [
                {
                    "step": "expand_xref",
                    "matrix": "story_place_matrix",
                    "output": "story_places",
                    "row_keys": "story_id",
                    "column_key": "place",
                    "value": "relation",
                    "drop_empty": True,
                }
            ]
        ),
    )

    recreated = json.loads(
        (reimport_out / "story_places.json").read_text(encoding="utf-8")
    )
    by_key = {(row["story_id"], row["place"]): row["relation"] for row in recreated}
    assert by_key == {
        ("s1", "cave"): "setting",
        ("s2", "camp"): "journey",
    }


def test_builder_keep_mode_replaces_consumer_keep_frames_plugin(
    tmp_path: Path,
) -> None:
    # Parity characterization for worldbuilding's keep_frames workaround: a
    # pipeline producing intermediates retains only the declared final set.
    in_dir = tmp_path / "in"
    _write_json_dir(in_dir, SAMPLE_DATA)
    out_dir = tmp_path / "out"

    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "json_dir", "path": str(out_dir)},
        steps=build_steps_from_config(
            [
                {
                    "step": "extract_frame",
                    "source": "story_places",
                    "output": "story_settings",
                    "where": {"column": "relation", "equals": "setting"},
                },
                {
                    "step": "configure_pipeline_cleanup",
                    "keep_frames": ["stories", "story_settings"],
                },
            ]
        ),
    )

    assert _json_frame_names(out_dir) == {"stories", "story_settings"}
    assert "pipeline_cleanup" not in _meta_sidecar(out_dir)


def test_stale_view_mapping_after_keep_mode_fails_before_spreadsheet_creation(
    tmp_path: Path,
) -> None:
    in_dir = tmp_path / "in"
    _write_json_dir(
        in_dir,
        {
            "stories": SAMPLE_DATA["stories"],
            "editable_story_view": SAMPLE_DATA["stories"],
        },
    )
    (in_dir / "_meta.yaml").write_text(
        yaml.safe_dump(
            {
                "workbook_view": {
                    "sheets": [
                        {"frame": "editable_story_view", "sheet": "Stories", "order": 0}
                    ],
                    "sheet_mappings": [
                        {"frame": "editable_story_view", "sheet": "Stories"}
                    ],
                }
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    structured_out = tmp_path / "structured"

    orchestrate(
        input={"kind": "json_dir", "path": str(in_dir)},
        output={"kind": "json_dir", "path": str(structured_out)},
        steps=build_steps_from_config(
            [
                {
                    "step": "configure_pipeline_cleanup",
                    "keep_frames": ["stories"],
                }
            ]
        ),
    )

    persisted_meta = _meta_sidecar(structured_out)
    assert persisted_meta["workbook_view"]["sheet_mappings"] == [
        {"frame": "editable_story_view", "sheet": "Stories"}
    ]
    assert "pipeline_cleanup" not in persisted_meta

    spreadsheet_out = tmp_path / "stale-view.xlsx"
    with pytest.raises(KeyError, match="missing frame 'editable_story_view'"):
        orchestrate(
            input={"kind": "json_dir", "path": str(structured_out)},
            output={"kind": "xlsx", "path": str(spreadsheet_out)},
        )

    assert not spreadsheet_out.exists()


def test_xref_intent_parity_and_recreation_across_xlsx_and_ods(
    tmp_path: Path,
) -> None:
    """Both spreadsheet carriers persist the same minimal XRef inverse intent.

    The relation is dropped after a lossless contraction; each carrier's
    persisted intent references the absent relation frame (required inverse
    intent, not a contradiction) without run-local Resolution facets, and
    the inverse expansion recreates identical relations from either carrier.
    """
    in_dir = tmp_path / "in"
    _write_json_dir(in_dir, SAMPLE_DATA)
    artifacts = {
        "xlsx": tmp_path / "out.xlsx",
        "ods": tmp_path / "out.ods",
    }

    for kind, path in artifacts.items():
        orchestrate(
            input={"kind": "json_dir", "path": str(in_dir)},
            output={"kind": kind, "path": str(path)},
            steps=build_steps_from_config([_XREF_STEP]),
        )

    recreated: dict[str, list[dict]] = {}
    for kind, path in artifacts.items():
        loaded = orchestrate(
            input={"kind": kind, "path": str(path)},
            output={"kind": "discard", "path": "-"},
        )
        assert "story_places" not in loaded
        intent = (loaded.get("_meta") or {}).get("xref_crosstable") or {}
        assert intent["story_places"] == {
            "relation": "story_places",
            "matrix": "story_place_matrix",
            "row_keys": ["story_id"],
        }

        reimport_out = tmp_path / f"reimported_{kind}"
        orchestrate(
            input={"kind": kind, "path": str(path)},
            output={"kind": "json_dir", "path": str(reimport_out)},
            steps=build_steps_from_config(
                [
                    {
                        "step": "expand_xref",
                        "matrix": "story_place_matrix",
                        "output": "story_places",
                        "row_keys": "story_id",
                        "column_key": "place",
                        "value": "relation",
                        "drop_empty": True,
                    }
                ]
            ),
        )
        rows = json.loads(
            (reimport_out / "story_places.json").read_text(encoding="utf-8")
        )
        recreated[kind] = sorted(rows, key=lambda row: (row["story_id"], row["place"]))

    assert recreated["xlsx"] == recreated["ods"]
    assert {(r["story_id"], r["place"], r["relation"]) for r in recreated["xlsx"]} == {
        ("s1", "cave", "setting"),
        ("s2", "camp", "journey"),
    }
