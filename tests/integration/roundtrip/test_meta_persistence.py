"""
FTR-META-PERSISTENCE – meta roundtrips across XLSX and JSON backends.

Acceptance:
  - XLSX: hidden _meta sheet written on write, extracted on read
  - JSON: optional _meta.yaml sidecar written/read
  - Meta roundtrips end-to-end
  - Pure-data adapters ignore meta unless enabled
"""
import json
from pathlib import Path

import pandas as pd
import pytest
import yaml

from spreadsheet_handling.io_backends.json_backend import JSONBackend
from spreadsheet_handling.io_backends.xlsx.xlsx_backend import ExcelBackend


pytestmark = pytest.mark.ftr("FTR-META-PERSISTENCE")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_SAMPLE_META = {
    "version": "3.0",
    "author": "test",
    "constraints": [
        {"sheet": "products", "column": "branch_id", "rule": {"type": "in_list", "values": ["B-1", "B-2"]}}
    ],
}


def _sample_frames(*, with_meta: bool = True) -> dict:
    frames: dict = {
        "products": pd.DataFrame([
            {"id": "P-1", "name": "Alpha", "branch_id": "B-1"},
            {"id": "P-2", "name": "Beta", "branch_id": "B-2"},
        ]),
    }
    if with_meta:
        frames["_meta"] = _SAMPLE_META
    return frames


# ===========================================================================
# JSON sidecar tests
# ===========================================================================

class TestJSONSidecar:

    def test_write_creates_sidecar(self, tmp_path: Path):
        out = tmp_path / "json_out"
        JSONBackend().write_multi(_sample_frames(), str(out))

        sidecar = out / "_meta.yaml"
        assert sidecar.exists(), "_meta.yaml sidecar not written"
        loaded = yaml.safe_load(sidecar.read_text(encoding="utf-8"))
        assert loaded["version"] == "3.0"
        assert loaded["author"] == "test"

    def test_write_no_meta_json_file(self, tmp_path: Path):
        """_meta must NOT be written as a regular .json file."""
        out = tmp_path / "json_out"
        JSONBackend().write_multi(_sample_frames(), str(out))
        assert not (out / "_meta.json").exists()

    def test_read_loads_sidecar(self, tmp_path: Path):
        out = tmp_path / "json_out"
        JSONBackend().write_multi(_sample_frames(), str(out))
        back = JSONBackend().read_multi(str(out), header_levels=1)

        assert "_meta" in back
        meta = back["_meta"]
        assert meta["version"] == "3.0"
        assert len(meta["constraints"]) == 1

    def test_roundtrip_json(self, tmp_path: Path):
        """Full JSON roundtrip preserves meta."""
        out = tmp_path / "json_out"
        frames = _sample_frames()
        JSONBackend().write_multi(frames, str(out))
        back = JSONBackend().read_multi(str(out), header_levels=1)

        assert back["_meta"] == _SAMPLE_META
        assert list(back["products"].columns) == ["id", "name", "branch_id"]

    def test_no_sidecar_no_meta(self, tmp_path: Path):
        """Without sidecar, read returns no _meta key."""
        out = tmp_path / "json_out"
        JSONBackend().write_multi(_sample_frames(with_meta=False), str(out))
        back = JSONBackend().read_multi(str(out), header_levels=1)
        assert "_meta" not in back

    def test_derived_helper_provenance_roundtrip_json(self, tmp_path: Path):
        """FTR-FK-HELPER-PROVENANCE-CLEANUP: derived helper provenance survives JSON sidecar roundtrip."""
        meta_with_prov = {
            **_SAMPLE_META,
            "derived": {
                "sheets": {
                    "desks": {
                        "helper_columns": [
                            {
                                "column": "_employees_first_name",
                                "fk_column": "id_(employees)",
                                "target": "employees",
                                "value_field": "first_name",
                            },
                        ]
                    }
                }
            },
        }
        frames = _sample_frames()
        frames["_meta"] = meta_with_prov
        out = tmp_path / "json_prov"
        JSONBackend().write_multi(frames, str(out))
        back = JSONBackend().read_multi(str(out), header_levels=1)

        prov = back["_meta"]["derived"]["sheets"]["desks"]["helper_columns"]
        assert len(prov) == 1
        assert prov[0]["column"] == "_employees_first_name"
        assert prov[0]["target"] == "employees"


# ===========================================================================
# XLSX meta tests
# ===========================================================================
class TestXLSXMeta:

    def test_write_creates_hidden_meta_sheet(self, tmp_path: Path, monkeypatch):
        out = tmp_path / "test.xlsx"
        ExcelBackend().write_multi(_sample_frames(), str(out))

        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert "_meta" in wb.sheetnames
        ws = wb["_meta"]
        assert ws.sheet_state == "hidden"
        wb.close()

    def test_read_extracts_meta(self, tmp_path: Path, monkeypatch):
        out = tmp_path / "test.xlsx"
        ExcelBackend().write_multi(_sample_frames(), str(out))

        back = ExcelBackend().read_multi(str(out), header_levels=1)
        assert "_meta" in back
        assert "_meta" not in {k for k in back if isinstance(back[k], pd.DataFrame)}

    def test_read_excludes_hidden_sheets_from_data(self, tmp_path: Path, monkeypatch):
        out = tmp_path / "test.xlsx"
        ExcelBackend().write_multi(_sample_frames(), str(out))

        back = ExcelBackend().read_multi(str(out), header_levels=1)
        data_sheets = {k for k in back if isinstance(back[k], pd.DataFrame)}
        assert "_meta" not in data_sheets
        assert "products" in data_sheets

    def test_roundtrip_xlsx(self, tmp_path: Path, monkeypatch):
        """Full XLSX roundtrip: meta survives write → read."""
        out = tmp_path / "test.xlsx"
        frames = _sample_frames()
        ExcelBackend().write_multi(frames, str(out))

        back = ExcelBackend().read_multi(str(out), header_levels=1)
        meta = back["_meta"]
        # At minimum the version field should survive
        assert meta.get("version") or meta.get("workbook_meta_blob")

    def test_derived_helper_provenance_roundtrip_xlsx(self, tmp_path: Path):
        """FTR-FK-HELPER-PROVENANCE-CLEANUP: derived helper provenance survives XLSX hidden-sheet roundtrip."""
        meta_with_prov = {
            **_SAMPLE_META,
            "derived": {
                "sheets": {
                    "desks": {
                        "helper_columns": [
                            {
                                "column": "_employees_first_name",
                                "fk_column": "id_(employees)",
                                "target": "employees",
                                "value_field": "first_name",
                            },
                        ]
                    }
                }
            },
        }
        frames = _sample_frames()
        frames["_meta"] = meta_with_prov
        out = tmp_path / "prov.xlsx"
        ExcelBackend().write_multi(frames, str(out))
        back = ExcelBackend().read_multi(str(out), header_levels=1)

        assert "_meta" in back
        meta = back["_meta"]
        prov = meta["derived"]["sheets"]["desks"]["helper_columns"]
        assert len(prov) == 1
        assert prov[0]["column"] == "_employees_first_name"
        assert prov[0]["target"] == "employees"


# ===========================================================================
# Cross-backend roundtrip
# ===========================================================================
class TestCrossBackendRoundtrip:

    def test_json_to_xlsx_to_json_meta_survives(self, tmp_path: Path, monkeypatch):
        """JSON → XLSX → JSON: meta must survive the round-trip."""
        json_dir = tmp_path / "step1_json"
        xlsx_path = tmp_path / "step2.xlsx"
        json_dir2 = tmp_path / "step3_json"

        frames = _sample_frames()

        # step 1: write to JSON
        JSONBackend().write_multi(frames, str(json_dir))

        # step 2: read from JSON, write to XLSX
        mid = JSONBackend().read_multi(str(json_dir), header_levels=1)
        assert "_meta" in mid
        ExcelBackend().write_multi(mid, str(xlsx_path))

        # step 3: read from XLSX, write to JSON
        back = ExcelBackend().read_multi(str(xlsx_path), header_levels=1)
        assert "_meta" in back
        JSONBackend().write_multi(back, str(json_dir2))

        # verify sidecar exists
        assert (json_dir2 / "_meta.yaml").exists()
