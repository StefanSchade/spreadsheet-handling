import json

import yaml

import spreadsheet_handling.application.orchestrator as orch_mod
import spreadsheet_handling.cli.apps.run as runmod


def test_main_with_real_steps_file(tmp_path, monkeypatch):
    steps_path = tmp_path / "steps.yml"
    steps_path.write_text("pipeline:
  - step: identity
")

    called = {}

    def _fake_orchestrate(**kwargs):
        called.update(kwargs)
        return {}

    monkeypatch.setattr(orch_mod, "orchestrate", _fake_orchestrate)
    monkeypatch.setattr(runmod, "orchestrate", _fake_orchestrate)

    rc = runmod.main([
        "--steps", str(steps_path),
        "--in-kind", "json_dir", "--in-path", "in",
        "--out-kind", "json_dir", "--out-path", "out",
    ])
    assert rc == 0
    assert called["input"]["kind"] == "json_dir"
    assert called["output"]["path"] == "out"


def test_main_with_top_level_pipeline(tmp_path, monkeypatch):
    cfg_path = tmp_path / "dummy.yml"
    cfg_path.write_text(
        "io:
"
        "  input: {kind: json_dir, path: in}
"
        "  output: {kind: json_dir, path: out}
"
        "pipeline:
"
        "  - step: identity
"
    )

    called = {}

    def _fake_orchestrate(**kwargs):
        called.update(kwargs)
        return {}

    monkeypatch.setattr(orch_mod, "orchestrate", _fake_orchestrate)
    monkeypatch.setattr(runmod, "orchestrate", _fake_orchestrate)

    rc = runmod.main(["--config", str(cfg_path)])
    assert rc == 0
    assert called["input"]["kind"] == "json_dir"


def test_main_run_path_honors_sheet_level_overrides(tmp_path):
    from openpyxl import load_workbook

    in_dir = tmp_path / "in"
    in_dir.mkdir()
    (in_dir / "product.json").write_text(
        json.dumps([
            {"id": "P-1", "name": "Starter", "status": "active"},
            {"id": "P-2", "name": "Pro", "status": "pilot"},
        ]),
        encoding="utf-8",
    )
    (in_dir / "branch.json").write_text(
        json.dumps([
            {"id": "B-1", "city": "Berlin"},
            {"id": "B-2", "city": "Hamburg"},
        ]),
        encoding="utf-8",
    )

    cfg_path = tmp_path / "run.yml"
    cfg = {
        "io": {
            "input": {"kind": "json_dir", "path": str(in_dir)},
            "output": {"kind": "xlsx", "path": str(tmp_path / "out.xlsx")},
        },
        "pipeline": [
            {
                "step": "apply_overrides",
                "overrides": {
                    "defaults": {"header_fill_rgb": "#DDDDDD"},
                    "sheets": {
                        "product": {
                            "freeze_header": True,
                            "auto_filter": True,
                            "header_fill_rgb": "#FFCC00",
                        },
                        "branch": {
                            "freeze_header": False,
                            "auto_filter": False,
                            "header_fill_rgb": "#00CCFF",
                        },
                    },
                },
            },
            {
                "step": "add_validations",
                "rules": [
                    {
                        "sheet": "product",
                        "column": "status",
                        "rule": {"type": "in_list", "values": ["active", "pilot"]},
                    }
                ],
            },
            {"step": "flatten_headers", "mode": "first_nonempty", "sep": "."},
        ],
    }
    cfg_path.write_text(yaml.safe_dump(cfg), encoding="utf-8")

    rc = runmod.main(["--config", str(cfg_path)])
    assert rc == 0

    wb = load_workbook(tmp_path / "out.xlsx")
    assert "_meta" in wb.sheetnames
    assert wb["_meta"].sheet_state == "hidden"

    product = wb["product"]
    branch = wb["branch"]

    assert product.freeze_panes == "A2"
    assert product.auto_filter and product.auto_filter.ref
    assert (product["A1"].fill.fgColor.rgb or "").endswith("FFCC00")

    assert branch.freeze_panes is None
    assert not (branch.auto_filter and branch.auto_filter.ref)
    assert (branch["A1"].fill.fgColor.rgb or "").endswith("00CCFF")

    wb.close()
